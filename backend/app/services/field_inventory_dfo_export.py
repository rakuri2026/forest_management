"""
Field Inventory DFO Summary Export — 4-sheet Nepali block/species summary Excel.
"""
import io
import math
import logging
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.comments import Comment as XLComment
from openpyxl.utils import get_column_letter

from ..models.field_inventory import (
    FieldInventoryCalculation,
    FieldInventoryBlockSummary,
    FieldInventoryMeasurement,
    FieldInventorySamplePlot,
)
from .field_inventory_mgmt_data import get_management_plan_data

logger = logging.getLogger(__name__)

# ── Style constants ──
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Calibri", italic=True, color="555555", size=9)
SUBHEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color="2F5496"),
    right=Side(style="medium", color="2F5496"),
    top=Side(style="medium", color="2F5496"),
    bottom=Side(style="medium", color="2F5496"),
)

AAH_MAP = {"Good": 75.0, "Moderate": 60.0, "Weak": 40.0}


def _style_header_row(ws, row: int, max_col: int, fill=None, font=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill or HEADER_FILL
        cell.font = font or HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int, bold: bool = False):
    cell = ws.cell(row=row, column=col)
    cell.font = BOLD_FONT if bold else DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="right" if col > 3 else "left", vertical="center")
    return cell


def _write_header_block(
    ws, row: int,
    ne_cols: List[str],
    en_cols: List[str],
    col_start: int = 1,
):
    for i, h in enumerate(ne_cols):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for i, h in enumerate(en_cols):
        c = col_start + i
        cell = ws.cell(row=row + 1, column=c, value=h)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _set_col_widths(ws, widths: List[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _get_aah_pct(condition: Optional[str]) -> float:
    if not condition:
        return 60.0
    return AAH_MAP.get(condition, 60.0)


def _bhari(fuelwood_m3_per_ha: float, wood_density: float) -> float:
    """Bhari/ha = Fuelwood_m³/ha × WoodDensity (t/m³) × 1150/30"""
    if fuelwood_m3_per_ha <= 0 or wood_density <= 0:
        return 0.0
    return fuelwood_m3_per_ha * wood_density * 1150.0 / 30.0


def _fetch_block_areas(db: Session, calculation_id: UUID) -> Dict[str, Dict[str, float]]:
    """Fetch block areas from calculation result_data blocks."""
    from ..models.calculation import Calculation
    calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    if not calc or not calc.result_data:
        return {}
    blocks = calc.result_data.get("blocks", [])
    result: Dict[str, Dict[str, float]] = {}
    for b in blocks:
        name = b.get("block_name", "")
        if name:
            result[name] = {
                "total_area_ha": float(b.get("area_hectares", 0)),
            }
    return result


def _fetch_species_breakdown(
    db: Session, field_inventory_id: UUID, coef_cache: Dict[str, float]
) -> List[Dict]:
    """Replicate the species-breakdown query logic, returning per-species per-block records."""
    fi = db.query(FieldInventoryCalculation).filter(
        FieldInventoryCalculation.id == field_inventory_id
    ).first()
    if not fi:
        return []
    regen_area = float(fi.regeneration_area_sqm)
    sapling_area = float(fi.sapling_area_sqm)
    pole_area = float(fi.pole_area_sqm)
    tree_area = float(fi.tree_area_sqm)

    query = text("""
        WITH plot_species AS (
            SELECT
                sp.block_name,
                m.species_scientific,
                m.species_local,
                m.stand_type,
                COUNT(DISTINCT sp.id) as total_plots_in_block,
                SUM(m.count) as total_count,
                SUM(COALESCE(m.net_volume, 0)) as total_timber,
                SUM(COALESCE(m.firewood_m3, 0)) as total_firewood,
                SUM(COALESCE(m.basal_area_m2, 0) * m.count) as total_basal_area
            FROM public.field_inventory_sample_plots sp
            JOIN public.field_inventory_measurements m ON m.sample_plot_id = sp.id
            WHERE sp.field_inventory_calculation_id = :fid
            GROUP BY sp.block_name, m.species_scientific, m.species_local, m.stand_type
        )
        SELECT
            block_name,
            species_scientific,
            species_local,
            SUM(CASE WHEN stand_type = 'Regeneration' THEN total_count ELSE 0 END) as regen_count,
            SUM(CASE WHEN stand_type = 'Sapling' THEN total_count ELSE 0 END) as sapling_count,
            SUM(CASE WHEN stand_type = 'Pole' THEN total_count ELSE 0 END) as pole_count,
            SUM(CASE WHEN stand_type = 'Tree' THEN total_count ELSE 0 END) as tree_count,
            SUM(CASE WHEN stand_type = 'Pole' THEN total_basal_area ELSE 0 END) as pole_basal_area,
            SUM(CASE WHEN stand_type = 'Tree' THEN total_basal_area ELSE 0 END) as tree_basal_area,
            SUM(CASE WHEN stand_type = 'Pole' THEN total_timber ELSE 0 END) as pole_timber,
            SUM(CASE WHEN stand_type = 'Pole' THEN total_firewood ELSE 0 END) as pole_firewood,
            SUM(CASE WHEN stand_type = 'Tree' THEN total_timber ELSE 0 END) as tree_timber,
            SUM(CASE WHEN stand_type = 'Tree' THEN total_firewood ELSE 0 END) as tree_firewood,
            MAX(total_plots_in_block) as total_plots
        FROM plot_species
        GROUP BY block_name, species_scientific, species_local
        ORDER BY block_name, species_scientific
    """)
    results = db.execute(query, {"fid": str(field_inventory_id)}).fetchall()
    rows = []
    for row in results:
        total_plots = row.total_plots or 1
        tp = float(row.pole_timber or 0) / total_plots / pole_area * 10000 if row.pole_timber else 0
        tf = float(row.pole_firewood or 0) / total_plots / pole_area * 10000 if row.pole_firewood else 0
        tt = float(row.tree_timber or 0) / total_plots / tree_area * 10000 if row.tree_timber else 0
        ttf = float(row.tree_firewood or 0) / total_plots / tree_area * 10000 if row.tree_firewood else 0

        net_timber = tp + tt
        fuelwood = tf + ttf
        wood_density = coef_cache.get(row.species_scientific, 0.65)

        pb = float(row.pole_basal_area or 0) / total_plots / pole_area * 10000 if row.pole_basal_area else 0
        tb = float(row.tree_basal_area or 0) / total_plots / tree_area * 10000 if row.tree_basal_area else 0

        rows.append({
            "block_name": row.block_name,
            "species_scientific": row.species_scientific,
            "species_local": row.species_local or "",
            "count_per_ha": int(
                (float(row.regen_count or 0) / total_plots / regen_area * 10000 if row.regen_count else 0) +
                (float(row.sapling_count or 0) / total_plots / sapling_area * 10000 if row.sapling_count else 0) +
                (float(row.pole_count or 0) / total_plots / pole_area * 10000 if row.pole_count else 0) +
                (float(row.tree_count or 0) / total_plots / tree_area * 10000 if row.tree_count else 0)
            ),
            "basal_area_m2_per_ha": round(pb + tb, 2),
            "growing_stock_m3_per_ha": round(net_timber, 2),
            "net_timber_m3_per_ha": round(net_timber, 2),
            "fuelwood_m3_per_ha": round(fuelwood, 2),
            "total_volume_m3_per_ha": round(net_timber + fuelwood, 2),
            "wood_density": wood_density,
        })
    return rows


def _load_species_coefficients(db: Session) -> Dict[str, float]:
    query = text("""
        SELECT scientific_name, wood_density_gm_cm3
        FROM public.tree_species_coefficients
        WHERE is_active = TRUE
    """)
    results = db.execute(query).fetchall()
    return {r.scientific_name: float(r.wood_density_gm_cm3 or 0.65) for r in results}


# =============================================================================
# Sheets 7-15: Management Plan Summary Writers
# =============================================================================


def _write_species_composition_sheet(ws, data):
    """Sheet 7: प्रजाती संरचना — forest-wide species composition."""
    ne_h = ["क्र.सं", "प्रजाती (वैज्ञानिक)", "स्थानीय नाम",
            "जम्मा आयतन\n(घ.मी./हे.)", "आयतन %",
            "रुख संख्या\n(प्रति हे.)", "संख्या %",
            "वृद्धि दर", "ब्लक संख्या"]
    en_h = ["SN", "Species (Scientific)", "Local Name",
            "Total Vol\n(m³/ha)", "Vol %",
            "Count\n(N/ha)", "Count %",
            "Growth Rate", "Blocks"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for i, s in enumerate(data.get("forest_wide", []), 1):
        vals = [
            i,
            s.get("scientific_name", ""),
            s.get("local_name", ""),
            s.get("total_volume_m3_per_ha", 0),
            s.get("volume_pct", 0),
            s.get("total_count_per_ha", 0),
            s.get("count_pct", 0),
            s.get("growth_rate", ""),
            s.get("block_count", 0),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, c)
            cell.value = v
        row += 1
    # Total row
    total_vol = data.get("total_volume_m3_per_ha", 0)
    total_cnt = data.get("total_count_per_ha", 0)
    total_vals = ["", "Total", "", total_vol, 100.0, total_cnt, 100.0, "", ""]
    for c, v in enumerate(total_vals, 1):
        cell = _style_data_cell(ws, row, c, bold=True)
        cell.value = v
    _set_col_widths(ws, [8, 30, 22, 14, 10, 14, 10, 12, 12])
    ws.freeze_panes = "A3"


def _write_block_comparison_sheet(ws, data):
    """Sheet 8: ब्लक तुलनात्मक — ranked block comparison."""
    ne_h = ["क्रम", "ब्लक", "क्षेत्र\n(हे.)",
            "आयतन\n(घ.मी./हे.)", "बेसल एरिया\n(व.मी./हे.)",
            "कार्बन\n(टन सी/हे.)",
            "AAH काठ\n(घ.मी./व.)", "AAH दाउरा\n(घ.मी./व.)",
            "MAI (%)", "AAH (%)", "वन स्थिति"]
    en_h = ["Rank", "Block", "Area\n(ha)",
            "Growing Stock\n(m³/ha)", "BA\n(m²/ha)",
            "Carbon\n(t C/ha)",
            "AAH Timber\n(m³/yr)", "AAH Fuelwood\n(m³/yr)",
            "MAI (%)", "AAH (%)", "Forest\nCondition"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for b in data.get("ranked", []):
        vals = [
            b.get("rank", ""),
            b.get("name", ""),
            b.get("area_ha", 0),
            b.get("growing_stock_m3ha", 0),
            b.get("basal_area_m2ha", 0),
            b.get("carbon_tcha", 0),
            b.get("aah_timber_m3yr", 0),
            b.get("aah_fuelwood_m3yr", 0),
            b.get("mai_pct", 0),
            b.get("aah_pct", 0),
            b.get("condition", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, c)
            cell.value = v
        row += 1
    _set_col_widths(ws, [8, 18, 12, 14, 12, 14, 14, 14, 10, 10, 14])
    ws.freeze_panes = "A3"


def _write_annual_harvest_sheet(ws, data):
    """Sheet 9: वार्षिक फसल योजना — harvest plan per block."""
    ne_h = ["ब्लक", "क्षेत्र\n(हे.)", "जम्मा आयतन\n(घ.मी./हे.)",
            "MAI (%)", "AAH (%)",
            "AAH काठ\n(घ.मी./व.)", "AAH दाउरा\n(घ.मी./व.)",
            "AAH काठ\n(क्यू.फी./व.)",
            "सुझाव फसल\nक्षेत्र (हे.)", "फसल चक्र\n(वर्ष)",
            "वन स्थिति"]
    en_h = ["Block", "Area\n(ha)", "Growing Stock\n(m³/ha)",
            "MAI (%)", "AAH (%)",
            "AAH Timber\n(m³/yr)", "AAH Fuelwood\n(m³/yr)",
            "AAH Timber\n(cft/yr)",
            "Coupe Area\n(ha)", "Rotation\n(yrs)",
            "Forest\nCondition"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for b in data.get("blocks", []):
        vals = [
            b.get("name", ""),
            b.get("area_ha", 0),
            b.get("growing_stock_m3ha", 0),
            b.get("mai_pct", 0),
            b.get("aah_pct", 0),
            b.get("aah_timber_m3yr", 0),
            b.get("aah_fuelwood_m3yr", 0),
            b.get("aah_timber_cftyr", 0),
            b.get("coupe_area_ha", 0),
            b.get("rotation_yrs", 0),
            b.get("condition", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, c)
            cell.value = v
        row += 1
    ft = data.get("forest_total", {})
    total_vals = [
        "Forest Total", ft.get("area_ha", 0), ft.get("growing_stock_m3ha", 0),
        "", "", ft.get("aah_timber_m3yr", 0), ft.get("aah_fuelwood_m3yr", 0),
        "", "", "", "",
    ]
    for c, v in enumerate(total_vals, 1):
        cell = _style_data_cell(ws, row, c, bold=True)
        cell.value = v
    _set_col_widths(ws, [18, 12, 14, 10, 10, 14, 14, 14, 14, 12, 14])
    ws.freeze_panes = "A3"


def _write_forest_condition_sheet(ws, data):
    """Sheet 10: वन स्थिति सारांश — condition summary + regeneration."""
    # Part A: Condition summary
    ne_h = ["वन स्थिति", "ब्लक संख्या", "क्षेत्र (हे.)",
            "क्षेत्र %", "जम्मा आयतन (m³)", "औसत आयतन (m³/ha)"]
    en_h = ["Condition", "Block Count", "Area (ha)",
            "Area %", "Total Volume (m³)", "Avg Volume (m³/ha)"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for c in data.get("by_condition", []):
        vals = [
            c.get("condition", ""),
            c.get("block_count", 0),
            c.get("area_ha", 0),
            c.get("area_pct", 0),
            c.get("total_volume_m3", 0),
            c.get("avg_volume_m3ha", 0),
        ]
        for ci, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, ci)
            cell.value = v
        row += 1
    total = data.get("total_area_ha", 0)
    total_vals = ["Total", "", total, 100.0, "", ""]
    for ci, v in enumerate(total_vals, 1):
        cell = _style_data_cell(ws, row, ci, bold=True)
        cell.value = v
    row += 2

    # Part B: Regeneration details
    regen_start = row
    ne_r = ["ब्लक", "पुनरुत्पादन अवस्था",
            "बिरुवा (N/ha)", "लाथ्रा (N/ha)", "जम्मा (N/ha)",
            "सिफारिस"]
    en_r = ["Block", "Regen Condition",
            "Seedling (N/ha)", "Sapling (N/ha)", "Total (N/ha)",
            "Recommendation"]
    _write_header_block(ws, row, ne_r, en_r)
    _style_header_row(ws, row, len(ne_r))
    _style_header_row(ws, row + 1, len(ne_r), fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row += 2
    for r in data.get("regeneration", []):
        vals = [
            r.get("block", ""),
            r.get("condition", ""),
            r.get("seedling_nha", 0),
            r.get("sapling_nha", 0),
            r.get("total_nha", 0),
            r.get("recommendation", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, ci)
            cell.value = v
        row += 1
    _set_col_widths(ws, [18, 20, 16, 16, 16, 50])
    ws.freeze_panes = "A3"


def _write_dbh_class_volume_sheet(ws, data):
    """Sheet 11: DBH वर्ग आयतन — DBH class volume distribution."""
    ne_h = ["ब्लक", "DBH वर्ग\n(से.मी.)", "नेपाली वर्ग",
            "संख्या\n(N/ha)", "काठ\n(मी.³/हे.)",
            "दाउरा\n(मी.³/हे.)", "जम्मा\n(मी.³/हे.)"]
    en_h = ["Block", "DBH Class\n(cm)", "Nepali Class",
            "Count\n(N/ha)", "Timber\n(m³/ha)",
            "Fuelwood\n(m³/ha)", "Total\n(m³/ha)"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for b in data.get("blocks", []):
        blk = b.get("block", "")
        for cls in b.get("classes", []):
            vals = [
                blk,
                cls.get("dbh_class", ""),
                cls.get("nepali_name", ""),
                cls.get("count_nha", 0),
                cls.get("timber_m3ha", 0),
                cls.get("fuelwood_m3ha", 0),
                cls.get("total_m3ha", 0),
            ]
            for c, v in enumerate(vals, 1):
                cell = _style_data_cell(ws, row, c)
                cell.value = v
            row += 1
    _set_col_widths(ws, [18, 14, 16, 12, 12, 12, 12])
    ws.freeze_panes = "A3"


def _write_carbon_per_block_sheet(ws, data):
    """Sheet 12: कार्बन भण्डार — carbon stock per block."""
    ne_h = ["ब्लक", "क्षेत्र\n(हे.)",
            "AGB\n(टन/हे.)", "BGB\n(टन/हे.)",
            "जम्मा बायोमास\n(टन/हे.)",
            "कार्बन\n(टन सी/हे.)",
            "CO₂e\n(टन/हे.)", "कुल CO₂e\n(टन)"]
    en_h = ["Block", "Area\n(ha)",
            "AGB\n(t/ha)", "BGB\n(t/ha)",
            "Total Biomass\n(t/ha)",
            "C Stock\n(t C/ha)",
            "CO₂e\n(t/ha)", "Total CO₂e\n(t)"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for b in data.get("blocks", []):
        vals = [
            b.get("block", ""),
            b.get("area_ha", 0),
            b.get("agb_tha", 0),
            b.get("bgb_tha", 0),
            b.get("total_biomass_tha", 0),
            b.get("c_stock_tcha", 0),
            b.get("co2e_tha", 0),
            b.get("total_co2e_t", 0),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, c)
            cell.value = v
        row += 1
    ft = data.get("forest_total", {})
    total_vals = [
        "Forest Total", "",
        ft.get("agb_t", 0), ft.get("bgb_t", 0),
        ft.get("agb_t", 0) + ft.get("bgb_t", 0),
        ft.get("c_stock_t", 0), "",
        ft.get("co2e_t", 0),
    ]
    for c, v in enumerate(total_vals, 1):
        cell = _style_data_cell(ws, row, c, bold=True)
        cell.value = v
    _set_col_widths(ws, [18, 12, 12, 12, 14, 14, 12, 14])
    ws.freeze_panes = "A3"


def _write_growth_rate_sheet(ws, data):
    """Sheet 13: वृद्धि दर वर्गीकरण — growth rate classification."""
    ne_h = ["वृद्धि दर", "प्रजाती संख्या",
            "आयतन (मी.³/हे.)", "आयतन %", "प्रमुख प्रजाती"]
    en_h = ["Growth Rate", "Species Count",
            "Volume (m³/ha)", "Volume %", "Key Species"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for c in data.get("classes", []):
        vals = [
            c.get("rate", ""),
            c.get("species_count", 0),
            c.get("volume_m3_per_ha", 0),
            c.get("volume_pct", 0),
            c.get("species", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, ci)
            cell.value = v
        row += 1
    total_vol = data.get("total_volume_m3_per_ha", 0)
    total_vals = ["Total", "", total_vol, 100.0, ""]
    for ci, v in enumerate(total_vals, 1):
        cell = _style_data_cell(ws, row, ci, bold=True)
        cell.value = v
    _set_col_widths(ws, [16, 14, 16, 10, 50])
    ws.freeze_panes = "A3"


def _write_stand_structure_sheet(ws, data):
    """Sheet 14: रुख संरचना प्रोफाइल — stand structure profile."""
    ne_h = ["ब्लक", "DBH वर्ग\n(से.मी.)", "नेपाली वर्ग",
            "वर्तमान\n(N/ha)", "आदर्श\n(N/ha)",
            "फरक", "अवस्था"]
    en_h = ["Block", "DBH Class\n(cm)", "Nepali Class",
            "Actual\n(N/ha)", "Ideal\n(N/ha)",
            "Difference", "Status"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for b in data.get("blocks", []):
        blk = b.get("block", "")
        for cls in b.get("classes", []):
            vals = [
                blk,
                cls.get("dbh_class", ""),
                cls.get("nepali_name", ""),
                cls.get("actual_nha", 0),
                cls.get("ideal_nha", 0),
                cls.get("difference", 0),
                cls.get("status", ""),
            ]
            for c, v in enumerate(vals, 1):
                cell = _style_data_cell(ws, row, c)
                cell.value = v
            row += 1
    row += 1
    assessment = data.get("assessment", "")
    ws.cell(row=row, column=1, value="Overall Assessment:").font = BOLD_FONT
    ws.cell(row=row, column=2, value=assessment).font = DATA_FONT
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    _set_col_widths(ws, [18, 14, 16, 12, 12, 12, 24])
    ws.freeze_panes = "A3"


def _write_productivity_sheet(ws, data):
    """Sheet 15: उत्पादकता वर्गीकरण — productivity classification."""
    ne_h = ["उत्पादकता", "सीमा (m³/ha)", "ब्लकहरू",
            "ब्लक संख्या", "कुल क्षेत्र\n(हे.)",
            "कुल आयतन\n(m³)", "सिफारिस"]
    en_h = ["Productivity", "Threshold (m³/ha)", "Blocks",
            "Block Count", "Total Area\n(ha)",
            "Total Volume\n(m³)", "Recommendation"]
    N = len(ne_h)
    _write_header_block(ws, 1, ne_h, en_h)
    _style_header_row(ws, 1, N)
    _style_header_row(ws, 2, N, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)
    row = 3
    for c in data.get("classes", []):
        vals = [
            c.get("class", ""),
            c.get("threshold", ""),
            ", ".join(c.get("blocks", [])),
            c.get("block_count", 0),
            c.get("area_ha", 0),
            c.get("volume_m3", 0),
            c.get("recommendation", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = _style_data_cell(ws, row, ci)
            cell.value = v
        row += 1
    _set_col_widths(ws, [16, 18, 30, 12, 14, 14, 60])
    ws.freeze_panes = "A3"


def generate_field_inventory_dfo_summary(
    db: Session,
    field_inventory_id: UUID,
    calculation_id: UUID,
    aah_good: float = 75.0,
    aah_moderate: float = 60.0,
    aah_weak: float = 40.0,
) -> bytes:
    """
    Generate 4-sheet DFO summary Excel:
      Sheet 1: प्रजातीगत विवरण (Species × Block)
      Sheet 2: ब्लक सारांश (Block Aggregate)
      Sheet 3: कार्बन सारांश (Carbon Summary)
      Sheet 4: DBH वर्ग विवरण (DBH Class Breakdown)
    """
    fi = db.query(FieldInventoryCalculation).filter(
        FieldInventoryCalculation.id == field_inventory_id
    ).first()
    if not fi:
        raise ValueError("Field inventory not found")

    check_meas = db.query(FieldInventoryMeasurement).join(
        FieldInventorySamplePlot,
        FieldInventoryMeasurement.sample_plot_id == FieldInventorySamplePlot.id
    ).filter(
        FieldInventorySamplePlot.field_inventory_calculation_id == field_inventory_id
    ).first()
    if not check_meas:
        raise ValueError("No measurements found in this field inventory.")

    block_summaries: List[FieldInventoryBlockSummary] = db.query(
        FieldInventoryBlockSummary
    ).filter(
        FieldInventoryBlockSummary.field_inventory_calculation_id == field_inventory_id
    ).order_by(FieldInventoryBlockSummary.block_name).all()

    if not block_summaries:
        raise ValueError("No block summaries found. Process the field inventory first.")

    coef_cache = _load_species_coefficients(db)
    species_rows = _fetch_species_breakdown(db, field_inventory_id, coef_cache)
    block_areas = _fetch_block_areas(db, calculation_id)

    bs_map: Dict[str, FieldInventoryBlockSummary] = {}
    for bs in block_summaries:
        bs_map[bs.block_name.strip()] = bs

    wb = Workbook()

    # =========================================================================
    # Sheet 1 — प्रजातीगत विवरण (Species × Block)
    # =========================================================================
    ws1 = wb.active
    ws1.title = "प्रजातीगत विवरण"

    ne_h1 = [
        "ब्लक", "प्रजाती (वैज्ञानिक)", "स्थानीय नाम",
        "रुख संख्या\n(प्रति हे.)", "बेसल एरिया\n(व.मी./हे.)",
        "काण्डको\nआयतन\n(घ.मी./हे.)",
        "काठ खप\n(घ.मी./हे.)", "दाउरा\n(घ.मी./हे.)",
        "जम्मा\nआयतन\n(घ.मी./हे.)",
        "काठ खप\n(क्यू.फी./हे.)",
        "दाउरा\n(भारी/हे.)",
        "मासिक दर\n(%)",
        "MAI काठ\n(घ.मी./हे./व.)",
        "MAI दाउरा\n(घ.मी./हे./व.)",
        "MAI जम्मा\n(घ.मी./हे./व.)",
        "संग्रह दर\n(%)",
        "AAH काठ\n(घ.मी./हे./व.)",
        "AAH दाउरा\n(घ.मी./हे./व.)",
        "AAH जम्मा\n(घ.मी./हे./व.)",
        "AAH काठ\n(क्यू.फी./हे./व.)",
    ]
    en_h1 = [
        "Block", "Species (Scientific)", "Local Name",
        "N/ha", "BA (m²/ha)",
        "Stem Vol\n(m³/ha)",
        "Net Timber\n(m³/ha)", "Fuelwood\n(m³/ha)",
        "Total Vol\n(m³/ha)",
        "Timber\n(cft/ha)",
        "Fuelwood\n(bhari/ha)",
        "MAI (%)",
        "MAI Timber\n(m³/ha/yr)",
        "MAI Fuelwood\n(m³/ha/yr)",
        "MAI Total\n(m³/ha/yr)",
        "AAH (%)",
        "AAH Timber\n(m³/ha/yr)",
        "AAH Fuelwood\n(m³/ha/yr)",
        "AAH Total\n(m³/ha/yr)",
        "AAH Timber\n(cft/ha/yr)",
    ]

    N1 = len(ne_h1)
    _write_header_block(ws1, 1, ne_h1, en_h1)
    _style_header_row(ws1, 1, N1)
    _style_header_row(ws1, 2, N1, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    # ── Excel hover comments for Sheet 1 ──
    s1c = {
        4: "Total count per hectare across all stand types (Regeneration + Sapling + Pole + Tree). Per-plot averaging method: sum counts across all sample plots, divide by total plots, divide by plot area in sqm, multiply by 10000 to get per-hectare value. This gives the density of individuals per hectare.",
        5: "Basal area per hectare calculated from Pole and Tree DBH measurements only. Per-tree basal area is π × (DBH/200)² in square meters, multiplied by count, then expanded to per-hectare using per-plot averaging. BA is a key metric for forest stocking density and competition.",
        6: "Stem (net timber) volume per hectare. Sum of Pole and Tree net timber volumes per hectare, where net timber = Gross Timber × Recovery Factor / 100. For Full Stem Merchantable species (Khair), net timber = Stem Volume with no deductions.",
        7: "Merchantable timber volume per hectare after waste deduction. Gross Timber = Stem Volume minus 10cm top-end volume (unless Full Stem Merchantable flag is set). Recovery Factor depends on tree quality class: Class 1=80%, Class 2=60%, Class 3=30%, Class 4=0%. This is the DFO timber reporting metric.",
        8: "Fuelwood volume per hectare comprising all non-merchantable wood. Calculated as Tree Volume minus Net Timber Volume. Includes 100% of branchwood volume plus any stem portion not recovered as timber. Used for firewood availability assessment.",
        9: "Total wood volume per hectare: merchantable timber plus fuelwood. Represents the complete biological productivity of the forest stand per hectare. Also known as Tree Volume per hectare (Stem Volume + Branch Volume).",
        10: "Timber volume converted to cubic feet per hectare for traditional reporting. Conversion factor: 1 cubic meter = 35.3147 cubic feet. Used for consistency with Nepal Forest Regulation volumetric reporting standards.",
        11: "Fuelwood expressed in Nepali bhari (headload) units per hectare. Formula: Fuelwood_m³ × WoodDensity × 1150 / 30, where 1150 is the air-dry correction factor (1000 kg/t × 1.15 for ~15% moisture), and 30 is the Regulation 2079 standard of 30 kg per bhari.",
        12: "Mean Annual Increment percent determined by dominant species growth rate and forest condition. The full matrix: Fast+Good=5.0%, Fast+Moderate=4.0%, Fast+Weak=3.0%, Moderate+Good=4.0%, Moderate+Moderate=3.0%, Moderate+Weak=2.0%, Slow+Good=3.0%, Slow+Moderate=2.0%, Slow+Weak=1.0%.",
        13: "Mean Annual Increment of timber: the annual timber volume increment per hectare. Formula: Net Timber (m³/ha) × MAI(%) / 100. Represents sustainable annual timber production capacity.",
        14: "Mean Annual Increment of fuelwood: the annual fuelwood volume increment per hectare. Formula: Fuelwood (m³/ha) × MAI(%) / 100. Used for assessing sustainable firewood supply.",
        15: "Mean Annual Increment total: combined annual increment of timber and fuelwood per hectare. Formula: MAI Timber + MAI Fuelwood. Total annual wood production capacity per hectare.",
        16: "Annual Allowable Harvest percent based on forest condition. Good forests = 75%, Moderate = 60%, Weak = 40%. These percentages are user-configurable via query parameters. AAH represents the sustainable harvest rate applied to the MAI.",
        17: "Annual Allowable Harvest of timber: sustainable annual timber harvest per hectare. Formula: MAI Timber × AAH(%) / 100. This is the recommended maximum annual timber extraction.",
        18: "Annual Allowable Harvest of fuelwood: sustainable annual fuelwood harvest per hectare. Formula: MAI Fuelwood × AAH(%) / 100. This is the recommended maximum annual firewood extraction.",
        19: "Annual Allowable Harvest total: combined sustainable annual harvest of timber and fuelwood. Formula: AAH Timber + AAH Fuelwood. The maximum sustainable wood extraction per year.",
        20: "Annual Allowable Harvest of timber in cubic feet per hectare per year. Formula: AAH Timber (m³/ha/yr) × 35.3147. Provides AAH in traditional cft units for operational planning.",
    }
    for col_idx, text in s1c.items():
        ws1.cell(row=1, column=col_idx).comment = XLComment(text, "System", width=400, height=150)

    row = 3
    for sr in species_rows:
        blk = sr["block_name"]
        bs = bs_map.get(blk.strip())
        mai_pct = float(bs.mai_percent) if bs and bs.mai_percent else 0.0
        cond = bs.forest_condition if bs else "Moderate"
        # Use local AAH map
        aah_pct_map = {"Good": aah_good, "Moderate": aah_moderate, "Weak": aah_weak}
        aah_pct = aah_pct_map.get(cond, aah_moderate)

        nt = sr["net_timber_m3_per_ha"]
        fw = sr["fuelwood_m3_per_ha"]
        wd = sr["wood_density"]

        mai_timber = nt * mai_pct / 100.0
        mai_fuelwood = fw * mai_pct / 100.0
        mai_total = mai_timber + mai_fuelwood

        aah_timber = mai_timber * aah_pct / 100.0
        aah_fuelwood = mai_fuelwood * aah_pct / 100.0
        aah_total = aah_timber + aah_fuelwood

        vals = [
            blk,
            sr["species_scientific"],
            sr["species_local"],
            sr["count_per_ha"],
            sr["basal_area_m2_per_ha"],
            sr["growing_stock_m3_per_ha"],
            nt, fw, sr["total_volume_m3_per_ha"],
            round(nt * 35.3147, 2),
            round(_bhari(fw, wd), 2),
            round(mai_pct, 1),
            round(mai_timber, 2), round(mai_fuelwood, 2), round(mai_total, 2),
            round(aah_pct, 1),
            round(aah_timber, 2), round(aah_fuelwood, 2), round(aah_total, 2),
            round(aah_timber * 35.3147, 2),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws1, row, c)
            cell.value = v
        row += 1

    _set_col_widths(ws1, [18, 30, 22, 12, 12, 12, 12, 12, 12, 12, 12, 10, 14, 14, 14, 10, 14, 14, 14, 14])
    ws1.freeze_panes = "A3"

    # =========================================================================
    # Sheet 2 — ब्लक सारांश (Block-wise Aggregate)
    # =========================================================================
    ws2 = wb.create_sheet("ब्लक सारांश")

    ne_h2 = [
        "ब्लक", "जम्मा क्षेत्र\n(हे.)", "प्रभावकारी\nक्षेत्र (हे.)",
        "वन स्थिति",
        "रुख संख्या\n(प्रति हे.)", "बेसल एरिया\n(व.मी./हे.)",
        "काण्डको\nआयतन\n(घ.मी./हे.)",
        "काठ खप\n(घ.मी./हे.)", "दाउरा\n(घ.मी./हे.)",
        "जम्मा\nआयतन\n(घ.मी./हे.)",
        "काठ खप\n(क्यू.फी./हे.)",
        "मासिक दर\n(%)",
        "MAI काठ\n(घ.मी./हे./व.)",
        "MAI दाउरा\n(घ.मी./हे./व.)",
        "MAI जम्मा\n(घ.मी./हे./व.)",
        "संग्रह दर\n(%)",
        "AAH काठ\n(घ.मी./हे./व.)",
        "AAH दाउरा\n(घ.मी./हे./व.)",
        "AAH जम्मा\n(घ.मी./हे./व.)",
    ]
    en_h2 = [
        "Block", "Total Area\n(ha)", "Effective\nArea (ha)",
        "Forest\nCondition",
        "N/ha", "BA (m²/ha)",
        "Stem Vol\n(m³/ha)",
        "Net Timber\n(m³/ha)", "Fuelwood\n(m³/ha)",
        "Total Vol\n(m³/ha)",
        "Timber\n(cft/ha)",
        "MAI (%)",
        "MAI Timber\n(m³/ha/yr)",
        "MAI Fuelwood\n(m³/ha/yr)",
        "MAI Total\n(m³/ha/yr)",
        "AAH (%)",
        "AAH Timber\n(m³/ha/yr)",
        "AAH Fuelwood\n(m³/ha/yr)",
        "AAH Total\n(m³/ha/yr)",
    ]

    N2 = len(ne_h2)
    _write_header_block(ws2, 1, ne_h2, en_h2)
    _style_header_row(ws2, 1, N2)
    _style_header_row(ws2, 2, N2, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    # ── Excel hover comments for Sheet 2 ──
    s2c = {
        4: "Overall forest condition based on growing stock volume and regeneration status. Growing stock >200 m³/ha with Good/Moderate regeneration = Good. Growing stock 50-200 with Good regeneration = Good, Moderate = Moderate, Weak = Weak. Growing stock <50 with Good regeneration = Moderate, otherwise Weak.",
        5: "Total count per hectare aggregated across all species in the block. Includes Regeneration, Sapling, Pole, and Tree stand types. Per-plot averaging applied across all sample plots in the block before expansion to per-hectare.",
        6: "Basal area per hectare for the entire block, aggregated across all species. Calculated from Pole and Tree DBH measurements only. BA is a key structural metric indicating forest stocking density and is used in stand density management.",
        7: "Total growing stock (net timber volume) per hectare for the block. Sum of Pole and Tree net timber volumes across all species. Net Timber = Gross Timber × Recovery Factor / 100. This is the standing merchantable volume.",
        8: "Total net timber volume per hectare aggregated across all species. Gross Timber = Stem Volume minus 10cm top-end (unless Full Stem Merchantable). Recovery Factor: Class 1=80%, 2=60%, 3=30%, 4=0%. Growing Stock for timber valuation.",
        9: "Total fuelwood volume per hectare aggregated across all species. Includes branchwood (100% fuelwood by regulation) plus non-recovered stem portion. Calculated as Tree Volume minus Net Timber Volume.",
        12: "Block-level Mean Annual Increment percent. Determined by dominant species growth rate and forest condition assessment for this block. Applied uniformly across all species to calculate MAI volumes.",
        13: "Block-level MAI timber volume per year. The annual increment of merchantable timber for the entire block per hectare. Formula: Net Timber (m³/ha) × MAI(%) / 100.",
        16: "Block-level Annual Allowable Harvest percent. Good = 75%, Moderate = 60%, Weak = 40%. Based on forest condition. These are user-configurable via query parameters aah_good, aah_moderate, aah_weak.",
        17: "Block-level AAH of timber per year. The recommended sustainable annual timber harvest per hectare. Formula: MAI Timber × AAH(%) / 100.",
        18: "Block-level AAH of fuelwood per year. The recommended sustainable annual firewood harvest per hectare. Formula: MAI Fuelwood × AAH(%) / 100.",
    }
    for col_idx, text in s2c.items():
        ws2.cell(row=1, column=col_idx).comment = XLComment(text, "System", width=400, height=150)

    row = 3
    for bs in block_summaries:
        blk = bs.block_name.strip()
        ba = block_areas.get(blk, {})
        total_ha = ba.get("total_area_ha", 0)
        # effective_area — placeholder until tree cover analysis is run
        eff_ha = ba.get("total_area_ha", 0)

        nt_m3 = float(bs.tree_timber_m3_per_ha or 0) + float(bs.pole_timber_m3_per_ha or 0)
        fw_m3 = float(bs.tree_firewood_m3_per_ha or 0) + float(bs.pole_firewood_m3_per_ha or 0)
        gs = float(bs.total_growing_stock_m3_per_ha or 0)
        ba_val = float(bs.basal_area_m2_per_ha or 0)
        n_ha = (bs.regeneration_per_ha or 0) + (bs.sapling_per_ha or 0) + (bs.pole_per_ha or 0) + (bs.tree_per_ha or 0)
        mai_pct = float(bs.mai_percent) if bs.mai_percent else 0.0
        cond = bs.forest_condition or "Moderate"
        aah_pct_map = {"Good": aah_good, "Moderate": aah_moderate, "Weak": aah_weak}
        aah_pct = aah_pct_map.get(cond, aah_moderate)

        mai_t = nt_m3 * mai_pct / 100.0
        mai_f = fw_m3 * mai_pct / 100.0
        mai_tot = mai_t + mai_f
        aah_t = mai_t * aah_pct / 100.0
        aah_f = mai_f * aah_pct / 100.0
        aah_tot = aah_t + aah_f

        vals = [
            blk,
            round(total_ha, 2), round(eff_ha, 2),
            cond,
            n_ha, round(ba_val, 2), round(gs, 2),
            round(nt_m3, 2), round(fw_m3, 2), round(nt_m3 + fw_m3, 2),
            round(nt_m3 * 35.3147, 2),
            round(mai_pct, 1),
            round(mai_t, 2), round(mai_f, 2), round(mai_tot, 2),
            round(aah_pct, 1),
            round(aah_t, 2), round(aah_f, 2), round(aah_tot, 2),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws2, row, c)
            cell.value = v
        row += 1

    _set_col_widths(ws2, [18, 14, 14, 14, 12, 12, 12, 12, 12, 12, 12, 10, 14, 14, 14, 10, 14, 14, 14])
    ws2.freeze_panes = "A3"

    # =========================================================================
    # Sheet 3 — कार्बन सारांश (Carbon Summary)
    # =========================================================================
    ws3 = wb.create_sheet("कार्बन सारांश")

    ne_h3 = [
        "ब्लक", "प्रभावकारी\nक्षेत्र (हे.)",
        "काठको मात्रा\n(घ.मी./हे.)",
        "ए.जी.बी.\n(टन/हे.)",
        "बी.जी.बी.\n(टन/हे.)",
        "जम्मा\nबायोमास\n(टन/हे.)",
        "कार्बन\nस्टक\n(टन सी/हे.)",
        "CO₂e\n(टन/हे.)",
        "भारित काठ\nघनत्व\n(टन/मी.³)",
    ]
    en_h3 = [
        "Block", "Effective\nArea (ha)",
        "Growing Stock\n(m³/ha)",
        "AGB\n(t/ha)",
        "BGB\n(t/ha)",
        "Total\nBiomass\n(t/ha)",
        "C Stock\n(t C/ha)",
        "CO₂e\n(t/ha)",
        "Weighted\nWood Density\n(t/m³)",
    ]

    N3 = len(ne_h3)
    _write_header_block(ws3, 1, ne_h3, en_h3)
    _style_header_row(ws3, 1, N3)
    _style_header_row(ws3, 2, N3, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    # ── Excel hover comments for Sheet 3 ──
    s3c = {
        3: "Total growing stock (net timber volume) per hectare. This is the merchantable stem volume that serves as the basis for biomass expansion. Growing stock is the standing volume of timber measured in cubic meters per hectare.",
        4: "Above-Ground Biomass in tonnes per hectare. Calculated as Growing Stock × Weighted Wood Density × Biomass Expansion Factor (BEF=1.40). The BEF accounts for bark, branches, foliage, and other above-ground components not included in stem volume. IPCC Tier 2 methodology for tropical broadleaf forests.",
        5: "Below-Ground Biomass (roots) in tonnes per hectare. Calculated as AGB × Root-to-Shoot Ratio (0.24). Root-to-shoot ratio of 0.24 means root biomass is 24% of above-ground biomass, based on IPCC default values for tropical moist forest ecosystems.",
        6: "Total tree biomass per hectare, combining above and below-ground components. Formula: AGB + BGB. This represents the complete carbon pool in living tree biomass, used for carbon stock estimation under REDD+ and IPCC guidelines.",
        7: "Total carbon stock in tonnes of carbon per hectare. Formula: Total Biomass × Carbon Fraction (0.47). The carbon fraction of 0.47 means 47% of dry tree biomass is elemental carbon, following IPCC 2006 Guidelines Table 4.3 for tropical forests.",
        8: "CO₂ equivalent sequestered in tonnes per hectare. Formula: Carbon Stock × 3.67 (molecular ratio 44/12). Each tonne of carbon stored equals 3.67 tonnes of CO₂ removed from the atmosphere. Used for carbon accounting and reporting under UNFCCC.",
        9: "Volume-weighted average wood density across all tree species in the block. Formula: Σ(Species Volume × Species Density) / Σ(Species Volume). Wood density converts volume to mass, with default value of 0.65 t/m³ if species coefficients are unavailable.",
    }
    for col_idx, text in s3c.items():
        ws3.cell(row=1, column=col_idx).comment = XLComment(text, "System", width=400, height=150)

    row = 3
    for bs in block_summaries:
        blk = bs.block_name.strip()
        ba = block_areas.get(blk, {})
        eff_h = ba.get("total_area_ha", 0)

        vals = [
            blk,
            round(eff_h, 2),
            round(float(bs.total_growing_stock_m3_per_ha or 0), 2),
            round(float(bs.agb_t_per_ha or 0), 2),
            round(float(bs.bgb_t_per_ha or 0), 2),
            round(float(bs.total_biomass_t_per_ha or 0), 2),
            round(float(bs.carbon_stock_tc_per_ha or 0), 2),
            round(float(bs.co2_equivalent_tco2_per_ha or 0), 2),
            round(float(bs.weighted_wood_density or 0), 3),
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws3, row, c)
            cell.value = v
        row += 1

    _set_col_widths(ws3, [18, 14, 14, 12, 12, 12, 12, 12, 14])
    ws3.freeze_panes = "A3"

    # =========================================================================
    # Sheet 4 — DBH वर्ग विवरण (DBH Class Breakdown)
    # =========================================================================
    ws4 = wb.create_sheet("DBH वर्ग विवरण")

    ne_h4 = [
        "ब्लक", "DBH वर्ग\n(से.मी.)",
        "नेपाली वर्ग",
        "संख्या\n(प्रति हे.)",
        "काठ\n(घ.मी./हे.)",
        "दाउरा\n(घ.मी./हे.)",
    ]
    en_h4 = [
        "Block", "DBH Class\n(cm)",
        "Nepali Class",
        "Count\n(N/ha)",
        "Timber\n(m³/ha)",
        "Firewood\n(m³/ha)",
    ]

    # DBH classes in display order (excluding regeneration 0-4 and 4-10 which go in Sheet 5)
    DBH_CLASS_ORDER_POLETREE = [
        ("10_20", "Small Pole", "सानो खाँवा"),
        ("20_30", "Large Pole", "ठुलो खाँवा"),
        ("30_40", "Small Tree", "सानो रुख"),
        ("40_50", "Medium Tree", "मझौला रुख"),
        ("50_60", "Large Tree", "ठुलो रुख"),
        ("60_plus", "V. Large Tree", "अति ठुलो रुख"),
    ]

    DBH_CLASS_ALL = [
        ("0_4", "Seedling", "बिरुवा"),
        ("4_10", "Sapling", "लाथ्रा"),
        ("10_20", "Small Pole", "सानो खाँवा"),
        ("20_30", "Large Pole", "ठुलो खाँवा"),
        ("30_40", "Small Tree", "सानो रुख"),
        ("40_50", "Medium Tree", "मझौला रुख"),
        ("50_60", "Large Tree", "ठुलो रुख"),
        ("60_plus", "V. Large Tree", "अति ठुलो रुख"),
    ]

    N4 = len(ne_h4)
    _write_header_block(ws4, 1, ne_h4, en_h4)
    _style_header_row(ws4, 1, N4)
    _style_header_row(ws4, 2, N4, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    # ── Excel hover comments for Sheet 4 ──
    s4c = {
        4: "Number of individuals per hectare in this DBH class. Per-plot averaging method: sum count for this class across all sample plots, divide by total plots, divide by plot area, multiply by 10000. This ensures DBH class totals are consistent with block-level per-hectare values.",
        5: "Net timber volume per hectare allocated to this DBH class. Uses per-plot averaging where timber volumes are summed by DBH class within each plot, then averaged across all plots, then expanded to per-hectare. Growing stock distribution by diameter class.",
        6: "Fuelwood volume per hectare allocated to this DBH class. Same per-plot averaging methodology as timber. Firewood volume distribution by diameter class, useful for understanding fuelwood availability by tree size.",
    }
    for col_idx, text in s4c.items():
        ws4.cell(row=1, column=col_idx).comment = XLComment(text, "System", width=400, height=150)

    row = 3
    for bs in block_summaries:
        blk = bs.block_name.strip()
        dbh_data = bs.dbh_class_breakdown or {}
        for key, en_label, np_label in DBH_CLASS_ORDER_POLETREE:
            cls_row = dbh_data.get(key, {})
            cnt = float(cls_row.get("count_per_ha", 0))
            tbr = float(cls_row.get("timber_m3_per_ha", 0))
            fw = float(cls_row.get("firewood_m3_per_ha", 0))
            if cnt == 0 and tbr == 0 and fw == 0:
                continue
            vals = [
                blk,
                key.replace("_", "–"),
                np_label,
                round(cnt, 2),
                round(tbr, 2),
                round(fw, 2),
            ]
            for c, v in enumerate(vals, 1):
                cell = _style_data_cell(ws4, row, c)
                cell.value = v
            row += 1

    _set_col_widths(ws4, [18, 14, 16, 12, 12, 12])
    ws4.freeze_panes = "A3"

    # =========================================================================
    # Sheet 5 — पुनरुत्पादन (Regeneration — 0-4 cm & 4-10 cm classes)
    # =========================================================================
    ws5 = wb.create_sheet("पुनरुत्पादन")

    ne_h5 = [
        "ब्लक",
        "बिरुवा संख्या\n(०–४ से.मी.)\n(प्रति हे.)",
        "लाथ्रा संख्या\n(४–१० से.मी.)\n(प्रति हे.)",
        "जम्मा पुनरुत्पादन\n(प्रति हे.)",
        "पुनरुत्पादन\nअवस्था",
        "मापदण्ड\n(बिरुवा/लाथ्रा)",
    ]
    en_h5 = [
        "Block",
        "Regeneration\n(0-4 cm)\n(N/ha)",
        "Sapling\n(4-10 cm)\n(N/ha)",
        "Total\nRegeneration\n(N/ha)",
        "Regeneration\nCondition",
        "Threshold\n(Regen/Sapling)",
    ]

    N5 = len(ne_h5)
    _write_header_block(ws5, 1, ne_h5, en_h5)
    _style_header_row(ws5, 1, N5)
    _style_header_row(ws5, 2, N5, fill=SUBHEADER_FILL, font=SUBHEADER_FONT)

    s5c = {
        5: "Regeneration condition based on combined Regen (0-4 cm) and Sapling (4-10 cm) densities. Good: Regen ≥5000/ha AND Sapling ≥2000/ha. Moderate: Regen ≥2000/ha AND Sapling ≥800/ha. Weak: below Moderate threshold. Condition stored in block_summary.regeneration_condition from Forest Regulation 2079 assessment.",
        6: "Threshold criteria used to determine the regeneration condition. Shows the minimum required densities for both Regen and Sapling classes. Each condition level requires BOTH thresholds to be met simultaneously (AND logic).",
    }
    for col_idx, text in s5c.items():
        ws5.cell(row=1, column=col_idx).comment = XLComment(text, "System", width=400, height=150)

    row = 3
    for bs in block_summaries:
        blk = bs.block_name.strip()
        dbh_data = bs.dbh_class_breakdown or {}

        # Get regen (0-4) and sapling (4-10) counts
        regen_cls = dbh_data.get("0_4", {})
        sapling_cls = dbh_data.get("4_10", {})
        regen_n = float(regen_cls.get("count_per_ha", 0))
        sapling_n = float(sapling_cls.get("count_per_ha", 0))
        total = regen_n + sapling_n

        cond = bs.regeneration_condition or "Weak"

        if cond == "Good":
            threshold_label = "≥5000 / ≥2000"
        elif cond == "Moderate":
            threshold_label = "≥2000 / ≥800"
        else:
            threshold_label = "<2000 / <800"

        vals = [
            blk,
            round(regen_n, 2),
            round(sapling_n, 2),
            round(total, 2),
            cond,
            threshold_label,
        ]
        for c, v in enumerate(vals, 1):
            cell = _style_data_cell(ws5, row, c)
            cell.value = v
        row += 1

    _set_col_widths(ws5, [18, 16, 16, 16, 18, 22])
    ws5.freeze_panes = "A3"
    ws5.sheet_properties.tabColor = "2ECC71"

    # =========================================================================
    # Sheet 6 — Descriptions (Calculation Methodology)
    # =========================================================================
    ws6 = wb.create_sheet("Descriptions")

    desc_headers = ["Sheet", "Col", "Nepali Header", "English Header", "Type", "Formula / Description"]
    for c, h in enumerate(desc_headers, 1):
        cell = ws6.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DESCRIPTIONS = [
        # ── Sheet 1 ──
        ("प्रजातीगत विवरण", "A", "ब्लक", "Block", "Static", "Block name from uploaded polygon/data."),
        ("प्रजातीगत विवरण", "B", "प्रजाती (वैज्ञानिक)", "Species (Scientific)", "Static", "Scientific name from field measurement. Used as lookup key for species coefficients and wood density."),
        ("प्रजातीगत विवरण", "C", "स्थानीय नाम", "Local Name", "Static", "Local/Nepali name of the species from field measurement."),
        ("प्रजातीगत विवरण", "D", "रुख संख्या (प्रति हे.)", "Count (N/ha)", "Formula",
         "Total individuals per hectare across all stand types.\n"
         "= (Σ Regen count / Regen_plots / Regen_area × 10000)\n"
         "+ (Σ Sapling count / Sapling_plots / Sapling_area × 10000)\n"
         "+ (Σ Pole count / Pole_plots / Pole_area × 10000)\n"
         "+ (Σ Tree count / Tree_plots / Tree_area × 10000).\n"
         "Per-plot averaging: averages counts across all sample plots then expands to per-hectare."),
        ("प्रजातीगत विवरण", "E", "बेसल एरिया (व.मी./हे.)", "BA (m²/ha)", "Formula",
         "Basal area per hectare (Pole + Tree only).\n"
         "= Σ (π × (DBH/200)² × count) / plots / plot_area_sqm × 10000.\n"
         "Pi × (DBH_cm ÷ 200)² gives basal area in m² per tree."),
        ("प्रजातीगत विवरण", "F", "काण्डको आयतन (घ.मी./हे.)", "Stem Vol (m³/ha)", "Formula",
         "Growing stock / stem volume per hectare.\n"
         "= Pole_net_m³_per_ha + Tree_net_m³_per_ha.\n"
         "Also called Net Timber Volume (m³/ha). Synonym: Growing Stock."),
        ("प्रजातीगत विवरण", "G", "काठ खप (घ.मी./हे.)", "Net Timber (m³/ha)", "Formula",
         "Merchantable timber volume per hectare.\n"
         "= Pole_timber_m³_per_ha + Tree_timber_m³_per_ha.\n"
         "Calculated per-tree: Net Timber = Gross Timber × Recovery Factor / 100.\n"
         "Gross Timber = IF(FullStemMerchantable=1, StemVol, StemVol − 10cmTopVol).\n"
         "Recovery Factor: Class 1=80%, Class 2=60%, Class 3=30%, Class 4=0%.\n"
         "For Khair (FSM=1): Net Timber = StemVol (no top deduction, 100% recovery)."),
        ("प्रजातीगत विवरण", "H", "दाउरा (घ.मी./हे.)", "Fuelwood (m³/ha)", "Formula",
         "Fuelwood volume per hectare (residual after timber extraction).\n"
         "= Pole_firewood_m³_per_ha + Tree_firewood_m³_per_ha.\n"
         "Per-tree: Fuelwood = TreeVol − NetTimberVol.\n"
         "Includes: Branchwood Volume (100% fuelwood by regulation) + non-merchantable stem portion.\n"
         "Branchwood = StemVol × BranchwoodRatio (interpolated by DBH class: s/m/bg coefficients)."),
        ("प्रजातीगत विवरण", "I", "जम्मा आयतन (घ.मी./हे.)", "Total Vol (m³/ha)", "Formula",
         "Total wood volume per hectare (timber + fuelwood).\n"
         "= Net Timber (m³/ha) + Fuelwood (m³/ha).\n"
         "Equivalent to Tree Volume (StemVol + BranchVol) per hectare."),
        ("प्रजातीगत विवरण", "J", "काठ खप (क्यू.फी./हे.)", "Timber (cft/ha)", "Formula",
         "Merchantable timber volume in cubic feet per hectare.\n"
         "= Net Timber (m³/ha) × 35.3147.\n"
         "1 m³ = 35.3147 cubic feet."),
        ("प्रजातीगत विवरण", "K", "दाउरा (भारी/हे.)", "Fuelwood (bhari/ha)", "Formula",
         "Fuelwood expressed in Nepali bhari (headload) units per hectare.\n"
         "= Fuelwood (m³/ha) × WoodDensity (t/m³) × 1150/30.\n"
         "Where: WoodDensity = species basic density (g/cm³ = t/m³),\n"
         "1150 = 1000 kg/t × 1.15 (air-dry moisture correction at ~15% MC),\n"
         "30 = kg per bhari (Forest Regulation 2079 standard).\n"
         "Note: 1 bhari = 30 kg air-dry firewood."),
        ("प्रजातीगत विवरण", "L", "मासिक दर (%)", "MAI (%)", "Static",
         "Mean Annual Increment percent.\n"
         "Determined by dominant growth rate + forest condition matrix:\n"
         "Fast+Good=5.0%, Fast+Moderate=4.0%, Fast+Weak=3.0%,\n"
         "Moderate+Good=4.0%, Moderate+Moderate=3.0%, Moderate+Weak=2.0%,\n"
         "Slow+Good=3.0%, Slow+Moderate=2.0%, Slow+Weak=1.0%.\n"
         "Dominant growth rate from most abundant species in block (count-weighted)."),
        ("प्रजातीगत विवरण", "M", "MAI काठ (घ.मी./हे./व.)", "MAI Timber (m³/ha/yr)", "Formula",
         "Mean Annual Increment of timber volume.\n"
         "= Net Timber (m³/ha) × MAI(%) / 100.\n"
         "Annual timber volume production per hectare."),
        ("प्रजातीगत विवरण", "N", "MAI दाउरा (घ.मी./हे./व.)", "MAI Fuelwood (m³/ha/yr)", "Formula",
         "Mean Annual Increment of fuelwood volume.\n"
         "= Fuelwood (m³/ha) × MAI(%) / 100.\n"
         "Annual fuelwood volume production per hectare."),
        ("प्रजातीगत विवरण", "O", "MAI जम्मा (घ.मी./हे./व.)", "MAI Total (m³/ha/yr)", "Formula",
         "Mean Annual Increment of total volume.\n"
         "= MAI Timber + MAI Fuelwood.\n"
         "Total annual wood volume production per hectare."),
        ("प्रजातीगत विवरण", "P", "संग्रह दर (%)", "AAH (%)", "Static",
         "Annual Allowable Harvest percent.\n"
         "Based on forest condition: Good = 75%, Moderate = 60%, Weak = 40%.\n"
         "User-configurable via query parameters (aah_good/moderate/weak).\n"
         "Represents the sustainable harvest rate applied to MAI."),
        ("प्रजातीगत विवरण", "Q", "AAH काठ (घ.मी./हे./व.)", "AAH Timber (m³/ha/yr)", "Formula",
         "Annual Allowable Harvest of timber volume.\n"
         "= MAI Timber × AAH(%) / 100.\n"
         "Sustainable annual timber harvest per hectare."),
        ("प्रजातीगत विवरण", "R", "AAH दाउरा (घ.मी./हे./व.)", "AAH Fuelwood (m³/ha/yr)", "Formula",
         "Annual Allowable Harvest of fuelwood volume.\n"
         "= MAI Fuelwood × AAH(%) / 100.\n"
         "Sustainable annual fuelwood harvest per hectare."),
        ("प्रजातीगत विवरण", "S", "AAH जम्मा (घ.मी./हे./व.)", "AAH Total (m³/ha/yr)", "Formula",
         "Annual Allowable Harvest of total volume.\n"
         "= AAH Timber + AAH Fuelwood.\n"
         "Sustainable annual total wood harvest per hectare."),
        ("प्रजातीगत विवरण", "T", "AAH काठ (क्यू.फी./हे./व.)", "AAH Timber (cft/ha/yr)", "Formula",
         "Annual Allowable Harvest of timber in cubic feet.\n"
         "= AAH Timber (m³/ha/yr) × 35.3147.\n"
         "Sustainable annual timber harvest in cft per hectare."),

        # ── Sheet 2 ──
        ("ब्लक सारांश", "A", "ब्लक", "Block", "Static", "Block name from uploaded polygon/data."),
        ("ब्लक सारांश", "B", "जम्मा क्षेत्र (हे.)", "Total Area (ha)", "Static",
         "Total geometric area of the block in hectares.\n"
         "Calculated from the block polygon geometry (area_hectares)."),
        ("ब्लक सारांश", "C", "प्रभावकारी क्षेत्र (हे.)", "Effective Area (ha)", "Static",
         "Effective forest area after excluding non-forest land.\n"
         "Currently equals total area. Will incorporate tree-cover analysis + protected/private exclusions\n"
         "from Table 5 (block-area-detail) in future User Group integration."),
        ("ब्लक सारांश", "D", "वन स्थिति", "Forest Condition", "Static",
         "Overall forest condition assessment.\n"
         "Based on growing stock (m³/ha) + regeneration condition (Good/Moderate/Weak).\n"
         "Growing stock >200 + Regen Good/Moderate → Good.\n"
         "Growing stock 50–200 + Regen Good → Good, Moderate → Moderate, Weak → Weak.\n"
         "Growing stock <50 + Regen Good → Moderate, Moderate → Weak, Weak → Weak."),
        ("ब्लक सारांश", "E", "रुख संख्या (प्रति हे.)", "Count (N/ha)", "Formula",
         "Total individuals per hectare across all stand types.\n"
         "= Σ(Regen + Sapling + Pole + Tree per-hectare counts).\n"
         "Summed across all species in the block."),
        ("ब्लक सारांश", "F", "बेसल एरिया (व.मी./हे.)", "BA (m²/ha)", "Formula",
         "Basal area per hectare (Pole + Tree only).\n"
         "= Pole_BA_m²_per_ha + Tree_BA_m²_per_ha.\n"
         "Summed across all species."),
        ("ब्लक सारांश", "G", "काण्डको आयतन (घ.मी./हे.)", "Stem Vol (m³/ha)", "Formula",
         "Growing stock (net timber volume) per hectare.\n"
         "= total_growing_stock_m3_per_ha from block summary.\n"
         "Sum of Pole + Tree net timber volumes per hectare."),
        ("ब्लक सारांश", "H", "काठ खप (घ.मी./हे.)", "Net Timber (m³/ha)", "Formula",
         "Tree_timber_m³_per_ha + Pole_timber_m³_per_ha.\n"
         "Summed across all species in the block. This is the net (growing stock) merchantable timber volume after waste deduction."),
        ("ब्लक सारांश", "I", "दाउरा (घ.मी./हे.)", "Fuelwood (m³/ha)", "Formula",
         "Tree_firewood_m³_per_ha + Pole_firewood_m³_per_ha.\n"
         "Summed across all species in the block. Residual fuelwood volume after timber extraction, includes all branchwood."),
        ("ब्लक सारांश", "J", "जम्मा आयतन (घ.मी./हे.)", "Total Vol (m³/ha)", "Formula",
         "Net Timber (m³/ha) + Fuelwood (m³/ha).\n"
         "Total wood volume per hectare combining merchantable timber and fuelwood."),
        ("ब्लक सारांश", "K", "काठ खप (क्यू.फी./हे.)", "Timber (cft/ha)", "Formula",
         "Net Timber (m³/ha) × 35.3147. Converts cubic meters to cubic feet for traditional reporting."),
        ("ब्लक सारांश", "L", "मासिक दर (%)", "MAI (%)", "Static",
         "Block-level Mean Annual Increment percent.\n"
         "Same method as Sheet 1: dominant growth rate × forest condition matrix.\n"
         "See Sheet 1 Col L for the full matrix."),
        ("ब्लक सारांश", "M", "MAI काठ (घ.मी./हे./व.)", "MAI Timber (m³/ha/yr)", "Formula",
         "Net Timber (m³/ha) × MAI(%) / 100. Mean Annual Increment of timber volume per hectare per year."),
        ("ब्लक सारांश", "N", "MAI दाउरा (घ.मी./हे./व.)", "MAI Fuelwood (m³/ha/yr)", "Formula",
         "Fuelwood (m³/ha) × MAI(%) / 100. Mean Annual Increment of fuelwood volume per hectare per year."),
        ("ब्लक सारांश", "O", "MAI जम्मा (घ.मी./हे./व.)", "MAI Total (m³/ha/yr)", "Formula",
         "MAI Timber + MAI Fuelwood. Total Mean Annual Increment combining timber and fuelwood per hectare per year."),
        ("ब्लक सारांश", "P", "संग्रह दर (%)", "AAH (%)", "Static",
         "Block-level Annual Allowable Harvest percent.\n"
         "Good = 75%, Moderate = 60%, Weak = 40%.\n"
         "User-configurable via query parameters."),
        ("ब्लक सारांश", "Q", "AAH काठ (घ.मी./हे./व.)", "AAH Timber (m³/ha/yr)", "Formula",
         "'= MAI Timber × AAH(%) / 100."),
        ("ब्लक सारांश", "R", "AAH दाउरा (घ.मी./हे./व.)", "AAH Fuelwood (m³/ha/yr)", "Formula",
         "'= MAI Fuelwood × AAH(%) / 100."),
        ("ब्लक सारांश", "S", "AAH जम्मा (घ.मी./हे./व.)", "AAH Total (m³/ha/yr)", "Formula",
         "'= AAH Timber + AAH Fuelwood."),

        # ── Sheet 3 ──
        ("कार्बन सारांश", "A", "ब्लक", "Block", "Static", "Block name."),
        ("कार्बन सारांश", "B", "प्रभावकारी क्षेत्र (हे.)", "Effective Area (ha)", "Static",
         "Effective forest area for carbon accounting.\n"
         "Currently equals total area. Future: area after excluding non-forest."),
        ("कार्बन सारांश", "C", "काठको मात्रा (घ.मी./हे.)", "Growing Stock (m³/ha)", "Static",
         "Total net timber volume per hectare.\n"
         "= total_growing_stock_m3_per_ha from block summary.\n"
         "Input value for biomass expansion."),
        ("कार्बन सारांश", "D", "ए.जी.बी. (टन/हे.)", "AGB (t/ha)", "Static",
         "Above-Ground Biomass in tonnes per hectare.\n"
         "= Growing_Stock × Weighted_Wood_Density × BEF.\n"
         "Where BEF (Biomass Expansion Factor) = 1.40 (IPCC default for tropical broadleaf).\n"
         "AGB converts stem volume into total above-ground tree biomass including bark, branches, foliage."),
        ("कार्बन सारांश", "E", "बी.जी.बी. (टन/हे.)", "BGB (t/ha)", "Static",
         "Below-Ground Biomass (roots) in tonnes per hectare.\n"
         "= AGB × Root-to-Shoot Ratio (0.24 for tropical moist forest, IPCC Table 4.4).\n"
         "Root biomass estimated as 24% of above-ground biomass."),
        ("कार्बन सारांश", "F", "जम्मा बायोमास (टन/हे.)", "Total Biomass (t/ha)", "Static",
         "Total tree biomass per hectare.\n"
         "= AGB + BGB.\n"
         "Total carbon pool in living tree biomass."),
        ("कार्बन सारांश", "G", "कार्बन स्टक (टन सी/हे.)", "C Stock (t C/ha)", "Static",
         "Total carbon stock in tree biomass per hectare.\n"
         "= Total_Biomass × Carbon_Fraction (0.47).\n"
         "IPCC default carbon fraction of 47% for tropical forest biomass (IPCC 2006 GL, Table 4.3)."),
        ("कार्बन सारांश", "H", "CO₂e (टन/हे.)", "CO₂e (t/ha)", "Static",
         "CO₂ equivalent sequestered per hectare.\n"
         "= Carbon_Stock × (44/12) = Carbon_Stock × 3.67.\n"
         "Molecular ratio CO₂/C: 44/12 = 3.67.\n"
         "One tonne of carbon equals 3.67 tonnes of CO₂ equivalent."),
        ("कार्बन सारांश", "I", "भारित काठ घनत्व (टन/मी.³)", "Weighted Wood Density (t/m³)", "Static",
         "Volume-weighted average wood density across all species in the block.\n"
         "= Σ(Species_Volume × Species_Wood_Density) / Σ(Species_Volume).\n"
         "Used in AGB calculation to convert stem volume to biomass.\n"
         "Default: 0.65 t/m³ if no species data available."),

        # ── Sheet 4 ──
        ("DBH वर्ग विवरण", "A", "ब्लक", "Block", "Static", "Block name."),
        ("DBH वर्ग विवरण", "B", "DBH वर्ग (से.मी.)", "DBH Class (cm)", "Static",
         "Diameter at Breast Height class range in cm.\n"
         "8-class system:\n"
         "0_4 = Seedling (<4 cm), 4_10 = Sapling (4–10 cm),\n"
         "10_20 = Small Pole, 20_30 = Large Pole,\n"
         "30_40 = Small Tree, 40_50 = Medium Tree,\n"
         "50_60 = Large Tree, 60_plus = Very Large Tree (≥60 cm)."),
        ("DBH वर्ग विवरण", "C", "नेपाली वर्ग", "Nepali Class", "Static",
         "Nepali name for the DBH class:\n"
         "बिरुवा / लाथ्रा / सानो खाँवा / ठुलो खाँवा /\n"
         "सानो रुख / मझौला रुख / ठुलो रुख / अति ठुलो रुख."),
        ("DBH वर्ग विवरण", "D", "संख्या (प्रति हे.)", "Count (N/ha)", "Formula",
         "Number of individuals per hectare in this DBH class.\n"
         "Per-plot averaging: sum per class across plots ÷ total plots ÷ plot_area × 10000."),
        ("DBH वर्ग विवरण", "E", "काठ (घ.मी./हे.)", "Timber (m³/ha)", "Formula",
         "Net timber volume per hectare in this DBH class.\n"
         "Per-plot averaging: same method as count.\n"
         "Growing stock allocated to each DBH class."),
        ("DBH वर्ग विवरण", "F", "दाउरा (घ.मी./हे.)", "Firewood (m³/ha)", "Formula",
         "Fuelwood volume per hectare in this DBH class.\n"
         "Per-plot averaging: same method as count.\n"
         "Firewood volume allocated to each DBH class."),

        # ── Sheet 5 ──
        ("पुनरुत्पादन", "A", "ब्लक", "Block", "Static", "Block name."),
        ("पुनरुत्पादन", "B", "बिरुवा संख्या (०–४ से.मी.)", "Regeneration (0-4 cm) (N/ha)", "Formula",
         "Number of regeneration (seedling) individuals per hectare in the 0-4 cm DBH class. Per-plot averaging: sum count for 0-4 cm class across all sample plots, divide by total plots, divide by plot area, multiply by 10000. Corresponds to Regeneration stand type."),
        ("पुनरुत्पादन", "C", "लाथ्रा संख्या (४–१० से.मी.)", "Sapling (4-10 cm) (N/ha)", "Formula",
         "Number of sapling individuals per hectare in the 4-10 cm DBH class. Per-plot averaging: same methodology as regeneration. Corresponds to Sapling stand type."),
        ("पुनरुत्पादन", "D", "जम्मा पुनरुत्पादन (प्रति हे.)", "Total Regeneration (N/ha)", "Formula",
         "Total regeneration density per hectare combining seedlings and saplings. Formula: Regen (0-4 cm) N/ha + Sapling (4-10 cm) N/ha. Represents the total natural regeneration stocking of the forest."),
        ("पुनरुत्पादन", "E", "पुनरुत्पादन अवस्था", "Regeneration Condition", "Static",
         "Regeneration condition assessment based on combined Regen and Sapling densities. Good: Regen ≥5000/ha AND Sapling ≥2000/ha. Moderate: Regen ≥2000/ha AND Sapling ≥800/ha. Weak: below Moderate threshold. Uses AND logic — both conditions must be met simultaneously."),
        ("पुनरुत्पादन", "F", "मापदण्ड (बिरुवा/लाथ्रा)", "Threshold (Regen/Sapling)", "Static",
         "Displays the minimum threshold values used to determine the regeneration condition. Good = ≥5000 Regen + ≥2000 Sapling per ha. Moderate = ≥2000 Regen + ≥800 Sapling per ha. Weak = below 2000 Regen or 800 Sapling per ha."),
    ]

    for i, (sheet, col, ne_h, en_h, typ, desc) in enumerate(DESCRIPTIONS, 2):
        ws6.cell(row=i, column=1, value=sheet).font = DATA_FONT
        ws6.cell(row=i, column=2, value=col).font = DATA_FONT
        ws6.cell(row=i, column=3, value=ne_h).font = DATA_FONT
        ws6.cell(row=i, column=4, value=en_h).font = DATA_FONT
        ws6.cell(row=i, column=5, value=typ).font = DATA_FONT
        cell_desc = ws6.cell(row=i, column=6, value=desc)
        cell_desc.font = DATA_FONT
        cell_desc.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, 7):
            ws6.cell(row=i, column=c).border = THIN_BORDER

    _set_col_widths(ws6, [22, 8, 28, 28, 10, 80])
    ws6.freeze_panes = "A2"

    # =========================================================================
    # Sheets 7-15 — Management Plan Summary (from field_inventory_mgmt_data)
    # =========================================================================
    mgmt_data = get_management_plan_data(
        db, field_inventory_id, calculation_id,
        aah_good, aah_moderate, aah_weak,
    )

    # ── Sheet 7: प्रजाती संरचना (Species Composition) ──
    ws7 = wb.create_sheet("प्रजाती संरचना")
    _write_species_composition_sheet(ws7, mgmt_data["species_composition"])

    # ── Sheet 8: ब्लक तुलनात्मक (Block Comparison) ──
    ws8 = wb.create_sheet("ब्लक तुलनात्मक")
    _write_block_comparison_sheet(ws8, mgmt_data["block_comparison"])

    # ── Sheet 9: वार्षिक फसल योजना (Annual Harvest Plan) ──
    ws9 = wb.create_sheet("वार्षिक फसल योजना")
    _write_annual_harvest_sheet(ws9, mgmt_data["annual_harvest_plan"])

    # ── Sheet 10: वन स्थिति सारांश (Forest Condition Summary) ──
    ws10 = wb.create_sheet("वन स्थिति सारांश")
    _write_forest_condition_sheet(ws10, mgmt_data["forest_condition_summary"])

    # ── Sheet 11: DBH वर्ग आयतन (DBH Class Volume Distribution) ──
    ws11 = wb.create_sheet("DBH वर्ग आयतन")
    _write_dbh_class_volume_sheet(ws11, mgmt_data["dbh_class_volume"])

    # ── Sheet 12: कार्बन भण्डार (Carbon Per Block) ──
    ws12 = wb.create_sheet("कार्बन भण्डार")
    _write_carbon_per_block_sheet(ws12, mgmt_data["carbon_per_block"])

    # ── Sheet 13: वृद्धि दर (Growth Rate Classification) ──
    ws13 = wb.create_sheet("वृद्धि दर वर्गीकरण")
    _write_growth_rate_sheet(ws13, mgmt_data["growth_rate_classification"])

    # ── Sheet 14: रुख संरचना (Stand Structure Profile) ──
    ws14 = wb.create_sheet("रुख संरचना प्रोफाइल")
    _write_stand_structure_sheet(ws14, mgmt_data["stand_structure"])

    # ── Sheet 15: उत्पादकता (Productivity Classification) ──
    ws15 = wb.create_sheet("उत्पादकता वर्गीकरण")
    _write_productivity_sheet(ws15, mgmt_data["productivity_classification"])

    # ── Save to bytes ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
