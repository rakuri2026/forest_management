"""
Export functionality for fieldbook and sampling data.
Supports CSV, Excel, GPX, GeoJSON, and KML formats.
"""
from sqlalchemy.orm import Session
from uuid import UUID
import csv
import io
from typing import List, Dict, Any
from xml.etree import ElementTree as ET

from app.models.fieldbook import Fieldbook
from app.models.sampling import SamplingDesign


# ===========================
# Fieldbook Export Functions
# ===========================

def _safe_string(val) -> str:
    s = str(val or '').strip()
    try:
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        s = ILLEGAL_CHARACTERS_RE.sub('', s)
    except ImportError:
        import re
        s = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', s)
    if s.startswith(('=', '+', '-')):
        s = f"'{s}"
    return s

def export_fieldbook_csv(db: Session, calculation_id: UUID) -> bytes:
    """
    Export fieldbook to CSV format with topographic features.
    Uses optimized pre-clipping algorithm (same as sampling export).
    """
    # CRITICAL: Rollback any previous failed transactions at the START
    try:
        db.rollback()
    except:
        pass

    points = db.query(Fieldbook).filter(
        Fieldbook.calculation_id == calculation_id
    ).order_by(Fieldbook.point_number).all()

    if not points:
        raise ValueError("No fieldbook points found for this calculation")

    # Get calculation boundary for pre-clipping topographic features
    from app.models.calculation import Calculation
    from sqlalchemy import text
    calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    boundary_wkt = None
    if calc:
        boundary_query = text("""
            SELECT ST_AsText(boundary_geom) as wkt
            FROM public.calculations
            WHERE id = :calc_id
        """)
        boundary_result = db.execute(boundary_query, {"calc_id": str(calculation_id)}).first()
        if boundary_result:
            boundary_wkt = boundary_result.wkt

    # Import required modules for topographic feature extraction
    from app.utils.geospatial_vector_optimized import (
        preclip_topographic_features,
        find_nearest_topographic_feature_optimized
    )
    import logging
    logger = logging.getLogger(__name__)

    # Pre-clip ridge/river data ONCE for performance (20-100x faster!)
    clipped_features = None
    if boundary_wkt:
        try:
            # Rollback any previous failed transactions
            try:
                db.rollback()
            except:
                pass

            clipped_features = preclip_topographic_features(
                db=db,
                boundary_wkt=boundary_wkt,
                buffer_meters=200.0
            )
            # DEBUG: Log what we got
            if clipped_features:
                ridge_count = len(clipped_features.get('ridges', []))
                river_count = len(clipped_features.get('rivers', []))
                logger.info(f"Pre-clipped {ridge_count} ridges and {river_count} rivers for fieldbook CSV export")
            else:
                logger.warning("Pre-clipping returned None - ridge/river columns will be empty!")
        except Exception as e:
            logger.error(f"Failed to pre-clip topographic features: {e}", exc_info=True)
            # Rollback the failed transaction
            try:
                db.rollback()
            except:
                pass
            clipped_features = None

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header with topographic features
    # If you see "TEST_NEW_CODE" as first column, the new code loaded successfully!
    writer.writerow([
        'TEST_NEW_CODE_LOADED',  # DEBUG: If you see this, new code is running!
        'Point No',
        'Type',
        'Block No',
        'Block Name',
        'Sub-area Name',
        'Is Excluded Zone',
        'Longitude',
        'Latitude',
        'Easting UTM',
        'Northing UTM',
        'UTM Zone',
        'Azimuth (deg)',
        'Distance (m)',
        'Elevation (m)',
        'Nearest Feature',
        'Feature Type',
        'Distance to Feature (m)',
        'Direction to Feature',
        'Verified',
        'Remarks'
    ])

    # Write data rows with topographic features
    for point in points:
        # Extract topographic features (ridge/river) using optimized algorithm
        feature_name = ''
        feature_type = ''
        feature_distance = ''
        feature_direction = ''

        if clipped_features and point.longitude and point.latitude:
            try:
                topo_feature = find_nearest_topographic_feature_optimized(
                    db=db,
                    longitude=float(point.longitude),
                    latitude=float(point.latitude),
                    clipped_features=clipped_features,
                    search_radius_meters=100.0,
                    prefer_rivers=True,
                    min_distance_threshold=20.0
                )

                if topo_feature:
                    feature_name = topo_feature.get("feature_name", "")
                    feature_type = topo_feature.get("feature_type", "")
                    feature_distance = f'{int(topo_feature.get("distance_meters", 0))}'
                    feature_direction = topo_feature.get("direction", "")
            except Exception as e:
                logger.warning(f"Failed to find topographic feature for point {point.point_number}: {e}")

        writer.writerow([
            'YES',
            _safe_string(f'P{point.point_number}'),
            _safe_string(point.point_type),
            _safe_string(point.block_number),
            _safe_string(point.block_name),
            _safe_string(point.sub_area_name),
            _safe_string('Yes' if point.is_excluded else 'No' if point.sub_area_name else ''),
            f'{point.latitude:.7f}' if point.latitude else '',
            f'{point.longitude:.7f}' if point.longitude else '',
            f'{point.easting_utm:.2f}' if point.easting_utm else '',
            f'{point.northing_utm:.2f}' if point.northing_utm else '',
            _safe_string(point.utm_zone),
            f'{point.azimuth_to_next:.2f}' if point.azimuth_to_next else '',
            f'{point.distance_to_next:.2f}' if point.distance_to_next else '',
            f'{point.elevation:.2f}' if point.elevation else '',
            _safe_string(feature_name),
            _safe_string(feature_type),
            feature_distance,
            _safe_string(feature_direction),
            _safe_string('Yes' if point.is_verified else 'No'),
            _safe_string(point.remarks)
        ])

    return output.getvalue().encode('utf-8')


def export_fieldbook_excel(db: Session, calculation_id: UUID) -> bytes:
    """
    Export fieldbook to Excel format with topographic features.
    Uses optimized pre-clipping algorithm (same as sampling export).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    # CRITICAL: Rollback any previous failed transactions at the START
    try:
        db.rollback()
    except:
        pass

    points = db.query(Fieldbook).filter(
        Fieldbook.calculation_id == calculation_id
    ).order_by(Fieldbook.point_number).all()

    if not points:
        raise ValueError("No fieldbook points found for this calculation")

    # Get calculation boundary for pre-clipping topographic features
    from app.models.calculation import Calculation
    from sqlalchemy import text
    calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    boundary_wkt = None
    if calc:
        boundary_query = text("""
            SELECT ST_AsText(boundary_geom) as wkt
            FROM public.calculations
            WHERE id = :calc_id
        """)
        boundary_result = db.execute(boundary_query, {"calc_id": str(calculation_id)}).first()
        if boundary_result:
            boundary_wkt = boundary_result.wkt

    # Import required modules for topographic feature extraction
    from app.utils.geospatial_vector_optimized import (
        preclip_topographic_features,
        find_nearest_topographic_feature_optimized
    )
    import logging
    logger = logging.getLogger(__name__)

    # Pre-clip ridge/river data ONCE for performance (20-100x faster!)
    clipped_features = None
    if boundary_wkt:
        try:
            # Rollback any previous failed transactions
            try:
                db.rollback()
            except:
                pass

            clipped_features = preclip_topographic_features(
                db=db,
                boundary_wkt=boundary_wkt,
                buffer_meters=200.0
            )
            # DEBUG: Log what we got
            if clipped_features:
                ridge_count = len(clipped_features.get('ridges', []))
                river_count = len(clipped_features.get('rivers', []))
                logger.info(f"Pre-clipped {ridge_count} ridges and {river_count} rivers for fieldbook Excel export")
            else:
                logger.warning("Pre-clipping returned None - ridge/river columns will be empty!")
        except Exception as e:
            logger.error(f"Failed to pre-clip topographic features: {e}", exc_info=True)
            # Rollback the failed transaction
            try:
                db.rollback()
            except:
                pass
            clipped_features = None

    # Create workbook
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = "Fieldbook Summary"
    ws_summary['A1'].font = Font(size=14, bold=True)

    summary_data = [
        ['Total Points', len(points)],
        ['Vertices', sum(1 for p in points if p.point_type == 'vertex')],
        ['Interpolated Points', sum(1 for p in points if p.point_type == 'interpolated')],
        ['Total Perimeter (m)', sum(p.distance_to_next or 0 for p in points)],
        ['Min Elevation (m)', min((p.elevation for p in points if p.elevation), default=None)],
        ['Max Elevation (m)', max((p.elevation for p in points if p.elevation), default=None)],
        ['Avg Elevation (m)', sum(p.elevation or 0 for p in points) / len(points) if points else None],
        ['Verified Points', sum(1 for p in points if p.is_verified)]
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        ws_summary[f'A{i}'] = label
        ws_summary[f'A{i}'].font = Font(bold=True)
        if value is not None:
            ws_summary[f'B{i}'] = round(value, 2) if isinstance(value, float) else value

# Points sheet
    ws_points = wb.create_sheet("Points")

    # Header row with topographic features
    headers = [
        'Point No', 'Type', 'Block No', 'Block Name', 'Sub-area Name', 'Is Excluded Zone',
        'Latitude',
        'Longitude',
        'Easting UTM', 'Northing UTM', 'UTM Zone',
        'Azimuth (deg)', 'Distance (m)', 'Elevation (m)',
        'Nearest Feature', 'Feature Type', 'Distance to Feature (m)', 'Direction to Feature',
        'Verified', 'Remarks'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws_points.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Data rows with topographic features
    for row, point in enumerate(points, start=2):
        # Extract topographic features (ridge/river) using optimized algorithm
        feature_name = ''
        feature_type = ''
        feature_distance = ''
        feature_direction = ''

        if clipped_features and point.longitude and point.latitude:
            try:
                topo_feature = find_nearest_topographic_feature_optimized(
                    db=db,
                    longitude=float(point.longitude),
                    latitude=float(point.latitude),
                    clipped_features=clipped_features,
                    search_radius_meters=100.0,
                    prefer_rivers=True,
                    min_distance_threshold=20.0
                )

                if topo_feature:
                    feature_name = topo_feature.get("feature_name", "")
                    feature_type = topo_feature.get("feature_type", "")
                    feature_distance = int(topo_feature.get("distance_meters", 0))
                    feature_direction = topo_feature.get("direction", "")
            except Exception as e:
                logger.warning(f"Failed to find topographic feature for point {point.point_number}: {e}")

        ws_points.cell(row=row, column=1, value=_safe_string(f'P{point.point_number}'))
        ws_points.cell(row=row, column=2, value=_safe_string(point.point_type))
        ws_points.cell(row=row, column=3, value=_safe_string(point.block_number))
        ws_points.cell(row=row, column=4, value=_safe_string(point.block_name))
        ws_points.cell(row=row, column=5, value=_safe_string(point.sub_area_name))
        ws_points.cell(row=row, column=6, value=_safe_string('Yes' if point.is_excluded else 'No' if point.sub_area_name else ''))
        ws_points.cell(row=row, column=7, value=round(point.latitude, 7) if point.latitude else '')
        ws_points.cell(row=row, column=8, value=round(point.longitude, 7) if point.longitude else '')
        ws_points.cell(row=row, column=9, value=round(point.easting_utm, 2) if point.easting_utm else '')
        ws_points.cell(row=row, column=10, value=round(point.northing_utm, 2) if point.northing_utm else '')
        ws_points.cell(row=row, column=11, value=_safe_string(point.utm_zone))
        ws_points.cell(row=row, column=12, value=round(point.azimuth_to_next, 2) if point.azimuth_to_next else '')
        ws_points.cell(row=row, column=13, value=round(point.distance_to_next, 2) if point.distance_to_next else '')
        ws_points.cell(row=row, column=14, value=round(point.elevation, 2) if point.elevation else '')
        ws_points.cell(row=row, column=15, value=_safe_string(feature_name))
        ws_points.cell(row=row, column=16, value=_safe_string(feature_type))
        ws_points.cell(row=row, column=17, value=feature_distance)
        ws_points.cell(row=row, column=18, value=_safe_string(feature_direction))
        ws_points.cell(row=row, column=19, value=_safe_string('Yes' if point.is_verified else 'No'))
        ws_points.cell(row=row, column=20, value=_safe_string(point.remarks))

        # Highlight verified points
        if point.is_verified:
            for col in range(1, 21):
                ws_points.cell(row=row, column=col).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )

    # Adjust column widths
    for col in range(1, 21):
        ws_points.column_dimensions[chr(64 + col)].width = 15

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def export_fieldbook_gpx(db: Session, calculation_id: UUID) -> bytes:
    """
    Export fieldbook to GPX format (GPS Exchange Format).
    """
    points = db.query(Fieldbook).filter(
        Fieldbook.calculation_id == calculation_id
    ).order_by(Fieldbook.point_number).all()

    if not points:
        raise ValueError("No fieldbook points found for this calculation")

    # Create GPX XML
    gpx = ET.Element('gpx', {
        'version': '1.1',
        'creator': 'Community Forest Management System',
        'xmlns': 'http://www.topografix.com/GPX/1/1',
        'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance',
        'xsi:schemaLocation': 'http://www.topografix.com/GPX/1/1 http://www.topografix.com/GPX/1/1/gpx.xsd'
    })

    metadata = ET.SubElement(gpx, 'metadata')
    ET.SubElement(metadata, 'name').text = f'Fieldbook - Calculation {calculation_id}'
    ET.SubElement(metadata, 'desc').text = f'Boundary fieldbook with {len(points)} points'

    for point in points:
        if point.latitude and point.longitude:
            wpt = ET.SubElement(gpx, 'wpt', {
                'lat': f'{point.latitude:.7f}',
                'lon': f'{point.longitude:.7f}'
            })

            ET.SubElement(wpt, 'name').text = f'P{point.point_number}'
            ET.SubElement(wpt, 'type').text = point.point_type

            if point.elevation:
                ET.SubElement(wpt, 'ele').text = f'{point.elevation:.2f}'

            desc_parts = []
            if point.block_name:
                desc_parts.append(f'Block: {point.block_name}')
            if point.sub_area_name:
                desc_parts.append(f'Sub-area: {point.sub_area_name}')
                if point.is_excluded:
                    desc_parts.append('Type: Excluded Zone (Private Land)')
            if point.azimuth_to_next:
                desc_parts.append(f'Azimuth: {point.azimuth_to_next:.2f}°')
            if point.distance_to_next:
                desc_parts.append(f'Distance: {point.distance_to_next:.2f}m')
            if point.remarks:
                desc_parts.append(f'Remarks: {point.remarks}')

            if desc_parts:
                ET.SubElement(wpt, 'desc').text = ' | '.join(desc_parts)

    # Convert to bytes
    tree = ET.ElementTree(gpx)
    output = io.BytesIO()
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)
    return output.read()


def export_fieldbook_geojson(db: Session, calculation_id: UUID) -> Dict[str, Any]:
    """
    Export fieldbook to GeoJSON format.
    """
    points = db.query(Fieldbook).filter(
        Fieldbook.calculation_id == calculation_id
    ).order_by(Fieldbook.point_number).all()

    if not points:
        raise ValueError("No fieldbook points found for this calculation")

    features = []
    for point in points:
        if point.latitude and point.longitude:
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [float(point.latitude), float(point.longitude)]
                },
                "properties": {
                    "point_number": point.point_number,
                    "point_type": point.point_type,
                    "elevation": round(float(point.elevation), 2) if point.elevation else None,
                    "azimuth_to_next": round(float(point.azimuth_to_next), 2) if point.azimuth_to_next else None,
                    "distance_to_next": round(float(point.distance_to_next), 2) if point.distance_to_next else None,
                    "easting_utm": round(float(point.easting_utm), 2) if point.easting_utm else None,
                    "northing_utm": round(float(point.northing_utm), 2) if point.northing_utm else None,
                    "utm_zone": point.utm_zone,
                    "is_verified": point.is_verified,
                    "remarks": point.remarks
                }
            }
            features.append(feature)

    return {
        "type": "FeatureCollection",
        "features": features
    }


# ===========================
# Sampling Export Functions
# ===========================

def export_sampling_csv(db: Session, design_id: UUID) -> bytes:
    """
    Export sampling points to CSV format with complete field data.
    Includes: Plot No, Block, Longitude, Latitude, Elevation, UTM Easting, UTM Northing, UTM Zone, Distance from Boundary
    """
    # CRITICAL: Rollback any previous failed transactions at the START
    # This prevents "InFailedSqlTransaction" errors from cascading
    try:
        db.rollback()
    except:
        pass

    design = db.query(SamplingDesign).filter(SamplingDesign.id == design_id).first()

    if not design or not design.points_geometry:
        raise ValueError("Sampling design not found or has no points")

    # Get geometry as WKT using PostGIS
    from sqlalchemy import text
    wkt_query = text("""
        SELECT ST_AsText(points_geometry) as wkt
        FROM public.sampling_designs
        WHERE id = :design_id
    """)
    result = db.execute(wkt_query, {"design_id": str(design_id)}).first()

    if not result or not result.wkt:
        raise ValueError("Failed to retrieve geometry as WKT")

    # Parse MultiPoint geometry
    from shapely import wkt as shapely_wkt
    multipoint = shapely_wkt.loads(result.wkt)

    # Get block assignment
    block_assignment = design.points_block_assignment or []

    # Get calculation boundary for distance calculation
    from app.models.calculation import Calculation
    calc = db.query(Calculation).filter(Calculation.id == design.calculation_id).first()
    boundary_wkt = None
    if calc:
        boundary_query = text("""
            SELECT ST_AsText(boundary_geom) as wkt
            FROM public.calculations
            WHERE id = :calc_id
        """)
        boundary_result = db.execute(boundary_query, {"calc_id": str(design.calculation_id)}).first()
        if boundary_result:
            boundary_wkt = boundary_result.wkt

    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header with all displayed fields including topographic features
    writer.writerow([
        'Plot No',
        'Block',
        'Zone Type',
        'Longitude',
        'Latitude',
        'Elevation (m)',
        'UTM Easting',
        'UTM Northing',
        'UTM Zone',
        'Distance from Boundary (m)',
        'Nearest Feature',
        'Feature Type',
        'Distance to Feature (m)',
        'Direction to Feature'
    ])

    # Import required modules for calculations
    from app.utils.geospatial import extract_elevation_at_point
    from app.utils.geospatial_vector_optimized import (
        preclip_topographic_features,
        find_nearest_topographic_feature_optimized
    )
    from pyproj import Transformer
    import logging
    logger = logging.getLogger(__name__)

    # Pre-clip ridge/river data ONCE for performance (20-100x faster!)
    # IMPORTANT: Wrap in transaction handling to prevent "InFailedSqlTransaction" errors
    clipped_features = None
    if boundary_wkt:
        try:
            # Rollback any previous failed transactions
            try:
                db.rollback()
            except:
                pass

            clipped_features = preclip_topographic_features(
                db=db,
                boundary_wkt=boundary_wkt,
                buffer_meters=200.0
            )
            # DEBUG: Log what we got
            if clipped_features:
                ridge_count = len(clipped_features.get('ridges', []))
                river_count = len(clipped_features.get('rivers', []))
                logger.info(f"Pre-clipped {ridge_count} ridges and {river_count} rivers for CSV export")
                if ridge_count == 0 and river_count == 0:
                    logger.warning("Pre-clipping returned 0 features - ridge/river columns will be empty!")
            else:
                logger.warning("Pre-clipping returned None - ridge/river columns will be empty!")
        except Exception as e:
            logger.error(f"Failed to pre-clip topographic features: {e}", exc_info=True)
            # Rollback the failed transaction
            try:
                db.rollback()
            except:
                pass
            clipped_features = None

    # Write points with complete data
    for i, point in enumerate(multipoint.geoms):
        lon, lat = point.x, point.y

        # Find block assignment for this point
        block_info = next((b for b in block_assignment if b.get('point_index') == i), None)
        block_name = block_info.get('block_name', f'Plot {i+1}') if block_info else f'Plot {i+1}'
        zone_type = block_info.get('zone_type', 'Productive') if block_info else 'Productive'
        zone_type_display = zone_type.capitalize() if zone_type else 'Productive'

        # Calculate UTM coordinates
        utm_zone = 44 if lon < 84 else 45
        transformer = Transformer.from_crs(f"EPSG:4326", f"EPSG:326{utm_zone}", always_xy=True)
        utm_easting, utm_northing = transformer.transform(lon, lat)

        # Extract elevation
        elevation_m = extract_elevation_at_point(db, lon, lat)

        # Calculate distance from boundary
        distance_from_boundary = None
        if boundary_wkt:
            try:
                boundary_geom = shapely_wkt.loads(boundary_wkt)
                distance_from_boundary = point.distance(boundary_geom.boundary) * 111320
            except:
                pass

        # Extract topographic features (ridge/river)
        topo_feature = None
        feature_name = ''
        feature_type = ''
        feature_distance = ''
        feature_direction = ''

        if clipped_features:
            try:
                topo_feature = find_nearest_topographic_feature_optimized(
                    db=db,
                    longitude=lon,
                    latitude=lat,
                    clipped_features=clipped_features,
                    search_radius_meters=100.0,
                    prefer_rivers=True,
                    min_distance_threshold=20.0
                )

                if topo_feature:
                    feature_name = topo_feature.get("feature_name", "")
                    feature_type = topo_feature.get("feature_type", "")
                    feature_distance = f'{int(topo_feature.get("distance_meters", 0))}'
                    feature_direction = topo_feature.get("direction", "")
                else:
                    if i == 0:
                        logger.warning(f"Point 1: No topographic feature found within 1000m")
            except Exception as e:
                logger.warning(f"Failed to find topographic feature for point {i+1}: {e}", exc_info=True)

        # Write CSV row
        writer.writerow([
            i + 1,  # Plot number starts at 1
            block_name,
            zone_type_display,
            f'{lon:.7f}',
            f'{lat:.7f}',
            f'{int(elevation_m)}' if elevation_m and elevation_m > 0 else 'N/A',
            f'{utm_easting:.2f}',
            f'{utm_northing:.2f}',
            f'{utm_zone}N',
            f'{distance_from_boundary:.2f}' if distance_from_boundary else '',
            feature_name,
            feature_type,
            feature_distance,
            feature_direction
        ])

    return output.getvalue().encode('utf-8')


def export_sampling_gpx(db: Session, design_id: UUID) -> bytes:
    """
    Export sampling points to GPX format with complete field data.
    Includes elevation, UTM coordinates, block information, and zone type.
    """
    # CRITICAL: Rollback any previous failed transactions at the START
    try:
        db.rollback()
    except:
        pass

    design = db.query(SamplingDesign).filter(SamplingDesign.id == design_id).first()

    if not design or not design.points_geometry:
        raise ValueError("Sampling design not found or has no points")

    # Get geometry as WKT using PostGIS
    from sqlalchemy import text
    wkt_query = text("""
        SELECT ST_AsText(points_geometry) as wkt
        FROM public.sampling_designs
        WHERE id = :design_id
    """)
    result = db.execute(wkt_query, {"design_id": str(design_id)}).first()

    if not result or not result.wkt:
        raise ValueError("Failed to retrieve geometry as WKT")

    # Parse MultiPoint geometry
    from shapely import wkt as shapely_wkt
    multipoint = shapely_wkt.loads(result.wkt)

    # Get block assignment
    block_assignment = design.points_block_assignment or []

    # Get calculation boundary for distance calculation
    from app.models.calculation import Calculation
    calc = db.query(Calculation).filter(Calculation.id == design.calculation_id).first()
    boundary_wkt = None
    if calc:
        boundary_query = text("""
            SELECT ST_AsText(boundary_geom) as wkt
            FROM public.calculations
            WHERE id = :calc_id
        """)
        boundary_result = db.execute(boundary_query, {"calc_id": str(design.calculation_id)}).first()
        if boundary_result:
            boundary_wkt = boundary_result.wkt

    # Import required modules for calculations
    from app.utils.geospatial import extract_elevation_at_point
    from pyproj import Transformer

    # Create GPX XML
    gpx = ET.Element('gpx', {
        'version': '1.1',
        'creator': 'Community Forest Management System',
        'xmlns': 'http://www.topografix.com/GPX/1/1'
    })

    metadata = ET.SubElement(gpx, 'metadata')
    ET.SubElement(metadata, 'name').text = f'Sampling Points - {design.sampling_type}'
    ET.SubElement(metadata, 'desc').text = f'{design.total_points} sampling plots'

    for i, point in enumerate(multipoint.geoms):
        lon, lat = point.x, point.y

        # Find block assignment for this point
        block_info = next((b for b in block_assignment if b.get('point_index') == i), None)
        block_name = block_info.get('block_name', f'Plot {i+1}') if block_info else f'Plot {i+1}'
        zone_type = block_info.get('zone_type', 'Productive') if block_info else 'Productive'

        # Calculate UTM coordinates
        utm_zone = 44 if lon < 84 else 45
        transformer = Transformer.from_crs(f"EPSG:4326", f"EPSG:326{utm_zone}", always_xy=True)
        utm_easting, utm_northing = transformer.transform(lon, lat)

        # Extract elevation
        elevation_m = extract_elevation_at_point(db, lon, lat)

        # Calculate distance from boundary
        distance_from_boundary = None
        if boundary_wkt:
            try:
                boundary_geom = shapely_wkt.loads(boundary_wkt)
                distance_from_boundary = point.distance(boundary_geom.boundary) * 111320
            except:
                pass

        wpt = ET.SubElement(gpx, 'wpt', {
            'lat': f'{lat:.7f}',
            'lon': f'{lon:.7f}'
        })
        ET.SubElement(wpt, 'name').text = f'Plot {i+1}'
        ET.SubElement(wpt, 'type').text = 'sampling_plot'

        # Add elevation
        if elevation_m:
            ET.SubElement(wpt, 'ele').text = f'{int(elevation_m)}'

        # Build description with all fields
        desc_parts = []
        desc_parts.append(f'Block: {block_name}')
        desc_parts.append(f'Zone: {zone_type.capitalize()}')
        desc_parts.append(f'UTM: {utm_easting:.2f}E, {utm_northing:.2f}N ({utm_zone}N)')
        if elevation_m:
            desc_parts.append(f'Elevation: {int(elevation_m)}m')
        if distance_from_boundary:
            desc_parts.append(f'Distance from Boundary: {distance_from_boundary:.2f}m')
        if design.plot_shape:
            desc_parts.append(f'Plot: {design.plot_shape}')
            if design.plot_radius_meters:
                desc_parts.append(f'Radius: {design.plot_radius_meters}m')

        ET.SubElement(wpt, 'desc').text = ' | '.join(desc_parts)

    # Convert to bytes
    tree = ET.ElementTree(gpx)
    output = io.BytesIO()
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)
    return output.read()


def export_sampling_kml(db: Session, design_id: UUID) -> bytes:
    """
    Export sampling points to KML format (Google Earth) with complete field data.
    Includes elevation, UTM coordinates, and block information.
    """
    design = db.query(SamplingDesign).filter(SamplingDesign.id == design_id).first()

    if not design or not design.points_geometry:
        raise ValueError("Sampling design not found or has no points")

    # Get geometry as WKT using PostGIS
    from sqlalchemy import text
    wkt_query = text("""
        SELECT ST_AsText(points_geometry) as wkt
        FROM public.sampling_designs
        WHERE id = :design_id
    """)
    result = db.execute(wkt_query, {"design_id": str(design_id)}).first()

    if not result or not result.wkt:
        raise ValueError("Failed to retrieve geometry as WKT")

    # Parse MultiPoint geometry
    from shapely import wkt as shapely_wkt
    multipoint = shapely_wkt.loads(result.wkt)

    # Get block assignment
    block_assignment = design.points_block_assignment or []

    # Get calculation boundary for distance calculation
    from app.models.calculation import Calculation
    calc = db.query(Calculation).filter(Calculation.id == design.calculation_id).first()
    boundary_wkt = None
    if calc:
        boundary_query = text("""
            SELECT ST_AsText(boundary_geom) as wkt
            FROM public.calculations
            WHERE id = :calc_id
        """)
        boundary_result = db.execute(boundary_query, {"calc_id": str(design.calculation_id)}).first()
        if boundary_result:
            boundary_wkt = boundary_result.wkt

    # Import required modules for calculations
    from app.utils.geospatial import extract_elevation_at_point
    from app.utils.geospatial_vector_optimized import (
        preclip_topographic_features,
        find_nearest_topographic_feature_optimized
    )
    from pyproj import Transformer
    import logging
    logger = logging.getLogger(__name__)
    import math

    # Create KML XML
    kml = ET.Element('kml', {'xmlns': 'http://www.opengis.net/kml/2.2'})
    document = ET.SubElement(kml, 'Document')

    ET.SubElement(document, 'name').text = f'Sampling Points - {design.sampling_type}'
    ET.SubElement(document, 'description').text = (
        f'Sampling design with {design.total_points} plots. '
        f'Type: {design.sampling_type}. '
        f'Intensity: {design.intensity_per_hectare}/ha.'
    )

    # Add style for sampling points
    style = ET.SubElement(document, 'Style', {'id': 'samplingPlot'})
    icon_style = ET.SubElement(style, 'IconStyle')
    ET.SubElement(icon_style, 'scale').text = '1.2'
    icon = ET.SubElement(icon_style, 'Icon')
    ET.SubElement(icon, 'href').text = 'http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png'

    # Add folder for points
    folder = ET.SubElement(document, 'Folder')
    ET.SubElement(folder, 'name').text = 'Sampling Plots'

    for i, point in enumerate(multipoint.geoms):
        lon, lat = point.x, point.y

        # Find block assignment for this point
        block_info = next((b for b in block_assignment if b.get('point_index') == i), None)
        block_name = block_info.get('block_name', f'Plot {i+1}') if block_info else f'Plot {i+1}'
        zone_type = block_info.get('zone_type', 'Productive') if block_info else 'Productive'

        # Calculate UTM coordinates
        utm_zone = 44 if lon < 84 else 45
        transformer = Transformer.from_crs(f"EPSG:4326", f"EPSG:326{utm_zone}", always_xy=True)
        utm_easting, utm_northing = transformer.transform(lon, lat)

        # Extract elevation
        elevation_m = extract_elevation_at_point(db, lon, lat)

        # Calculate distance from boundary
        distance_from_boundary = None
        if boundary_wkt:
            try:
                boundary_geom = shapely_wkt.loads(boundary_wkt)
                distance_from_boundary = point.distance(boundary_geom.boundary) * 111320
            except:
                pass

        placemark = ET.SubElement(folder, 'Placemark')
        ET.SubElement(placemark, 'name').text = f'Plot {i+1}'
        ET.SubElement(placemark, 'styleUrl').text = '#samplingPlot'

        # Build HTML description with all fields
        desc = f'<b>Plot {i+1}</b><br/>'
        desc += f'Block: {block_name}<br/>'
        desc += f'Zone Type: {zone_type.capitalize()}<br/>'
        desc += f'Longitude: {lon:.7f}<br/>'
        desc += f'Latitude: {lat:.7f}<br/>'
        if elevation_m:
            desc += f'Elevation: {int(elevation_m)}m ASLM<br/>'
        desc += f'UTM Easting: {utm_easting:.2f}<br/>'
        desc += f'UTM Northing: {utm_northing:.2f}<br/>'
        desc += f'UTM Zone: {utm_zone}N<br/>'
        if distance_from_boundary:
            desc += f'Distance from Boundary: {distance_from_boundary:.2f}m<br/>'

        if design.plot_shape:
            desc += f'<br/><b>Plot Details:</b><br/>'
            desc += f'Shape: {design.plot_shape}<br/>'
            if design.plot_radius_meters:
                plot_area = math.pi * float(design.plot_radius_meters) ** 2
                desc += f'Radius: {design.plot_radius_meters}m<br/>'
                desc += f'Area: {plot_area:.2f} m²'
            elif design.plot_length_meters and design.plot_width_meters:
                plot_area = float(design.plot_length_meters) * float(design.plot_width_meters)
                desc += f'Length: {design.plot_length_meters}m x Width: {design.plot_width_meters}m<br/>'
                desc += f'Area: {plot_area:.2f} m²'

        ET.SubElement(placemark, 'description').text = desc

        point_elem = ET.SubElement(placemark, 'Point')
        # KML format: lon,lat,elevation
        elev_val = int(elevation_m) if elevation_m else 0
        ET.SubElement(point_elem, 'coordinates').text = f'{lon:.7f},{lat:.7f},{elev_val}'

    # Convert to bytes
    tree = ET.ElementTree(kml)
    output = io.BytesIO()
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)
    return output.read()
