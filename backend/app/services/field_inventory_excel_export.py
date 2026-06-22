"""
Field Inventory Excel Export — entry/coefficients/plot sizes static, volume + per-hectare as formulas.
"""
import io
import math
import logging
from datetime import datetime
from typing import Dict, Optional
from uuid import UUID
from sqlalchemy.orm import Session
from sqlalchemy import text

logger = logging.getLogger(__name__)


def cl(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s
COL = [cl(i) for i in range(1, 100)]

# ── Column letter references ──
C_ = COL[2]      # 3: Stand Type
H_ = COL[7]      # 8: DBH
I_ = COL[8]      # 9: Height
K_ = COL[10]     # 11: Tree Class
N_ = COL[13]     # 14: Total Sample Plots

CA = COL[14]     # 15: Coeff_a
CB = COL[15]     # 16: Coeff_b
CC = COL[16]     # 17: Coeff_c
CA1 = COL[17]    # 18: Coeff_a1
CB1 = COL[18]    # 19: Coeff_b1
CS = COL[19]     # 20: Coeff_s
CM = COL[20]     # 21: Coeff_m
CBG = COL[21]    # 22: Coeff_bg

WD = COL[22]     # 23: Wood Density
SV = COL[25]     # 26: Stem Volume
BR = COL[26]     # 27: Branchwood Ratio
BV = COL[27]     # 28: Branchwood Volume
TV = COL[28]     # 29: Tree Volume
TDR = COL[29]    # 30: 10-cm Top Diameter Ratio
TVOL = COL[30]   # 31: 10-cm Top Volume
GT = COL[31]     # 32: Gross Timber Volume
RF = COL[32]     # 33: Recovery Factor (%)
NT = COL[33]     # 34: Net Timber Volume
NTCFT = COL[34]  # 35: Net Timber Volume (cft)
FW = COL[35]     # 36: Fuelwood Volume (m3)
FWC = COL[36]    # 37: Firewood (chatta)
BA = COL[37]     # 38: Basal Area
FSM = COL[24]    # 25: Full Stem Merchantable
AGB = COL[39]    # 40: AGB (t)
BGB = COL[40]    # 41: BGB (t)
TB = COL[41]     # 42: Total Biomass (t)
CARB = COL[42]   # 43: Carbon Stock (t C)
CO2 = COL[43]    # 44: CO\u2082e (t)
MAI_PCT = COL[64]   # 65: MAI (%)
AAH_PCT = COL[65]   # 66: AAH (%)

PERHA_COLS = [COL[i] for i in range(44, 60)]   # 45-60: per-ha letters (16 cols: 11 vol + 5 carbon)
MAI_COLS  = [COL[i] for i in range(66, 82)]    # 67-82: MAI per-ha letters (16 cols)
AAH_COLS  = [COL[i] for i in range(82, 98)]    # 83-98: AAH per-ha letters (16 cols)

HEADERS = [
    "Block Name", "Sample Plot Number", "Stand Type",
    "X Coordinate", "Y Coordinate",
    "Species (Scientific)", "Species (Local)",
    "DBH (cm)", "Height (m)", "Height Estimated",
    "Tree Class (Quality)", "Count", "Serial Number (SN)",
    "Total Sample Plots",
    "Coeff_a", "Coeff_b", "Coeff_c", "Coeff_a1", "Coeff_b1",
    "Coeff_s", "Coeff_m", "Coeff_bg",
    "Wood Density (g/cm\u00b3)", "Growth Rate",
    "Full Stem Merchantable (0/1)",
    "Stem Volume (m\u00b3)", "Branchwood Ratio", "Branchwood Volume (m\u00b3)",
    "Tree Volume (m\u00b3)", "10-cm Top Diameter Ratio", "10-cm Top Volume (m\u00b3)",
    "Gross Timber Volume (m\u00b3)", "Recovery Factor (%)",
    "Net Timber Volume (m\u00b3)", "Net Timber Volume (cft)",
    "Fuelwood Volume (m\u00b3)", "Firewood (chatta)",
    "Basal Area (m\u00b2)", "DBH Class",
    "AGB (t)", "BGB (t)", "Total Biomass (t)",
    "Carbon Stock (t C)", "CO\u2082e (t)",
    "Count (Individuals)/ha",
    "Stem Volume (m\u00b3)/ha", "Branchwood Volume (m\u00b3)/ha",
    "Tree Volume (m\u00b3)/ha", "10-cm Top Volume (m\u00b3)/ha",
    "Gross Timber Volume (m\u00b3)/ha",
    "Net Timber Volume (m\u00b3)/ha", "Net Timber Volume (cft)/ha",
    "Fuelwood Volume (m\u00b3)/ha", "Firewood (chatta)/ha",
    "Basal Area (m\u00b2)/ha",
    "AGB (t)/ha", "BGB (t)/ha", "Total Biomass (t)/ha",
    "Carbon Stock (t C)/ha", "CO\u2082e (t)/ha",
    "Regen Plot (sqm)", "Sapling Plot (sqm)", "Pole Plot (sqm)", "Tree Plot (sqm)",
    "MAI (%)", "AAH (%)",
    "MAI_Count (Individuals)/ha",
    "MAI_Stem Volume (m\u00b3)/ha", "MAI_Branchwood Volume (m\u00b3)/ha",
    "MAI_Tree Volume (m\u00b3)/ha", "MAI_10-cm Top Volume (m\u00b3)/ha",
    "MAI_Gross Timber Volume (m\u00b3)/ha",
    "MAI_Net Timber Volume (m\u00b3)/ha", "MAI_Net Timber Volume (cft)/ha",
    "MAI_Fuelwood Volume (m\u00b3)/ha", "MAI_Firewood (chatta)/ha",
    "MAI_Basal Area (m\u00b2)/ha",
    "MAI_AGB (t)/ha", "MAI_BGB (t)/ha", "MAI_Total Biomass (t)/ha",
    "MAI_Carbon Stock (t C)/ha", "MAI_CO\u2082e (t)/ha",
    "AAH_Count (Individuals)/ha",
    "AAH_Stem Volume (m\u00b3)/ha", "AAH_Branchwood Volume (m\u00b3)/ha",
    "AAH_Tree Volume (m\u00b3)/ha", "AAH_10-cm Top Volume (m\u00b3)/ha",
    "AAH_Gross Timber Volume (m\u00b3)/ha",
    "AAH_Net Timber Volume (m\u00b3)/ha", "AAH_Net Timber Volume (cft)/ha",
    "AAH_Fuelwood Volume (m\u00b3)/ha", "AAH_Firewood (chatta)/ha",
    "AAH_Basal Area (m\u00b2)/ha",
    "AAH_AGB (t)/ha", "AAH_BGB (t)/ha", "AAH_Total Biomass (t)/ha",
    "AAH_Carbon Stock (t C)/ha", "AAH_CO\u2082e (t)/ha",
    "Species Code", "Species Regulation",
]

NUM_COLS = len(HEADERS)

DESCRIPTIONS = [
    (1, "Block Name", "Static", "Block name from uploaded data."),
    (2, "Sample Plot Number", "Static", "Sequential plot number within the block."),
    (3, "Stand Type", "Static", "Growth stage: Regeneration, Sapling, Pole, Tree."),
    (4, "X Coordinate", "Static", "Longitude/X from GPS."),
    (5, "Y Coordinate", "Static", "Latitude/Y from GPS."),
    (6, "Species (Scientific)", "Static", "Scientific name. Lookup key for coefficients."),
    (7, "Species (Local)", "Static", "Local/Nepali name."),
    (8, "DBH (cm)", "Static", "Diameter at Breast Height at 1.3m."),
    (9, "Height (m)", "Static", "Total tree height."),
    (10, "Height Estimated", "Static", "Yes=estimated, No=measured."),
    (11, "Tree Class (Quality)", "Static", "Quality: 1/i/a=80%, 2/ii/b=60%, 3/iii/c=30%, 4/iv/d=0%."),
    (12, "Count", "Static", "Number of individuals this record represents."),
    (13, "Serial Number (SN)", "Static", "Serial number within plot (auto-generated if missing)."),
    (14, "Total Sample Plots", "Static", "Total unique sample plots in this block."),
    (15, "Coeff_a", "Static", "Stem volume coefficient a. V = exp(a+b*ln(DBH)+c*ln(Ht))/1000."),
    (16, "Coeff_b", "Static", "Stem volume coefficient b."),
    (17, "Coeff_c", "Static", "Stem volume coefficient c."),
    (18, "Coeff_a1", "Static", "Bark deduction coefficient a1."),
    (19, "Coeff_b1", "Static", "Bark deduction coefficient b1."),
    (20, "Coeff_s", "Static", "Branch ratio for DBH<10cm."),
    (21, "Coeff_m", "Static", "Branch ratio for 10<=DBH<=40cm."),
    (22, "Coeff_bg", "Static", "Branch ratio for DBH>70cm."),
    (23, "Wood Density (g/cm3)", "Static", "Species wood density for biomass/carbon (IPCC)."),
    (24, "Growth Rate", "Static", "Growth rate classification (Fast/Moderate/Slow)."),
    (25, "Full Stem Merchantable (0/1)", "Static", "Flag: 1 = entire stem is merchantable. Khair (Acacia catechu) qualifies per Regulation 2079."),
    (26, "Stem Volume (m3)", "Formula", "Per-tree stem volume. =IF(DBH>=10&Ht>0, EXP(a+b*LN(DBH)+c*LN(Ht))/1000, 0)."),
    (27, "Branchwood Ratio", "Formula", "Branch proportion interpolated by DBH class (s/m/bg)."),
    (28, "Branchwood Volume (m3)", "Formula", "= StemVol * BranchwoodRatio. 100% fuelwood by regulation."),
    (29, "Tree Volume (m3)", "Formula", "Total tree = StemVol + BranchwoodVol. Input for AGB calculation."),
    (30, "10-cm Top Diameter Ratio", "Formula", "= EXP(a1 + b1*LN(DBH)). Top-end deduction ratio."),
    (31, "10-cm Top Volume (m3)", "Formula", "= StemVol * TopDiaRatio. Non-merchantable top-end."),
    (32, "Gross Timber Volume (m3)", "Formula", "= IF(FSM=1, StemVol, StemVol-TopVol). For Khair, entire stem is merchantable."),
    (33, "Recovery Factor (%)", "Formula", "= IF(FSM=1, 100, <quality>). Class 1=80%, 2=60%, 3=30%, 4=0%."),
    (34, "Net Timber Volume (m3)", "Formula", "= GrossTimber * RF/100. This IS Growing Stock."),
    (35, "Net Timber Volume (cft)", "Formula", "= NetTimber * 35.3147."),
    (36, "Fuelwood Volume (m3)", "Formula", "= TreeVol - NetTimber. Residual fuelwood."),
    (37, "Firewood (chatta)", "Formula", "= Fuelwood_m3 / 9.486. 1 chatta = 500 cft stacked x 67% packing = 9.486 m3."),
    (38, "Basal Area (m2)", "Formula", "= PI() * (DBH/200)^2."),
    (39, "DBH Class", "Formula", "Seedling to 60+ V.Lg.Tree."),
    (40, "AGB (t)", "Formula", "Above-Ground Biomass. = GrossTimber * WoodDensity * 1.3 (BEF=1.3, IPCC 2006 GL Table 4.4). VOB = gross merchantable stem."),
    (41, "BGB (t)", "Formula", "Below-Ground Biomass (roots). = AGB * 0.24 (R/S=0.24, IPCC 2006 GL Table 4.4)."),
    (42, "Total Biomass (t)", "Formula", "= AGB + BGB. Total tree biomass (above + below ground)."),
    (43, "Carbon Stock (t C)", "Formula", "= TotalBiomass * 0.47 (IPCC carbon fraction for tropical forest)."),
    (44, "CO\u2082e (t)", "Formula", "= CarbonStock * 3.67 (44/12, CO\u2082/C molecular ratio). CO\u2082 equivalent sequestered."),
    (45, "Count (Individuals)/ha", "Formula", "= Count / TotalPlots * PlotFactor."),
    (46, "Stem Volume (m3)/ha", "Formula", "= StemVol / TotalPlots * PlotFactor."),
    (47, "Branchwood Volume (m3)/ha", "Formula", "= BranchwoodVol / TotalPlots * PlotFactor."),
    (48, "Tree Volume (m3)/ha", "Formula", "= TreeVol / TotalPlots * PlotFactor."),
    (49, "10-cm Top Volume (m3)/ha", "Formula", "= TopVolume / TotalPlots * PlotFactor."),
    (50, "Gross Timber Volume (m3)/ha", "Formula", "= GrossTimber / TotalPlots * PlotFactor."),
    (51, "Net Timber Volume (m3)/ha", "Formula", "= NetTimber / TotalPlots * PlotFactor. Growing Stock per hectare."),
    (52, "Net Timber Volume (cft)/ha", "Formula", "= NetTimber_cft / TotalPlots * PlotFactor."),
    (53, "Fuelwood Volume (m3)/ha", "Formula", "= Fuelwood_m3 / TotalPlots * PlotFactor."),
    (54, "Firewood (chatta)/ha", "Formula", "= Firewood_chatta / TotalPlots * PlotFactor."),
    (55, "Basal Area (m2)/ha", "Formula", "= BasalArea / TotalPlots * PlotFactor."),
    (56, "AGB (t)/ha", "Formula", "= AGB / TotalPlots * PlotFactor. Above-ground biomass per hectare."),
    (57, "BGB (t)/ha", "Formula", "= BGB / TotalPlots * PlotFactor. Below-ground biomass per hectare."),
    (58, "Total Biomass (t)/ha", "Formula", "= TotalBiomass / TotalPlots * PlotFactor. Total tree biomass per hectare."),
    (59, "Carbon Stock (t C)/ha", "Formula", "= CarbonStock / TotalPlots * PlotFactor. Total carbon stock per hectare."),
    (60, "CO\u2082e (t)/ha", "Formula", "= CO2e / TotalPlots * PlotFactor. CO\u2082 equivalent per hectare."),
    (61, "Regen Plot (sqm)", "Static", "Plot area in sqm."),
    (62, "Sapling Plot (sqm)", "Static", "Plot area in sqm."),
    (63, "Pole Plot (sqm)", "Static", "Plot area in sqm."),
    (64, "Tree Plot (sqm)", "Static", "Plot area in sqm."),
    (65, "MAI (%)", "Static", "Mean Annual Increment percent. MAI = Growing Stock x MAI%/100 (m3/ha/yr)."),
    (66, "AAH (%)", "Static", "Annual Allowable Harvest percent (Good=75, Moderate=60, Weak=40)."),
    (67, "MAI_Count (Individuals)/ha", "Formula", "= Count/ha * MAI%/100."),
    (68, "MAI_Stem Volume (m3)/ha", "Formula", "= StemVol/ha * MAI%/100."),
    (69, "MAI_Branchwood Volume (m3)/ha", "Formula", "= BranchwoodVol/ha * MAI%/100."),
    (70, "MAI_Tree Volume (m3)/ha", "Formula", "= TreeVol/ha * MAI%/100."),
    (71, "MAI_10-cm Top Volume (m3)/ha", "Formula", "= TopVolume/ha * MAI%/100."),
    (72, "MAI_Gross Timber Volume (m3)/ha", "Formula", "= GrossTimber/ha * MAI%/100."),
    (73, "MAI_Net Timber Volume (m3)/ha", "Formula", "= NetTimber/ha * MAI%/100. MAI of Growing Stock."),
    (74, "MAI_Net Timber Volume (cft)/ha", "Formula", "= NetTimber_cft/ha * MAI%/100."),
    (75, "MAI_Fuelwood Volume (m3)/ha", "Formula", "= Fuelwood_m3/ha * MAI%/100."),
    (76, "MAI_Firewood (chatta)/ha", "Formula", "= Firewood_chatta/ha * MAI%/100."),
    (77, "MAI_Basal Area (m2)/ha", "Formula", "= BasalArea/ha * MAI%/100."),
    (78, "MAI_AGB (t)/ha", "Formula", "= AGB/ha * MAI%/100. MAI of above-ground biomass."),
    (79, "MAI_BGB (t)/ha", "Formula", "= BGB/ha * MAI%/100. MAI of below-ground biomass."),
    (80, "MAI_Total Biomass (t)/ha", "Formula", "= TotalBiomass/ha * MAI%/100."),
    (81, "MAI_Carbon Stock (t C)/ha", "Formula", "= CarbonStock/ha * MAI%/100. MAI of carbon stock."),
    (82, "MAI_CO\u2082e (t)/ha", "Formula", "= CO2e/ha * MAI%/100. MAI of CO\u2082 equivalent."),
    (83, "AAH_Count (Individuals)/ha", "Formula", "= MAI_Count/ha * AAH%/100."),
    (84, "AAH_Stem Volume (m3)/ha", "Formula", "= MAI_StemVol/ha * AAH%/100."),
    (85, "AAH_Branchwood Volume (m3)/ha", "Formula", "= MAI_BranchwoodVol/ha * AAH%/100."),
    (86, "AAH_Tree Volume (m3)/ha", "Formula", "= MAI_TreeVol/ha * AAH%/100."),
    (87, "AAH_10-cm Top Volume (m3)/ha", "Formula", "= MAI_TopVolume/ha * AAH%/100."),
    (88, "AAH_Gross Timber Volume (m3)/ha", "Formula", "= MAI_GrossTimber/ha * AAH%/100."),
    (89, "AAH_Net Timber Volume (m3)/ha", "Formula", "= MAI_NetTimber/ha * AAH%/100. AAH of Growing Stock."),
    (90, "AAH_Net Timber Volume (cft)/ha", "Formula", "= MAI_NetTimber_cft/ha * AAH%/100."),
    (91, "AAH_Fuelwood Volume (m3)/ha", "Formula", "= MAI_Fuelwood_m3/ha * AAH%/100."),
    (92, "AAH_Firewood (chatta)/ha", "Formula", "= MAI_Firewood_chatta/ha * AAH%/100."),
    (93, "AAH_Basal Area (m2)/ha", "Formula", "= MAI_BasalArea/ha * AAH%/100."),
    (94, "AAH_AGB (t)/ha", "Formula", "= MAI_AGB/ha * AAH%/100. AAH of above-ground biomass."),
    (95, "AAH_BGB (t)/ha", "Formula", "= MAI_BGB/ha * AAH%/100. AAH of below-ground biomass."),
    (96, "AAH_Total Biomass (t)/ha", "Formula", "= MAI_TotalBiomass/ha * AAH%/100."),
    (97, "AAH_Carbon Stock (t C)/ha", "Formula", "= MAI_CarbonStock/ha * AAH%/100. AAH of carbon stock."),
    (98, "AAH_CO\u2082e (t)/ha", "Formula", "= MAI_CO2e/ha * AAH%/100. AAH of CO\u2082 equivalent."),
    (99, "Species Code", "Static", "Regulation species code per Forest Regulation 2079. 1-21 = named species, 22 = Terai spp, 23 = Hill spp."),
    (100, "Species Regulation", "Static", "Regulation species name: scientific name for codes 1-21, 'Terai Spp' for code 22, 'Hill Spp' for code 23."),
]


def _bd(v):
    return float(v) if v is not None else None


def _safe(v, precision=4):
    if v is None:
        return None
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return 0.0
    return round(v, precision) if isinstance(v, float) else v


def _tree_class_to_letter(val) -> str:
    """Convert tree_class from any format to lowercase letter: a, b, c, d"""
    if not val:
        return ''
    s = str(val).strip().lower().replace('.0', '')
    return {'1': 'a', 'i': 'a', 'a': 'a',
            '2': 'b', 'ii': 'b', 'b': 'b',
            '3': 'c', 'iii': 'c', 'c': 'c',
            '4': 'd', 'iv': 'd', 'd': 'd'}.get(s, s)


def generate_field_inventory_excel(
    db: Session,
    field_inventory_id: UUID,
    aah_good: float = 75.0,
    aah_moderate: float = 60.0,
    aah_weak: float = 40.0,
    custom_multipliers: Optional[Dict[str, float]] = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment as XLComment
    from openpyxl.utils import get_column_letter

    rows = db.execute(text("""
        SELECT
            sp.block_name, sp.sample_plot_number,
            ST_X(sp.location::geometry) as x_coord,
            ST_Y(sp.location::geometry) as y_coord,
            m.stand_type, m.species_scientific, m.species_local,
            m.dbh_cm, m.height_m, m.height_estimated, m.tree_class,
            m.count,
            COALESCE(m.sn, ROW_NUMBER() OVER (
                PARTITION BY sp.block_name, sp.sample_plot_number, m.stand_type
                ORDER BY m.id
            )) as sn,
            fi.regeneration_area_sqm, fi.sapling_area_sqm,
            fi.pole_area_sqm, fi.tree_area_sqm
        FROM public.field_inventory_measurements m
        JOIN public.field_inventory_sample_plots sp
            ON m.sample_plot_id = sp.id
        JOIN public.field_inventory_calculations fi
            ON sp.field_inventory_calculation_id = fi.id
        WHERE fi.id = :fid
        ORDER BY sp.block_name, sp.sample_plot_number, m.stand_type, m.id
    """), {"fid": str(field_inventory_id)}).fetchall()

    if not rows:
        raise ValueError("No measurements found")

    sp_rows = db.execute(text("""
        SELECT scientific_name, species_code, a, b, c, a1, b1, s, m, bg,
               growth_rate, wood_density_gm_cm3, full_stem_merchantable
        FROM public.tree_species_coefficients WHERE is_active = TRUE
    """)).fetchall()

    sp_map: Dict[str, dict] = {}
    for sr in sp_rows:
        def f(v):
            return float(v) if v is not None else None
        sc = int(sr.species_code) if sr.species_code is not None else 23
        if sc <= 21:
            reg = str(sr.scientific_name)
        elif sc == 22:
            reg = "Terai Spp"
        else:
            reg = "Hill Spp"
        sp_map[str(sr.scientific_name)] = {
            'a': f(sr.a), 'b': f(sr.b), 'c': f(sr.c),
            'a1': f(sr.a1), 'b1': f(sr.b1),
            's': f(sr.s), 'm': f(sr.m), 'bg': f(sr.bg),
            'gr': str(sr.growth_rate or ''),
            'wd': f(sr.wood_density_gm_cm3),
            'fsm': bool(sr.full_stem_merchantable) if sr.full_stem_merchantable is not None else False,
            'species_code': sc,
            'species_regulation': reg,
        }

    block_summaries = db.execute(text("""
        SELECT block_name, mai_percent, forest_condition
        FROM public.field_inventory_block_summary
        WHERE field_inventory_calculation_id = :fid
    """), {"fid": str(field_inventory_id)}).fetchall()

    block_mai_aah: Dict[str, dict] = {}
    for bs in block_summaries:
        fc = str(bs.forest_condition or 'Moderate').strip().lower()
        if fc == 'good':
            aah_pct = aah_good
        elif fc == 'weak':
            aah_pct = aah_weak
        else:
            aah_pct = aah_moderate
        bn = str(bs.block_name).strip()
        if custom_multipliers and bn in custom_multipliers:
            aah_pct = custom_multipliers[bn]
        block_mai_aah[bn] = {
            'mai': float(bs.mai_percent) if bs.mai_percent is not None else 0.0,
            'aah': aah_pct,
        }

    block_plot_counts: Dict[str, int] = {}
    for r in rows:
        bn = str(r.block_name).strip()
        if bn not in block_plot_counts:
            plots = set()
            for r2 in rows:
                if str(r2.block_name).strip() == bn:
                    plots.add(r2.sample_plot_number)
            block_plot_counts[bn] = len(plots)

    hf = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hfont = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    df = Font(name='Calibri', size=10)
    entry_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    coeff_fill = PatternFill(start_color='E8DAEF', end_color='E8DAEF', fill_type='solid')
    coeff_font = Font(name='Calibri', size=9, color='6C3483', italic=True)
    formula_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
    formula_font = Font(name='Calibri', size=10, color='1A6B1A')
    perha_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    perha_font = Font(name='Calibri', size=10, color='004080')
    meta_font = Font(name='Calibri', size=10, color='666666')
    meta_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    thin = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hf
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 36

    # ── Excel comments with definitions on headers ──
    header_comments = {
        25: "Full Stem Merchantable flag. 1 = entire stem is merchantable (Khair).",
        26: "Stem Volume. Per-tree stem volume. Regulation 2079 formula.",
        29: "Tree Volume = Stem + Branchwood. Complete above-ground tree.",
        30: "= exp(a1 + b1*ln(DBH)). Fraction above 10cm top diameter.",
        31: "= StemVol x TopDiaRatio. Non-merchantable top-end.",
        32: "= IF(FSM=1, StemVol, StemVol-TopVol). Gross merchantable.",
        33: "= IF(FSM=1, 100, <quality>). Class 1=80%, 2=60%, 3=30%, 4=0%.",
        34: "【DFO REPORT — TIMBER】Net Timber = GrossTimber x RF/100. This IS GROWING STOCK — the standing merchantable volume. Use this column for DFO timber reporting and MAI/AAH calculations.",
        35: "Net Timber in cubic feet (x 35.3147).",
        36: "【DFO REPORT — FIREWOOD】Fuelwood Volume = TreeVol - NetTimber. All residual wood (branchwood + non-merchantable stem). Use this column for DFO firewood reporting.",
        37: "Firewood in Nepali chatta units. 1 chatta = 9.486 m3 (500 cft x 67% packing).",
        38: "Basal Area = pi x (DBH/200)^2.",
        40: "AGB = GrossTimber x WoodDensity x 1.3 (BEF=1.3, IPCC 2006 GL Table 4.4). VOB = gross merchantable stem volume.",
        41: "BGB = AGB x 0.24 (R/S=0.24, IPCC 2006 GL Table 4.4).",
        42: "Total Biomass = AGB + BGB.",
        43: "【DFO REPORT — CARBON】Carbon Stock = TotalBiomass x 0.47. Use for carbon reporting.",
        44: "CO2e = CarbonStock x 3.67. Use for carbon offset reporting.",
        51: "【DFO REPORT — TIMBER per ha】Net Timber /ha. GROWING STOCK per hectare. Primary metric for forest valuation and MAI/AAH.",
        53: "【DFO REPORT — FIREWOOD per ha】Fuelwood /ha. Use for per-hectare firewood reporting.",
        59: "【DFO REPORT — CARBON per ha】Carbon Stock (t C)/ha. Use for per-hectare carbon reporting.",
        65: "MAI (%) = Growing Stock x MAI%/100. Yearly increment (m3/ha/yr).",
        66: "AAH (%) = MAI x AAH%/100. Good=75%, Moderate=60%, Weak=40%.",
        73: "【DFO — MAI Timber】MAI_Net Timber = NetTimber/ha x MAI%/100. MAI of Growing Stock (m3/ha/yr).",
        75: "【DFO — MAI Firewood】MAI_Fuelwood = Fuelwood/ha x MAI%/100.",
        81: "【DFO — MAI Carbon】MAI_Carbon Stock = C/ha x MAI%/100.",
        89: "【DFO — AAH Timber】AAH_Net Timber = MAI_NetTimber/ha x AAH%/100. AAH of Growing Stock (m3/ha/yr).",
        91: "【DFO — AAH Firewood】AAH_Fuelwood = MAI_Fuelwood/ha x AAH%/100.",
        97: "【DFO — AAH Carbon】AAH_Carbon Stock = MAI_CStock/ha x AAH%/100.",
    }
    for col_idx, def_text in header_comments.items():
        cell = ws.cell(row=1, column=col_idx)
        cell.comment = XLComment(def_text, "System", width=400, height=150)

    # ── Highlight reporting column headers (DFO reporting guide) ──
    timber_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    firewood_fill = PatternFill(start_color='FCE4CC', end_color='FCE4CC', fill_type='solid')
    carbon_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    reporting_highlight = {
        34: timber_fill,    # Net Timber Volume — DFO Timber
        36: firewood_fill,  # Fuelwood Volume — DFO Firewood
        37: firewood_fill,  # Firewood chatta
        43: carbon_fill,    # Carbon Stock
        44: carbon_fill,    # CO2e
        51: timber_fill,    # Net Timber /ha
        53: firewood_fill,  # Fuelwood /ha
        54: firewood_fill,  # Firewood chatta /ha
        59: carbon_fill,    # Carbon Stock /ha
        60: carbon_fill,    # CO2e /ha
        73: timber_fill,    # MAI Net Timber
        75: firewood_fill,  # MAI Fuelwood
        81: carbon_fill,    # MAI Carbon
        89: timber_fill,    # AAH Net Timber
        91: firewood_fill,  # AAH Fuelwood
        97: carbon_fill,    # AAH Carbon
    }
    report_font = Font(name='Calibri', bold=True, size=9, color='000000')
    for col_idx, fill in reporting_highlight.items():
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = report_font

    # ── Shared plot factor formula fragment ──
    def plot_factor(dr):
        return (
            f'IF({C_}{dr}="Regeneration",1000,'
            f'IF({C_}{dr}="Sapling",400,'
            f'IF({C_}{dr}="Pole",100,20)))'
        )

    for ri, r in enumerate(rows):
        dr = ri + 2
        sci_name = str(r.species_scientific or '').strip()
        coef = sp_map.get(sci_name, {})
        stand_type = str(r.stand_type or '')
        dbh = _bd(r.dbh_cm)
        height = _bd(r.height_m)
        block_name = str(r.block_name).strip()

        a = coef.get('a')
        b = coef.get('b')
        c = coef.get('c')
        a1 = coef.get('a1')
        b1 = coef.get('b1')
        s = coef.get('s')
        m_v = coef.get('m')
        bg = coef.get('bg')
        wd = coef.get('wd')
        gr = coef.get('gr', '')
        fsm = 1 if coef.get('fsm') else 0

        # ── 1-14: Entry values (static) ──
        entry_vals = [
            block_name,
            int(r.sample_plot_number) if r.sample_plot_number is not None else None,
            stand_type,
            _safe(r.x_coord, 6),
            _safe(r.y_coord, 6),
            sci_name,
            str(r.species_local or ''),
            _safe(dbh, 2),
            _safe(height, 2),
            'Yes' if r.height_estimated else 'No',
            _tree_class_to_letter(r.tree_class),
            int(r.count) if r.count is not None else 1,
            int(r.sn) if r.sn is not None else 0,
            block_plot_counts.get(block_name, 0),
        ]
        for ci, v in enumerate(entry_vals, 1):
            cell = ws.cell(row=dr, column=ci, value=v)
            cell.font = df
            cell.fill = entry_fill
            cell.border = thin
            if isinstance(v, (int, float)) and v is not None:
                cell.alignment = Alignment(horizontal='right')

        # ── 15-25: Coefficients + Full Stem Merchantable (static) ──
        coeff_vals = [
            _safe(a, 6), _safe(b, 6), _safe(c, 6),
            _safe(a1, 6), _safe(b1, 6),
            _safe(s, 6), _safe(m_v, 6), _safe(bg, 6),
            _safe(wd, 4),
            gr if gr else None,
            fsm,
        ]
        for ci, v in enumerate(coeff_vals, 15):
            cell = ws.cell(row=dr, column=ci, value=v)
            cell.font = coeff_font
            cell.fill = coeff_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 26-44: Volume + Carbon / computed (FORMULAS) ──
        is_vol = (
            f'AND({H_}{dr}>=10,{I_}{dr}>0,'
            f'NOT({C_}{dr}="Regeneration"),'
            f'NOT({C_}{dr}="Sapling"))'
        )

        # Col 26: Stem Volume
        ws.cell(row=dr, column=26).value = (
            f'=IF({is_vol},'
            f'EXP({CA}{dr}+{CB}{dr}*LN({H_}{dr})+{CC}{dr}*LN({I_}{dr}))/1000,0)'
        )
        # Col 27: Branchwood Ratio
        ws.cell(row=dr, column=27).value = (
            f'=IF(OR({C_}{dr}="Regeneration",{C_}{dr}="Sapling"),0,'
            f'IF({H_}{dr}<10,{CS}{dr},'
            f'IF({H_}{dr}<=40,(({H_}{dr}-10)*{CM}{dr}+(40-{H_}{dr})*{CS}{dr})/30,'
            f'IF({H_}{dr}<=70,(({H_}{dr}-40)*{CBG}{dr}+(70-{H_}{dr})*{CM}{dr})/30,{CBG}{dr})))'
            f')'
        )
        # Col 28: Branchwood Volume
        ws.cell(row=dr, column=28).value = f'={SV}{dr}*{BR}{dr}'
        # Col 29: Tree Volume
        ws.cell(row=dr, column=29).value = f'={SV}{dr}+{BV}{dr}'
        # Col 30: 10-cm Top Diameter Ratio
        ws.cell(row=dr, column=30).value = (
            f'=IF({is_vol},EXP({CA1}{dr}+{CB1}{dr}*LN({H_}{dr})),0)'
        )
        # Col 31: 10-cm Top Volume
        ws.cell(row=dr, column=31).value = f'=IF({is_vol},{SV}{dr}*{TDR}{dr},0)'
        # Col 32: Gross Timber Volume — conditional: if FSM=1, StemVol, else StemVol-TopVol
        ws.cell(row=dr, column=32).value = (
            f'=IF({is_vol},IF({FSM}{dr}=1,{SV}{dr},{SV}{dr}-{TVOL}{dr}),0)'
        )
        # Col 33: Recovery Factor % — quality class lookup (applies to ALL species)
        ws.cell(row=dr, column=33).value = (
            f'=IF({K_}{dr}="",60,'
            f'IF(OR({K_}{dr}="1",{K_}{dr}="i",{K_}{dr}="a"),80,'
            f'IF(OR({K_}{dr}="2",{K_}{dr}="ii",{K_}{dr}="b"),60,'
            f'IF(OR({K_}{dr}="3",{K_}{dr}="iii",{K_}{dr}="c"),30,'
            f'IF(OR({K_}{dr}="4",{K_}{dr}="iv",{K_}{dr}="d"),0,60)))))'
        )
        # Col 34: Net Timber Volume
        ws.cell(row=dr, column=34).value = f'={GT}{dr}*{RF}{dr}/100'
        # Col 35: Net Timber Volume (cft)
        ws.cell(row=dr, column=35).value = f'={NT}{dr}*35.3147'
        # Col 36: Fuelwood Volume (m3) = TreeVol - NetTimber
        ws.cell(row=dr, column=36).value = f'={TV}{dr}-{NT}{dr}'
        # Col 37: Firewood (chatta) = Fuelwood_m3 / 9.486
        ws.cell(row=dr, column=37).value = f'={FW}{dr}/9.486'
        # Col 38: Basal Area
        ws.cell(row=dr, column=38).value = f'=PI()*({H_}{dr}/200)^2'
        # Col 39: DBH Class
        ws.cell(row=dr, column=39).value = (
            f'=IF({H_}{dr}="","",IF({H_}{dr}<4,"0-4 Seedling",'
            f'IF({H_}{dr}<10,"4-10 Sapling",IF({H_}{dr}<20,"10-20 Sm.Pole",'
            f'IF({H_}{dr}<30,"20-30 Lg.Pole",IF({H_}{dr}<40,"30-40 Sm.Tree",'
            f'IF({H_}{dr}<50,"40-50 Med.Tree",IF({H_}{dr}<60,"50-60 Lg.Tree",'
            f'"60+ V.Lg.Tree"))))))))'
        )
        # ── 40-44: Carbon (FORMULAS, IPCC Tier 2) ──
        # AGB = GrossTimber * WoodDensity * 1.3 (BEF, IPCC 2006 Table 4.4)
        ws.cell(row=dr, column=40).value = (
            f'=IF({is_vol},{GT}{dr}*{WD}{dr}*1.3,0)'
        )
        # BGB = AGB * 0.24 (R/S ratio, IPCC 2006 Table 4.4)
        ws.cell(row=dr, column=41).value = (
            f'=IF({is_vol},{AGB}{dr}*0.24,0)'
        )
        # Total Biomass = AGB + BGB
        ws.cell(row=dr, column=42).value = (
            f'=IF({is_vol},{AGB}{dr}+{BGB}{dr},0)'
        )
        # Carbon Stock = TotalBiomass * 0.47 (IPCC carbon fraction)
        ws.cell(row=dr, column=43).value = (
            f'=IF({is_vol},{TB}{dr}*0.47,0)'
        )
        # CO2e = CarbonStock * 3.67 (44/12)
        ws.cell(row=dr, column=44).value = (
            f'=IF({is_vol},{CARB}{dr}*3.67,0)'
        )

        for ci in range(26, 45):
            cell = ws.cell(row=dr, column=ci)
            cell.font = formula_font
            cell.fill = formula_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 45-60: Per-hectare (FORMULAS) = Value / TotalPlots * PlotFactor ──
        pf = plot_factor(dr)
        ha_cols = [
            (45, COL[11]),  # Count /ha
            (46, SV),   # Stem Vol /ha
            (47, BV),   # Branchwood Vol /ha
            (48, TV),   # Tree Vol /ha
            (49, TVOL), # 10-cm Top Vol /ha
            (50, GT),   # Gross Timber Vol /ha
            (51, NT),   # Net Timber Vol /ha
            (52, NTCFT),  # Net Timber Vol cft /ha
            (53, FW),   # Fuelwood Vol m3 /ha
            (54, FWC),  # Firewood chatta /ha
            (55, BA),   # Basal Area /ha
            (56, AGB),  # AGB /ha
            (57, BGB),  # BGB /ha
            (58, TB),   # Total Biomass /ha
            (59, CARB), # Carbon Stock /ha
            (60, CO2),  # CO2e /ha
        ]
        for col_num, val_col in ha_cols:
            ws.cell(row=dr, column=col_num).value = (
                f'=IF({N_}{dr}>0,{val_col}{dr}/{N_}{dr}*{pf},0)'
            )
            cell = ws.cell(row=dr, column=col_num)
            cell.font = perha_font
            cell.fill = perha_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 67-82: MAI per-ha (FORMULAS) = per-ha * MAI% / 100 ──
        for i in range(16):
            ws.cell(row=dr, column=67 + i).value = (
                f'={PERHA_COLS[i]}{dr}*{MAI_PCT}{dr}/100'
            )
            cell = ws.cell(row=dr, column=67 + i)
            cell.font = formula_font
            cell.fill = formula_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 83-98: AAH per-ha (FORMULAS) = MAI per-ha * AAH% / 100 ──
        for i in range(16):
            ws.cell(row=dr, column=83 + i).value = (
                f'={MAI_COLS[i]}{dr}*{AAH_PCT}{dr}/100'
            )
            cell = ws.cell(row=dr, column=83 + i)
            cell.font = formula_font
            cell.fill = formula_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 61-64: Plot sizes (static) ──
        def fv(v):
            return float(v) if v is not None else 0.0
        for ci, val in [
            (61, round(fv(r.regeneration_area_sqm), 1)),
            (62, round(fv(r.sapling_area_sqm), 1)),
            (63, round(fv(r.pole_area_sqm), 1)),
            (64, round(fv(r.tree_area_sqm), 1)),
        ]:
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.font = meta_font
            cell.fill = meta_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 65-66: MAI%, AAH% (static per-block) ──
        blk_mai_aah = block_mai_aah.get(block_name, {'mai': 0.0, 'aah': 75.0})
        for ci, val in [
            (65, blk_mai_aah['mai']),
            (66, blk_mai_aah['aah']),
        ]:
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.font = meta_font
            cell.fill = meta_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right')

        # ── 99-100: Species Code & Species Regulation (static) ──
        for ci, val in [
            (99, coef.get('species_code')),
            (100, coef.get('species_regulation')),
        ]:
            cell = ws.cell(row=dr, column=ci, value=val)
            cell.font = df
            cell.fill = coeff_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='right' if ci == 99 else 'left')

    for ci in range(1, NUM_COLS + 1):
        w = min(max(len(HEADERS[ci - 1]) + 4, 12), 28)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = ws.cell(row=2, column=1)
    ws.auto_filter.ref = f"A1:{get_column_letter(NUM_COLS)}{1 + len(rows)}"

    # ══════════════════════════════════
    #  DESCRIPTIONS SHEET
    # ══════════════════════════════════
    ws_desc = wb.create_sheet("Descriptions", 1)

    dhf = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    dhfont = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    d_title = Font(name='Calibri', bold=True, size=14, color='2F5496')
    d_sub = Font(name='Calibri', size=10, color='666666')
    d_font = Font(name='Calibri', size=10)
    d_bold = Font(name='Calibri', bold=True, size=10)
    d_static_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    d_formula_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

    ws_desc.cell(row=1, column=1, value="Field Inventory Excel Export \u2014 Column Descriptions").font = d_title
    ws_desc.merge_cells("A1:D1")
    ws_desc.cell(row=2, column=1, value=f"Generated: {datetime.now():%Y-%m-%d %H:%M}  |  Cols 1-24 = coeffs  |  25 = FSM  |  26-39 = volume  |  40-44 = carbon  |  45-60 = per-ha  |  61-66 = meta  |  67-82 = MAI  |  83-98 = AAH").font = d_sub
    ws_desc.merge_cells("A2:D2")

    for ci, h in enumerate(["#", "Column Name", "Type", "Description"], 1):
        cell = ws_desc.cell(row=4, column=ci, value=h)
        cell.font = dhfont
        cell.fill = dhf
        cell.alignment = Alignment(horizontal='center')

    for ri, (num, name, vtype, desc) in enumerate(DESCRIPTIONS, 5):
        ws_desc.cell(row=ri, column=1, value=num).font = d_font
        ws_desc.cell(row=ri, column=2, value=name).font = d_bold
        ws_desc.cell(row=ri, column=3, value=vtype).font = d_font
        desc_val = str(desc or '')
        if desc_val.startswith(('=', '+', '-')):
            desc_val = "'" + desc_val
        c4 = ws_desc.cell(row=ri, column=4, value=desc_val)
        c4.font = d_font
        c4.alignment = Alignment(wrap_text=True)
        fill = d_static_fill if vtype == "Static" else d_formula_fill
        for c in range(1, 5):
            ws_desc.cell(row=ri, column=c).fill = fill

    # ── IPCC Methodology Reference ──
    ref_row = ri + 2
    ws_desc.cell(row=ref_row, column=1, value="IPCC Methodology Reference").font = Font(name='Calibri', bold=True, size=12, color='2F5496')
    ws_desc.merge_cells(f"A{ref_row}:D{ref_row}")
    refs = [
        ("Parameter", "Value", "Source", "Notes"),
        ("VOB (Volume of growing stock)", "Gross merchantable stem volume (gross_volume = stem_vol + branch_vol)", "IPCC 2006 GL Vol 4 Ch 4 §2.2", "Excludes 10-cm top; NOT net_volume (quality-reduced)"),
        ("BEF (Biomass Expansion Factor)", "1.3", "IPCC 2006 GL Vol 4 Table 4.4", "Tropical moist deciduous forest; converts VOB to AGB"),
        ("R/S (Root-to-Shoot Ratio)", "0.24", "IPCC 2006 GL Vol 4 Table 4.4", "Tropical moist forest; AGB > 50 t/ha"),
        ("CF (Carbon Fraction)", "0.47", "IPCC 2006 GL Vol 4 Table 4.3", "47% of biomass is carbon; tropical forest default"),
        ("CO\u2082/C ratio", "3.67", "Molecular weight (44/12)", "CO\u2082 equivalent = C stock \u00d7 3.67"),
        ("Wood Density", "Species-specific (g/cm\u00b3)", "Global Wood Density DB / local data", "Per-tree gross-volume-weighted sum across all species"),
        ("AGB formula", "AGB = VOB \u00d7 WD \u00d7 BEF", "IPCC 2006 GL Vol 4 Eq 2.2.1", "Above-ground biomass in t/ha"),
        ("BGB formula", "BGB = AGB \u00d7 R/S", "IPCC 2006 GL Vol 4 Eq 2.2.2", "Below-ground (root) biomass in t/ha"),
    ]
    ref_fill_header = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for ci, h in enumerate(refs[0], 1):
        c = ws_desc.cell(row=ref_row + 1, column=ci, value=h)
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = ref_fill_header
    for i, (param, val, src, notes) in enumerate(refs[1:], ref_row + 2):
        ws_desc.cell(row=i, column=1, value=param).font = d_bold
        ws_desc.cell(row=i, column=2, value=val).font = d_font
        ws_desc.cell(row=i, column=3, value=src).font = d_font
        ws_desc.cell(row=i, column=4, value=notes).font = d_font
        for c in range(1, 5):
            ws_desc.cell(row=i, column=c).alignment = Alignment(wrap_text=True)

    # ── Forest Condition Assessment Algorithm ──
    cond_row = i + 2
    ws_desc.cell(row=cond_row, column=1, value="Regeneration & Forest Condition Algorithm").font = Font(name='Calibri', bold=True, size=12, color='2F5496')
    ws_desc.merge_cells(f"A{cond_row}:D{cond_row}")

    cond_notes = [
        ("Assessment Criteria", "Source: Nepal Forest Regulation 2075/2079"),
        ("", ""),
        ("A) Regeneration Condition (पुनरोत्पादनको अवस्था)", "AND logic: both thresholds must be met"),
        ("  राम्रो (Good)", "Regen (0-4 cm) >= 5000/ha AND Sapling (4-10 cm) >= 2000/ha"),
        ("  मध्यम (Moderate)", "Regen >= 2000/ha AND Sapling >= 800/ha"),
        ("  कमजोर (Weak)", "All other cases"),
        ("", ""),
        ("B) Forest Condition (वनको अवस्था)", "3x3 matrix: Growing Stock + Regeneration Condition"),
        ("  GS > 200 m3/ha + Good regen", "=> राम्रो (Good)"),
        ("  GS > 200 m3/ha + Moderate regen", "=> राम्रो (Good)"),
        ("  GS > 200 m3/ha + Weak regen", "=> मध्यम (Moderate)"),
        ("  GS 50-200 m3/ha + Good regen", "=> राम्रो (Good)"),
        ("  GS 50-200 m3/ha + Moderate regen", "=> मध्यम (Moderate)"),
        ("  GS 50-200 m3/ha + Weak regen", "=> कमजोर (Weak)"),
        ("  GS < 50 m3/ha + Good regen", "=> मध्यम (Moderate)"),
        ("  GS < 50 m3/ha + Moderate regen", "=> कमजोर (Weak)"),
        ("  GS < 50 m3/ha + Weak regen", "=> कमजोर (Weak)"),
    ]
    for ri, (label, desc) in enumerate(cond_notes, cond_row + 1):
        ws_desc.cell(row=ri, column=1, value=label).font = d_bold if label and not label.startswith(' ') else d_font
        ws_desc.cell(row=ri, column=2, value=desc).font = d_font
        for c in range(1, 5):
            ws_desc.cell(row=ri, column=c).alignment = Alignment(wrap_text=True)

    ws_desc.column_dimensions['A'].width = 50
    ws_desc.column_dimensions['B'].width = 65
    ws_desc.column_dimensions['C'].width = 40
    ws_desc.column_dimensions['D'].width = 60

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
