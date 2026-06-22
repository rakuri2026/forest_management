"""
Synthetic Tree Distribution Model Generator

Generates individual tree points (GPKG) from canopy height raster data
combined with species proportions and forestry standards for Nepal.

Algorithm Version: v1.0_prototype
Author: Community Forest Management System
Date: February 18, 2026
"""

import logging
import random
import math
import os
import uuid
from typing import List, Dict, Tuple, Optional, Any
from datetime import datetime
from pathlib import Path

from sqlalchemy.orm import Session
from sqlalchemy import text
from geoalchemy2.shape import to_shape
from shapely.geometry import Point, Polygon, MultiPolygon, box
from shapely.ops import unary_union
import numpy as np
import geopandas as gpd
import pandas as pd

from ..models.calculation import Calculation
from ..models.synthetic_tree_model import SyntheticTreeModel
from ..models.sampling import SamplingDesign
from .volume_calculator import calculate_tree_volumes

# Debug logging for volume calculation comparison
DEBUG_VOLUME_CALC = os.environ.get('DEBUG_VOLUME_CALC', 'false').lower() == 'true'

def _debug_log(msg: str):
    """Print debug message if DEBUG_VOLUME_CALC is enabled"""
    if DEBUG_VOLUME_CALC:
        print(f"[TREE_MODEL_VOLUME] {msg}")
from ..core.config import settings


# Configuration Constants (Nepal-specific standards)
MIN_DBH_CM = 10.0          # Commercial inventory threshold
MIN_HEIGHT_M = 5.0         # Minimum tree height
MAX_TREES_PER_HA = 500     # Upper cap on density
CANOPY_HEIGHT_PIXEL_SIZE = 5   # meters (5m resolution canopy height)
PIXEL_AREA_HA = 0.0025     # 25m² = 0.0025 hectares (no longer used in generation, kept for reference)


# Tree Density Lookup Tables (trees per hectare by canopy height)
TREE_DENSITY_DEFAULT = {
    (0, 2): 0,           # Below 2m = no commercial trees
    (2, 5): 0,           # Below 5m = excluded (regeneration/poles)
    (5, 10): 600,        # Young timber
    (10, 15): 450,       # Maturing timber
    (15, 20): 320,       # Mature timber
    (20, 25): 250,       # Old growth
    (25, 30): 200,       # Very tall canopy
    (30, 100): 150,      # Exceptional heights
}

# Forest type-specific densities (override default)
TREE_DENSITY_BY_FOREST_TYPE = {
    'Tropical Mixed Deciduous Forest': {
        (0, 5): 0,
        (5, 10): 550,
        (10, 15): 420,
        (15, 20): 300,
        (20, 30): 230,
        (30, 100): 180,
    },
    'Lower Temperate Pine Forest': {
        (0, 5): 0,
        (5, 10): 300,
        (10, 15): 240,
        (15, 20): 200,
        (20, 30): 160,
        (30, 100): 130,
    },
    'Lower Temperate Mixed Broadleaved Forest': {
        (0, 5): 0,
        (5, 10): 500,
        (10, 15): 400,
        (15, 20): 300,
        (20, 30): 240,
        (30, 100): 180,
    },
    'Sal Forest': {
        (0, 5): 0,
        (5, 10): 450,
        (10, 15): 350,
        (15, 20): 280,
        (20, 30): 220,
        (30, 100): 170,
    },
}


def get_species_regulation(species_code: Any, scientific_name: str) -> str:
    code = int(species_code) if species_code and str(species_code).strip() else 0
    if code == 22:
        return "Terai spp"
    elif code == 23:
        return "Hill spp"
    return scientific_name


def get_tree_density(canopy_height: float, forest_type: str, max_density: int = MAX_TREES_PER_HA) -> int:
    """
    Get trees per hectare for given canopy height and forest type

    Args:
        canopy_height: Canopy height in meters
        forest_type: Forest type classification string
        max_density: Maximum trees per hectare cap (user-configurable)

    Returns:
        Trees per hectare (integer)
    """
    # Select appropriate lookup table
    lookup = TREE_DENSITY_BY_FOREST_TYPE.get(forest_type, TREE_DENSITY_DEFAULT)

    # Find matching height range
    for (min_h, max_h), density in lookup.items():
        if min_h <= canopy_height < max_h:
            return min(density, max_density)  # Use user's max_density, not hardcoded constant

    # Fallback for heights outside defined ranges
    fallback = 150 if canopy_height >= 5 else 0
    return min(fallback, max_density)


def generate_random_point_in_pixel(pixel_bounds: Tuple[float, float, float, float]) -> Tuple[float, float]:
    """
    Generate random point within pixel boundaries

    Args:
        pixel_bounds: (min_x, min_y, max_x, max_y) in EPSG:4326

    Returns:
        (longitude, latitude) tuple
    """
    min_x, min_y, max_x, max_y = pixel_bounds
    x = random.uniform(min_x, max_x)
    y = random.uniform(min_y, max_y)
    return (x, y)


def weighted_random_choice(species_list: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Select species based on weighted probabilities from availability_rank

    Args:
        species_list: List of species dictionaries with 'availability_rank'

    Returns:
        Selected species dictionary
    """
    if not species_list:
        raise ValueError("Species list is empty")

    # Calculate weights (higher rank = higher probability)
    # availability_rank: 1=dominant, 2=co-dominant, 3=associate, 4=occasional, 5=rare
    # Invert so dominant species get higher weight
    weights = []
    for species in species_list:
        rank = species.get('availability_rank', 3)
        # Weight formula: 6 - rank (so 1 becomes 5, 5 becomes 1)
        weight = max(1, 6 - rank)
        weights.append(weight)

    # Weighted random selection
    total_weight = sum(weights)
    r = random.uniform(0, total_weight)
    cumulative = 0
    for species, weight in zip(species_list, weights):
        cumulative += weight
        if r <= cumulative:
            return species

    # Fallback to last species
    return species_list[-1]


def assign_tree_height(canopy_height: float, species: Dict[str, Any], role: str) -> float:
    """
    Assign tree height based on canopy height, species, and vertical position

    Args:
        canopy_height: Pixel canopy height (meters)
        species: Species dictionary with max_height_m
        role: 'dominant', 'co-dominant', 'associate', 'occasional', 'rare'

    Returns:
        Tree height in meters
    """
    # Get species maximum height constraint
    max_species_height = species.get('max_height_m', 40)

    # Apply vertical stratification based on role
    if role == 'dominant':
        # Dominant species: 80-100% of canopy height
        height = canopy_height * random.uniform(0.80, 1.00)
    elif role == 'co-dominant':
        # Co-dominant: 60-85% of canopy height
        height = canopy_height * random.uniform(0.60, 0.85)
    elif role == 'associate':
        # Associate: 40-70% of canopy height
        height = canopy_height * random.uniform(0.40, 0.70)
    else:  # occasional/rare
        # Understory: 25-55% of canopy height
        height = canopy_height * random.uniform(0.25, 0.55)

    # Constrain to species maximum
    height = min(height, max_species_height)

    # Ensure meets minimum threshold
    height = max(height, MIN_HEIGHT_M)

    return round(height, 1)


def calculate_dbh_from_height(height: float, species: Dict[str, Any]) -> float:
    """
    Calculate DBH from height using species-specific relationships

    Priority order:
    1. H:D ratio from database (typical_hd_ratio)
    2. Allometric coefficients (a, b, c)
    3. Fallback: growth rate-based ratio

    Args:
        height: Tree height in meters
        species: Species dictionary with allometric data

    Returns:
        DBH in centimeters
    """
    # Option 1: Use H:D ratio if available
    hd_min = species.get('typical_hd_ratio_min')
    hd_max = species.get('typical_hd_ratio_max')

    if hd_min is not None and hd_max is not None and hd_min > 0 and hd_max > 0:
        hd_ratio = random.uniform(hd_min, hd_max)
        dbh = height / hd_ratio * 100  # Convert to cm

    # Option 2: Use allometric equation (if coefficients exist)
    elif species.get('a') is not None and species.get('b') is not None:
        a = species['a']
        b = species['b']
        c = species.get('c', 1.0)  # Default c=1 if not specified

        # Inverse of Height = a + b*DBH^c
        # Simplified for c=1: DBH = (Height - a) / b
        if c == 1.0 and b != 0:
            dbh = (height - a) / b
        else:
            # For other c values, use approximation
            dbh = height / 35  # Fallback

    # Option 3: Fallback - use growth rate-based ratio
    else:
        growth_rate = species.get('growth_rate', 'Moderate')

        if growth_rate == 'Fast':
            hd_ratio = random.uniform(0.25, 0.35)  # Fast growers: relatively thicker
        elif growth_rate == 'Moderate':
            hd_ratio = random.uniform(0.30, 0.40)
        else:  # Slow
            hd_ratio = random.uniform(0.35, 0.50)  # Slow growers: relatively thinner

        dbh = height / hd_ratio

    # Constrain to species maximum
    max_dbh = species.get('max_dbh_cm', 200)
    dbh = min(dbh, max_dbh)

    # Ensure meets minimum threshold
    dbh = max(dbh, MIN_DBH_CM)

    return round(dbh, 1)


# calculate_tree_volumes() is imported from .volume_calculator
# Single shared source of truth for all three tabs.


def calculate_tree_biomass(dbh: float, height: Optional[float], species: Dict[str, Any]) -> float:
    """
    Calculate above-ground biomass (AGB) for a single tree using allometric equations.

    Uses Nepal-appropriate allometric equations:
    1. Generic equation: AGB (kg) = 0.0673 × (DBH^2.7395)  [Chave et al. 2005]
    2. If height available: height adjustment for large trees

    Args:
        dbh: Diameter at breast height in centimeters
        height: Tree height in meters (optional)
        species: Species dictionary (for future species-specific coefficients)

    Returns:
        Above-ground biomass in megagrams (Mg) = metric tons
    """
    if dbh < 1.0:
        return 0.0  # Regeneration has negligible biomass

    # Calculate biomass in kilograms using Chave et al. (2005) equation
    # AGB (kg) = 0.0673 × (DBH^2.7395)
    agb_kg = 0.0673 * (dbh ** 2.7395)

    # Optional: Adjust for height if available and tree is large
    if height and height > 10 and dbh > 30:
        # Height adjustment factor (empirical for Nepal)
        height_factor = 1.0 + (height - 15) * 0.02  # +2% per meter above 15m
        height_factor = max(0.8, min(1.3, height_factor))  # Constrain to 0.8-1.3
        agb_kg *= height_factor

    # Convert kg to Mg (megagrams = metric tons)
    agb_mg = agb_kg / 1000.0

    return round(agb_mg, 4)  # 4 decimal places for precision


def assign_tree_class(dbh: float, height: float, species: Dict[str, Any]) -> int:
    """
    Assign tree class (1=25%, 2=50%, 3=75%, 4=100% firewood potential)

    Nepal standard:
    - Class 4 (100%): Small trees, low economic value
    - Class 1 (25%): Large trees, high economic value
    - Class 2-3: Mid-range

    Args:
        dbh: DBH in centimeters
        height: Height in meters
        species: Species dictionary with economic_value

    Returns:
        Tree class (1-4)
    """
    # Base class on tree size
    if dbh < 15 or height < 8:
        base_class = 4  # Small = mostly firewood
    elif dbh < 25 or height < 12:
        base_class = 3
    elif dbh < 40 or height < 16:
        base_class = 2
    else:
        base_class = 1  # Large = mostly timber

    # Adjust by economic value
    econ_value = species.get('economic_value', 'Moderate')

    if econ_value in ['High', 'Very High']:
        base_class = max(1, base_class - 1)  # Shift towards timber
    elif econ_value == 'Low':
        base_class = min(4, base_class + 1)  # Shift towards firewood

    return base_class


def get_species_role(availability_rank: int) -> str:
    """
    Convert availability_rank to role string

    Args:
        availability_rank: 1=dominant, 2=co-dominant, 3=associate, 4=occasional, 5=rare

    Returns:
        Role string
    """
    rank_to_role = {
        1: 'dominant',
        2: 'co-dominant',
        3: 'associate',
        4: 'occasional',
        5: 'rare'
    }
    return rank_to_role.get(availability_rank, 'associate')


def group_species_by_role(species_list: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Group species by their role based on availability_rank

    Returns:
        Dict with keys: 'dominant', 'co-dominant', 'associate', 'occasional', 'rare'
    """
    rank_to_role = {1: 'dominant', 2: 'co-dominant', 3: 'associate', 4: 'occasional', 5: 'rare'}
    groups = {role: [] for role in rank_to_role.values()}
    for sp in species_list:
        rank = sp.get('availability_rank', 3)
        role = rank_to_role.get(rank, 'associate')
        groups[role].append(sp)
    return groups


def weighted_random_choice_with_ratio(
    species_list: List[Dict[str, Any]],
    role_target_ratio: Optional[Dict[str, float]] = None
) -> Dict[str, Any]:
    """
    Select species with optional role-ratio override.

    Mode 1 — No ratio (default): Same as weighted_random_choice().
    Uses availability_rank across ALL species (current behavior).

    Mode 2 — Ratio provided:
    1. Pick role first using target ratio (roulette wheel)
    2. Within the role group, pick species using frequency-weighted ranks
       Formula: within_role_weight = (6 - rank) * (frequency_percent / 100 + 0.5)

    Args:
        species_list: List of species dicts with 'availability_rank', 'frequency_percent'
        role_target_ratio: e.g. {'dominant': 0.50, 'co-dominant': 0.30, ...}
                          or None for default behavior

    Returns:
        Selected species dictionary
    """
    if not species_list:
        raise ValueError("Species list is empty")

    # Mode 1: Default — use existing weighted_random_choice logic
    if role_target_ratio is None:
        return weighted_random_choice(species_list)

    # Mode 2: Role-ratio override
    # Step 1: Group species by role
    rank_to_role = {1: 'dominant', 2: 'co-dominant', 3: 'associate', 4: 'occasional', 5: 'rare'}
    role_to_ranks = {role: [r for r, rl in rank_to_role.items() if rl == role] for role in role_target_ratio}

    role_groups = {}
    for sp in species_list:
        rank = sp.get('availability_rank', 3)
        role = rank_to_role.get(rank, 'associate')
        if role not in role_groups:
            role_groups[role] = []
        role_groups[role].append(sp)

    # Step 2: Pick role by target ratio
    roles = list(role_target_ratio.keys())
    ratios = [role_target_ratio[r] for r in roles]
    total_ratio = sum(ratios)
    r = random.uniform(0, total_ratio)
    cumulative = 0.0
    selected_role = roles[-1]
    for role, ratio in zip(roles, ratios):
        cumulative += ratio
        if r <= cumulative:
            selected_role = role
            break

    # Step 3: Get species in selected role
    role_species = role_groups.get(selected_role, [])
    if not role_species:
        # Fallback: if no species found for selected role, pick from any
        role_species = species_list

    # Step 4: Within-role weighted selection
    weights = []
    for sp in role_species:
        rank = sp.get('availability_rank', 3)
        freq = sp.get('frequency_percent', 50) or 50
        weight = (6 - rank) * (freq / 100.0 + 0.5)
        weights.append(max(weight, 0.1))

    total_weight = sum(weights)
    r2 = random.uniform(0, total_weight)
    cumulative = 0.0
    for sp, weight in zip(role_species, weights):
        cumulative += weight
        if r2 <= cumulative:
            return sp

    return role_species[-1] if role_species else species_list[-1]


def extract_canopy_pixels(boundary_wkt: str, db: Session) -> List[Dict[str, Any]]:
    """
    Extract canopy height pixels within boundary from PostGIS raster.
    Uses ST_Clip to clip raster to polygon boundary BEFORE extracting pixels.

    PERFORMANCE FIX: Previous version processed 99.6% wasted pixels from bounding box extent.
    New version clips raster first, reducing processing by 100x for elongated polygons.

    Args:
        boundary_wkt: WKT string of boundary geometry (EPSG:4326)
        db: Database session

    Returns:
        List of pixel dictionaries with {height, bounds, center}
    """
    query = text("""
        WITH boundary AS (
            SELECT ST_GeomFromText(:boundary_wkt, 4326) AS geom
        ),
        clipped_rasters AS (
            -- CRITICAL FIX: Clip raster tiles to polygon boundary FIRST
            -- This eliminates pixels outside the polygon before extraction
            SELECT ST_Clip(rast, boundary.geom, true) AS clipped_rast
            FROM rasters.canopy_height, boundary
            WHERE ST_Intersects(rast, boundary.geom)
        ),
        pixels AS (
            -- Extract pixels only from clipped rasters (already within boundary)
            SELECT (ST_PixelAsCentroids(clipped_rast, 1)).*
            FROM clipped_rasters
            WHERE clipped_rast IS NOT NULL
        )
        SELECT
            x, y, val AS height,
            ST_XMin(geom) AS min_x,
            ST_YMin(geom) AS min_y,
            ST_XMax(geom) AS max_x,
            ST_YMax(geom) AS max_y
        FROM pixels
        WHERE val IS NOT NULL
          AND val > 0
    """)

    result = db.execute(query, {"boundary_wkt": boundary_wkt})
    pixels = []

    for row in result:
        pixels.append({
            'height': float(row.height),
            'bounds': (row.min_x, row.min_y, row.max_x, row.max_y),
            'center': (row.x, row.y)
        })

    return pixels


def assign_block_names_to_trees(
    trees: List[Dict[str, Any]],
    result_data: Dict[str, Any]
) -> List[Dict[str, Any]]:
    """
    Assign correct block names to trees via spatial join (uses STRtree index).

    Args:
        trees: List of tree dictionaries with geometry (x, y)
        result_data: Calculation result_data with 'blocks' array

    Returns:
        Trees list with correct block_name assigned
    """
    from shapely import wkt

    blocks_data = result_data.get('blocks', [])
    if not blocks_data:
        return trees

    from shapely.geometry import shape
    from shapely import STRtree

    block_geoms = []
    block_names = []
    for idx, block in enumerate(blocks_data):
        block_geojson = block.get('geometry')
        if not block_geojson:
            continue
        try:
            block_geom = shape(block_geojson)
            block_name = block.get('block_name', f"Block_{idx+1}")
            block_geoms.append(block_geom)
            block_names.append(block_name)
        except Exception as e:
            print(f"Warning: Could not parse block {idx}: {e}")
            continue

    if not block_geoms:
        return trees

    tree_idx = STRtree(block_geoms)

    for tree in trees:
        tree_point = Point(tree['geometry'])
        indices = tree_idx.query(tree_point)
        for idx in indices:
            if block_geoms[int(idx)].contains(tree_point):
                tree['block_name'] = block_names[int(idx)]
                break

    return trees


def assign_sample_plots_to_trees(
    trees: List[Dict[str, Any]],
    sampling_design: 'SamplingDesign',
    buffer_meters: float,
    db: Session
) -> List[Dict[str, Any]]:
    """
    Assign sample plot numbers to trees using STRtree spatial index.

    Keeps only trees within at least one plot buffer (discards outside trees).

    Args:
        trees: List of tree dictionaries with geometry (x, y)
        sampling_design: SamplingDesign object with sample points
        buffer_meters: Buffer distance around each plot
        db: Database session

    Returns:
        FILTERED trees list with sample_plot_number assigned
    """
    from geoalchemy2.shape import to_shape
    from shapely import STRtree

    if not sampling_design.points_geometry:
        return trees

    sample_points_geom = to_shape(sampling_design.points_geometry)
    plot_assignments = sampling_design.points_block_assignment or []
    buffer_deg = buffer_meters / 111320.0

    plot_buffers = []
    plot_numbers = []

    if hasattr(sample_points_geom, 'geoms'):
        for idx, point in enumerate(sample_points_geom.geoms):
            plot_info = next(
                (p for p in plot_assignments if p.get('point_index') == idx),
                None
            )
            plot_number = plot_info.get('plot_number', idx + 1) if plot_info else idx + 1
            plot_buffers.append(point.buffer(buffer_deg))
            plot_numbers.append(plot_number)
    else:
        plot_buffers.append(sample_points_geom.buffer(buffer_deg))
        plot_numbers.append(1)

    buffer_tree = STRtree(plot_buffers)
    filtered_trees = []

    for tree in trees:
        tree_point = Point(tree['geometry'])
        indices = buffer_tree.query(tree_point)
        intersecting_plots = []
        for idx in indices:
            if plot_buffers[int(idx)].contains(tree_point):
                intersecting_plots.append(str(plot_numbers[int(idx)]))
        if intersecting_plots:
            tree['sample_plot_number'] = ','.join(intersecting_plots)
            filtered_trees.append(tree)

    return filtered_trees


def generate_regeneration_entries(
    sampling_design: 'SamplingDesign',
    species_list: List[Dict[str, Any]],
    buffer_meters: float,
    tree_id_start: int
) -> List[Dict[str, Any]]:
    """
    Generate regeneration entries (1-9.9 cm DBH) for each sample plot.

    Per sample plot:
    - Unestablished regeneration (1-3.9 cm DBH): 2-5 species
    - Established regeneration/sapling (4-9.9 cm DBH): 1-4 species

    Regeneration entries do NOT have height_m or tree_class (field verification needed).

    Args:
        sampling_design: SamplingDesign object with sample points
        species_list: List of species dictionaries
        buffer_meters: Buffer distance around each plot
        tree_id_start: Starting tree_id number

    Returns:
        List of regeneration tree dictionaries
    """
    from geoalchemy2.shape import to_shape

    if not sampling_design.points_geometry:
        return []

    # Get sample plot points
    sample_points_geom = to_shape(sampling_design.points_geometry)
    plot_assignments = sampling_design.points_block_assignment or []

    # Create list of plot locations
    plot_locations = []
    if hasattr(sample_points_geom, 'geoms'):
        # MultiPoint
        for idx, point in enumerate(sample_points_geom.geoms):
            plot_info = next(
                (p for p in plot_assignments if p.get('point_index') == idx),
                None
            )
            plot_number = plot_info.get('plot_number', idx + 1) if plot_info else idx + 1
            plot_locations.append({
                'plot_number': plot_number,
                'center': (point.x, point.y)
            })
    else:
        # Single Point
        plot_locations.append({
            'plot_number': 1,
            'center': (sample_points_geom.x, sample_points_geom.y)
        })

    # Generate regeneration entries
    regeneration_trees = []
    tree_id = tree_id_start

    for plot in plot_locations:
        plot_number = str(plot['plot_number'])
        center_x, center_y = plot['center']

        # 1. Unestablished Regeneration (1-3.9 cm DBH): 2-5 species
        num_unestablished = random.randint(2, 5)
        for _ in range(num_unestablished):
            species = weighted_random_choice(species_list)

            # Random position within plot buffer
            angle = random.uniform(0, 2 * 3.14159)
            distance = random.uniform(0, buffer_meters / 111320.0)  # Convert to degrees
            x = center_x + distance * math.cos(angle)
            y = center_y + distance * math.sin(angle)

            regeneration_trees.append({
                'tree_id': tree_id,
                'geometry': (x, y),
                'species_scientific': species.get('scientific_name'),
                'species_local': species.get('local_name'),
                'dbh_cm': round(random.uniform(1.0, 3.9), 1),
                'height_m': None,  # Not measured for regeneration
                'tree_class': None,  # Not applicable
                'block_name': '',  # Will be assigned via spatial join
                'sample_plot_number': plot_number,
                'generated_date': datetime.now().isoformat(),
                'model_version': 'v1.0',
                'notes': 'REGENERATION (1-3.9cm DBH) - Field verification required',
            })
            tree_id += 1

        # 2. Established Regeneration/Sapling (4-9.9 cm DBH): 1-4 species
        num_established = random.randint(1, 4)
        for _ in range(num_established):
            species = weighted_random_choice(species_list)

            # Random position within plot buffer
            angle = random.uniform(0, 2 * 3.14159)
            distance = random.uniform(0, buffer_meters / 111320.0)
            x = center_x + distance * math.cos(angle)
            y = center_y + distance * math.sin(angle)

            regeneration_trees.append({
                'tree_id': tree_id,
                'geometry': (x, y),
                'species_scientific': species.get('scientific_name'),
                'species_local': species.get('local_name'),
                'dbh_cm': round(random.uniform(4.0, 9.9), 1),
                'height_m': None,  # Not measured for regeneration
                'tree_class': None,  # Not applicable
                'block_name': '',  # Will be assigned via spatial join
                'sample_plot_number': plot_number,
                'generated_date': datetime.now().isoformat(),
                'model_version': 'v1.0',
                'notes': 'SAPLING (4-9.9cm DBH) - Field verification required',
            })
            tree_id += 1

    return regeneration_trees


def export_to_gpkg(
    trees: List[Dict[str, Any]],
    calculation_id: uuid.UUID,
    db: Session = None,
    output_dir: str = "exports"
) -> Tuple[str, float]:
    """
    Export trees to GPKG file using GeoPandas in REGULATION FORMAT

    Forest Regulation 2079 - Standard Format
    16 columns organized by size class with count columns

    Size Classes (based on DBH):
    - Regeneration: 1 cm <= DBH < 4 cm
    - Sapling: 4 cm <= DBH < 10 cm
    - Pole: 10 cm <= DBH < 30 cm
    - Tree: DBH >= 30 cm

    Each tree populates only its size class columns.
    Count columns default to 1 (one individual tree per row).

    Args:
        trees: List of tree dictionaries
        calculation_id: UUID of calculation
        db: Database session (optional, for forest name)
        output_dir: Directory to save GPKG files

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    # Create output directory if needed
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # Get forest name from calculation for filename
    forest_name = "forest"
    if db is not None:
        try:
            from app.models.calculation import Calculation
            calculation = db.query(Calculation).filter(Calculation.id == calculation_id).first()
            if calculation and calculation.forest_name:
                # Sanitize forest name for filename (remove special characters)
                forest_name = calculation.forest_name.replace(' ', '_')
                forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')
        except Exception as e:
            print(f"Warning: Could not retrieve forest name: {e}")

    # Generate filename with new format: forest_name_TreeModel_SyntheticTrees_date
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{forest_name}_TreeModel_SyntheticTrees_{timestamp}.gpkg"
    filepath = os.path.join(output_dir, filename)

    # Create list of records - REGULATION FORMAT (16 columns)
    records = []
    fid = 1

    for tree in trees:
        dbh = tree['dbh_cm']
        species_sci = tree.get('species_scientific', '')
        height = tree.get('height_m')
        tree_class = tree.get('tree_class')

        # Get volumes (Forest Regulation 2079)
        stem_vol = tree.get('stem_volume', 0.0)
        branch_vol = tree.get('branch_volume', 0.0)
        tree_vol = tree.get('tree_volume', 0.0)
        gross_vol = tree.get('gross_volume', 0.0)
        net_vol = tree.get('net_volume', 0.0)
        firewood_vol = tree.get('firewood_m3', 0.0)

        # Initialize all columns as None
        record = {
            'fid': fid,
            'block_name': tree.get('block_name', ''),
            'sample_plot_number': tree.get('sample_plot_number', ''),
            # Regeneration columns
            'regen_species_scientific': None,
            'regen_dbh': None,
            'regen_count': None,
            # Sapling columns
            'sapling_species_scientific': None,
            'sapling_dbh_cm': None,
            'sapling_count': None,
            # Pole columns
            'pole_species_scientific': None,
            'pole_dbh_cm': None,
            'pole_height_m': None,
            'pole_class': None,
            'pole_stem_volume_m3': None,
            'pole_branch_volume_m3': None,
            'pole_tree_volume_m3': None,
            'pole_gross_volume_m3': None,
            'pole_net_volume_m3': None,
            'pole_firewood_m3': None,
            # Tree columns
            'tree_species_scientific': None,
            'tree_dbh_cm': None,
            'tree_height_m': None,
            'tree_class': None,
            'tree_stem_volume_m3': None,
            'tree_branch_volume_m3': None,
            'tree_tree_volume_m3': None,
            'tree_gross_volume_m3': None,
            'tree_net_volume_m3': None,
            'tree_firewood_m3': None,
            # Geometry
            'geometry': Point(tree['geometry'])
        }

        # Populate appropriate columns based on DBH (size class)
        # FIXED: Correct DBH thresholds for regeneration (1-3.99) and sapling (4-9.99)
        if dbh < 4:
            # Regeneration (1-3.99 cm DBH - unestablished regeneration)
            record['regen_species_scientific'] = species_sci
            record['regen_dbh'] = dbh
            record['regen_count'] = 1
        elif dbh < 10:
            # Sapling (4-9.99 cm DBH - established regeneration)
            record['sapling_species_scientific'] = species_sci
            record['sapling_dbh_cm'] = dbh
            record['sapling_count'] = 1
        elif dbh < 20:
            # Small Pole (10-19.99 cm DBH - with volumes)
            record['pole_species_scientific'] = species_sci
            record['pole_dbh_cm'] = dbh
            record['pole_height_m'] = height
            record['pole_class'] = tree_class
            record['pole_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['pole_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['pole_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['pole_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['pole_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['pole_firewood_m3'] = firewood_vol if firewood_vol > 0 else None
        elif dbh < 30:
            # Large Pole (20-29.99 cm DBH - with volumes)
            record['pole_species_scientific'] = species_sci
            record['pole_dbh_cm'] = dbh
            record['pole_height_m'] = height
            record['pole_class'] = tree_class
            record['pole_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['pole_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['pole_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['pole_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['pole_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['pole_firewood_m3'] = firewood_vol if firewood_vol > 0 else None
        else:
            # Tree (>=30 cm DBH - with volumes)
            record['tree_species_scientific'] = species_sci
            record['tree_dbh_cm'] = dbh
            record['tree_height_m'] = height
            record['tree_class'] = tree_class
            record['tree_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['tree_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['tree_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['tree_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['tree_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['tree_firewood_m3'] = firewood_vol if firewood_vol > 0 else None

        records.append(record)
        fid += 1

    # Create GeoDataFrame with regulation column order
    gdf = gpd.GeoDataFrame(records, crs='EPSG:4326')

    # Ensure column order matches regulation format (with volume columns)
    column_order = [
        'fid',
        'block_name',
        'sample_plot_number',
        'regen_species_scientific',
        'regen_dbh',
        'regen_count',
        'sapling_species_scientific',
        'sapling_dbh_cm',
        'sapling_count',
        'pole_species_scientific',
        'pole_dbh_cm',
        'pole_height_m',
        'pole_class',
        'pole_stem_volume_m3',
        'pole_branch_volume_m3',
        'pole_tree_volume_m3',
        'pole_gross_volume_m3',
        'pole_net_volume_m3',
        'pole_firewood_m3',
        'tree_species_scientific',
        'tree_dbh_cm',
        'tree_height_m',
        'tree_class',
        'tree_stem_volume_m3',
        'tree_branch_volume_m3',
        'tree_tree_volume_m3',
        'tree_gross_volume_m3',
        'tree_net_volume_m3',
        'tree_firewood_m3',
        'geometry'
    ]
    gdf = gdf[column_order]

    # Write to GPKG
    gdf.to_file(filepath, driver='GPKG', layer='tree_model')

    # Calculate file size
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)

    return filepath, file_size_mb


def export_to_excel(
    trees: List[Dict[str, Any]],
    calculation_id: uuid.UUID,
    db: Session = None,
    output_dir: str = "exports"
) -> Tuple[str, float]:
    """
    Export trees to Excel file (.xlsx) in REGULATION FORMAT

    Same structure as GPKG export but in Excel format for field teams.
    Includes lat/lon columns for easy coordinate reference.

    Forest Regulation 2079 - Standard Format
    16 columns + latitude/longitude for convenience

    Args:
        trees: List of tree dictionaries
        calculation_id: UUID of calculation
        db: Database session (optional, for validation enhancement)
        output_dir: Directory to save Excel files

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    # Create output directory if needed
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # Get forest name from calculation for filename
    forest_name = "forest"
    if db is not None:
        try:
            from app.models.calculation import Calculation
            calculation = db.query(Calculation).filter(Calculation.id == calculation_id).first()
            if calculation and calculation.forest_name:
                # Sanitize forest name for filename (remove special characters)
                forest_name = calculation.forest_name.replace(' ', '_')
                forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')
        except Exception as e:
            print(f"Warning: Could not retrieve forest name: {e}")

    # Generate filename with new format: forest_name_TreeModel_SyntheticTrees_date
    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{forest_name}_TreeModel_SyntheticTrees_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Get sample plot center coordinates from sampling design
    plot_coordinates = {}
    if db is not None:
        try:
            sampling_design = db.query(SamplingDesign).filter(
                SamplingDesign.calculation_id == calculation_id
            ).first()
            
            if sampling_design and sampling_design.points_geometry:
                # Extract coordinates from MULTIPOINT geometry
                # Query uses ST_DumpPoints to get each point with its index
                query = text("""
                    SELECT 
                        (ST_DumpPoints(points_geometry)).path[1] - 1 as point_index,
                        ST_X((ST_DumpPoints(points_geometry)).geom) as lon,
                        ST_Y((ST_DumpPoints(points_geometry)).geom) as lat
                    FROM sampling_designs
                    WHERE id = :design_id
                """)
                result = db.execute(query, {"design_id": sampling_design.id})
                for row in result:
                    # point_index is 0-based, plot numbers are typically 1-based
                    plot_number = str(row[0] + 1)
                    plot_coordinates[plot_number] = (row[1], row[2])
                
                print(f"[Excel Export] Loaded {len(plot_coordinates)} plot coordinates from sampling design")
        except Exception as e:
            print(f"[Excel Export] Warning: Could not load plot coordinates: {e}")

    # Create list of records - REGULATION FORMAT (16 columns + lat/lon)
    records = []
    fid = 1

    for tree in trees:
        dbh = tree['dbh_cm']
        species_sci = tree.get('species_scientific', '')
        species_role = tree.get('species_role', 'associate')  # Get role for sorting
        height = tree.get('height_m')
        tree_class = tree.get('tree_class')
        
        # Get sample plot number for coordinate lookup
        sample_plot_num = str(tree.get('sample_plot_number', ''))
        
        # Use sample plot center coordinates instead of individual tree coordinates
        if sample_plot_num and sample_plot_num in plot_coordinates:
            lon, lat = plot_coordinates[sample_plot_num]
        else:
            # Fallback to tree's own coordinates if plot not found
            lon, lat = tree['geometry']

        # Get volumes (Forest Regulation 2079)
        stem_vol = tree.get('stem_volume', 0.0)
        branch_vol = tree.get('branch_volume', 0.0)
        tree_vol = tree.get('tree_volume', 0.0)
        gross_vol = tree.get('gross_volume', 0.0)
        net_vol = tree.get('net_volume', 0.0)
        firewood_vol = tree.get('firewood_m3', 0.0)

        # Initialize all columns
        record = {
            'fid': fid,
            'block_name': tree.get('block_name', ''),
            'block_number': tree.get('block_number', 0),
            'sample_plot_number': tree.get('sample_plot_number', ''),
            'total_sample_plots': tree.get('total_sample_plots', 0),
            # Hidden sorting column (removed before export)
            'species_role': species_role,
            # Regeneration columns
            'regen_species_scientific': None,
            'regen_dbh': None,
            'regen_count': None,
            # Sapling columns
            'sapling_species_scientific': None,
            'sapling_dbh_cm': None,
            'sapling_count': None,
            # Pole columns
            'pole_species_scientific': None,
            'pole_dbh_cm': None,
            'pole_height_m': None,
            'pole_class': None,
            'pole_stem_volume_m3': None,
            'pole_branch_volume_m3': None,
            'pole_tree_volume_m3': None,
            'pole_gross_volume_m3': None,
            'pole_net_volume_m3': None,
            'pole_firewood_m3': None,
            # Tree columns
            'tree_species_scientific': None,
            'tree_dbh_cm': None,
            'tree_height_m': None,
            'tree_class': None,
            'tree_stem_volume_m3': None,
            'tree_branch_volume_m3': None,
            'tree_tree_volume_m3': None,
            'tree_gross_volume_m3': None,
            'tree_net_volume_m3': None,
            'tree_firewood_m3': None,
            # Coordinates (for Excel convenience)
            'longitude': lon,
            'latitude': lat,
            # Additional resource columns (populated once per sample plot)
            'firewood_kg_per_100sqm_per_year': None,
            'grass_kg_per_100sqm_per_year': None,
            'bedding_material_kg_per_100sqm_per_year': None,
            'ntfp_kg_per_100sqm_per_year': None,
        }

        # Populate appropriate columns based on DBH (size class)
        # FIXED: Correct DBH thresholds for regeneration (1-3.99) and sapling (4-9.99)
        if dbh < 4:
            # Regeneration (1-3.99 cm DBH - unestablished regeneration)
            record['regen_species_scientific'] = species_sci
            record['regen_dbh'] = dbh
            record['regen_count'] = 1
        elif dbh < 10:
            # Sapling (4-9.99 cm DBH - established regeneration)
            record['sapling_species_scientific'] = species_sci
            record['sapling_dbh_cm'] = dbh
            record['sapling_count'] = 1
        elif dbh < 20:
            # Small Pole (10-19.99 cm DBH - with volumes)
            record['pole_species_scientific'] = species_sci
            record['pole_dbh_cm'] = dbh
            record['pole_height_m'] = height
            record['pole_class'] = tree_class
            record['pole_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['pole_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['pole_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['pole_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['pole_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['pole_firewood_m3'] = firewood_vol if firewood_vol > 0 else None
        elif dbh < 30:
            # Large Pole (20-29.99 cm DBH - with volumes)
            record['pole_species_scientific'] = species_sci
            record['pole_dbh_cm'] = dbh
            record['pole_height_m'] = height
            record['pole_class'] = tree_class
            record['pole_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['pole_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['pole_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['pole_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['pole_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['pole_firewood_m3'] = firewood_vol if firewood_vol > 0 else None
        else:
            # Tree (>=30 cm DBH - with volumes)
            record['tree_species_scientific'] = species_sci
            record['tree_dbh_cm'] = dbh
            record['tree_height_m'] = height
            record['tree_class'] = tree_class
            record['tree_stem_volume_m3'] = stem_vol if stem_vol > 0 else None
            record['tree_branch_volume_m3'] = branch_vol if branch_vol > 0 else None
            record['tree_tree_volume_m3'] = tree_vol if tree_vol > 0 else None
            record['tree_gross_volume_m3'] = gross_vol if gross_vol > 0 else None
            record['tree_net_volume_m3'] = net_vol if net_vol > 0 else None
            record['tree_firewood_m3'] = firewood_vol if firewood_vol > 0 else None

        records.append(record)
        fid += 1

    # Create DataFrame with regulation column order
    df = pd.DataFrame(records)

    # SORTING LOGIC - Sort by importance before export
    # Convert sample_plot_number to numeric for proper sorting (1, 2, 3... not 1, 10, 11, 2)
    df['sample_plot_number_numeric'] = pd.to_numeric(df['sample_plot_number'], errors='coerce').fillna(999999)

    # Add serial numbers (SN) for each category that reset per sample plot
    # Initialize SN columns
    df['regen_sn'] = None
    df['sapling_sn'] = None
    df['pole_sn'] = None
    df['tree_sn'] = None

    # Calculate serial numbers per plot per category
    for plot_num in df['sample_plot_number'].unique():
        plot_mask = df['sample_plot_number'] == plot_num

        # Regeneration SN (reset per plot)
        regen_mask = plot_mask & df['regen_species_scientific'].notna()
        if regen_mask.any():
            df.loc[regen_mask, 'regen_sn'] = range(1, regen_mask.sum() + 1)

        # Sapling SN (reset per plot)
        sapling_mask = plot_mask & df['sapling_species_scientific'].notna()
        if sapling_mask.any():
            df.loc[sapling_mask, 'sapling_sn'] = range(1, sapling_mask.sum() + 1)

        # Pole SN (reset per plot)
        pole_mask = plot_mask & df['pole_species_scientific'].notna()
        if pole_mask.any():
            df.loc[pole_mask, 'pole_sn'] = range(1, pole_mask.sum() + 1)

        # Tree SN (reset per plot)
        tree_mask = plot_mask & df['tree_species_scientific'].notna()
        if tree_mask.any():
            df.loc[tree_mask, 'tree_sn'] = range(1, tree_mask.sum() + 1)

    # Define species role priority (dominant first, rare last)
    role_priority = {
        'dominant': 1,
        'co-dominant': 2,
        'associate': 3,
        'occasional': 4,
        'rare': 5
    }
    df['role_priority'] = df['species_role'].map(role_priority).fillna(3)  # Default to associate

    # Get species columns for sorting (whichever is not None)
    df['species_for_sorting'] = (
        df['tree_species_scientific'].fillna('') +
        df['pole_species_scientific'].fillna('') +
        df['sapling_species_scientific'].fillna('') +
        df['regen_species_scientific'].fillna('')
    )

    # COMPACT FORMAT: Sort BEFORE merge by sample_plot_number, then regen_sn, sapling_sn, pole_sn, tree_sn
    # This groups rows so that after merge, they consolidate properly
    # Use numeric sorting to match Excel's "Sort anything that looks like a number, as a number"
    df['_s1'] = pd.to_numeric(df['sample_plot_number'], errors='coerce').fillna(999999)
    df['_s2'] = pd.to_numeric(df['regen_sn'], errors='coerce').fillna(999999)
    df['_s3'] = pd.to_numeric(df['sapling_sn'], errors='coerce').fillna(999999)
    df['_s4'] = pd.to_numeric(df['pole_sn'], errors='coerce').fillna(999999)
    df['_s5'] = pd.to_numeric(df['tree_sn'], errors='coerce').fillna(999999)
    
    df = df.sort_values(by=['_s1', '_s2', '_s3', '_s4', '_s5'])
    df = df.drop(columns=['_s1', '_s2', '_s3', '_s4', '_s5'])

    # Use min_sn as merge key to combine different size classes with same SN
    df['min_sn'] = df[['tree_sn', 'pole_sn', 'sapling_sn', 'regen_sn']].min(axis=1)
    df['merge_key'] = df['sample_plot_number'].astype(str) + '_' + df['min_sn'].astype(str)
    
    # Define aggregation functions - take first non-null value for each column
    agg_dict = {
        'fid': 'first',  # Will reassign later
        'block_name': 'first',
        'sample_plot_number': 'first',
        'block_number': 'first',
        'total_sample_plots': 'first',
        'longitude': 'first',
        'latitude': 'first',
        # Regeneration columns
        'regen_sn': 'first',
        'regen_species_scientific': 'first',
        'regen_dbh': 'first',
        'regen_count': 'first',
        # Sapling columns
        'sapling_sn': 'first',
        'sapling_species_scientific': 'first',
        'sapling_dbh_cm': 'first',
        'sapling_count': 'first',
        # Pole columns
        'pole_sn': 'first',
        'pole_species_scientific': 'first',
        'pole_dbh_cm': 'first',
        'pole_height_m': 'first',
        'pole_class': 'first',
        # Pole volumes
        'pole_stem_volume_m3': 'first',
        'pole_branch_volume_m3': 'first',
        'pole_tree_volume_m3': 'first',
        'pole_gross_volume_m3': 'first',
        'pole_net_volume_m3': 'first',
        'pole_firewood_m3': 'first',
        # Tree columns
        'tree_sn': 'first',
        'tree_species_scientific': 'first',
        'tree_dbh_cm': 'first',
        'tree_height_m': 'first',
        'tree_class': 'first',
        # Tree volumes
        'tree_stem_volume_m3': 'first',
        'tree_branch_volume_m3': 'first',
        'tree_tree_volume_m3': 'first',
        'tree_gross_volume_m3': 'first',
        'tree_net_volume_m3': 'first',
        'tree_firewood_m3': 'first',
        # Additional resource columns
        'firewood_kg_per_100sqm_per_year': 'first',
        'grass_kg_per_100sqm_per_year': 'first',
        'bedding_material_kg_per_100sqm_per_year': 'first',
        'ntfp_kg_per_100sqm_per_year': 'first',
    }
    
    # Group by merge_key and aggregate
    df = df.groupby('merge_key', as_index=False).agg(agg_dict)
    
    # FINAL SORT: After groupby, sort by sample_plot_number, regen_sn, sapling_sn, pole_sn, tree_sn numerically
    df['_s1'] = pd.to_numeric(df['sample_plot_number'], errors='coerce').fillna(999999)
    df['_s2'] = pd.to_numeric(df['regen_sn'], errors='coerce').fillna(999999)
    df['_s3'] = pd.to_numeric(df['sapling_sn'], errors='coerce').fillna(999999)
    df['_s4'] = pd.to_numeric(df['pole_sn'], errors='coerce').fillna(999999)
    df['_s5'] = pd.to_numeric(df['tree_sn'], errors='coerce').fillna(999999)
    df = df.sort_values(by=['_s1', '_s2', '_s3', '_s4', '_s5']).drop(columns=['_s1', '_s2', '_s3', '_s4', '_s5'])
    
    # Drop temporary columns (use errors='ignore' in case column doesn't exist)
    cols_to_drop = [col for col in ['merge_key', 'sample_plot_number_numeric'] if col in df.columns]
    df = df.drop(columns=cols_to_drop)

    # Reset index and reassign fid sequentially
    df = df.reset_index(drop=True)
    df['fid'] = range(1, len(df) + 1)

    # Column order for Tree Model sheet (original format - NO volume columns)
    tree_model_column_order = [
        'fid',
        'block_name',
        'sample_plot_number',
        'longitude',
        'latitude',
        'regen_sn',
        'regen_species_scientific',
        'regen_dbh',
        'regen_count',
        'sapling_sn',
        'sapling_species_scientific',
        'sapling_dbh_cm',
        'sapling_count',
        'pole_sn',
        'pole_species_scientific',
        'pole_dbh_cm',
        'pole_height_m',
        'pole_class',
        'tree_sn',
        'tree_species_scientific',
        'tree_dbh_cm',
        'tree_height_m',
        'tree_class',
        'firewood_kg_per_100sqm_per_year',
        'grass_kg_per_100sqm_per_year',
        'bedding_material_kg_per_100sqm_per_year',
        'ntfp_kg_per_100sqm_per_year',
    ]
    
    # REMOVE EMPTY ROWS (rows with no species data in any category)
    # A row is empty if all species columns are None/NaN
    has_species_data = (
        df['regen_species_scientific'].notna() |
        df['sapling_species_scientific'].notna() |
        df['pole_species_scientific'].notna() |
        df['tree_species_scientific'].notna()
    )
    df = df[has_species_data].reset_index(drop=True)

    # Reassign FID after removing empty rows
    df['fid'] = range(1, len(df) + 1)

    # Populate default values for additional resource columns (only first row per sample plot)
    # Track which sample plots have been populated
    populated_plots = set()
    for idx in df.index:
        sample_plot = df.loc[idx, 'sample_plot_number']
        if sample_plot not in populated_plots:
            # First row of this sample plot - populate default values
            df.loc[idx, 'firewood_kg_per_100sqm_per_year'] = 50
            df.loc[idx, 'grass_kg_per_100sqm_per_year'] = 50
            df.loc[idx, 'bedding_material_kg_per_100sqm_per_year'] = 50
            df.loc[idx, 'ntfp_kg_per_100sqm_per_year'] = 1
            populated_plots.add(sample_plot)

    # DBH columns - keep full precision for Field Inventory recalculation
    # Round DBH and Height to 6 decimal places (preserves precision for volume calculation)
    # DBH and Height columns - keep original precision from tree generation
    # Tree Model calculates these with full precision (DBH rounded to 1 decimal, height to 1 decimal)
    # Field Inventory will use pre-calculated volume columns from Excel instead of recalculating
    
    # Filter to only include columns that exist in df
    tree_model_column_order = [col for col in tree_model_column_order if col in df.columns]
    df_tree_model = df[tree_model_column_order]

    # Column order for Volumes sheet (includes block_number and total_sample_plots)
    volumes_column_order = [
        'fid',
        'block_name',
        'block_number',
        'sample_plot_number',
        'total_sample_plots',
        'pole_species_scientific',
        'pole_dbh_cm',
        'pole_height_m',
        'pole_class',
        'pole_stem_volume_m3',
        'pole_branch_volume_m3',
        'pole_tree_volume_m3',
        'pole_gross_volume_m3',
        'pole_net_volume_m3',
        'pole_firewood_m3',
        'tree_species_scientific',
        'tree_dbh_cm',
        'tree_height_m',
        'tree_class',
        'tree_stem_volume_m3',
        'tree_branch_volume_m3',
        'tree_tree_volume_m3',
        'tree_gross_volume_m3',
        'tree_net_volume_m3',
        'tree_firewood_m3',
    ]
    
    volumes_column_order = [col for col in volumes_column_order if col in df.columns]
    df_volumes = df[volumes_column_order]
    
    # Write to Excel - optimized for speed
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Sheet 1: Data Template (renamed from Tree Model)
        df_tree_model.to_excel(writer, sheet_name='Data Template', index=False)

        # Sheet 2: Volumes (for verification) - includes block_number and total_sample_plots
        df_volumes.to_excel(writer, sheet_name='Volumes', index=False)

    # Apply formatting to Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = load_workbook(filepath)

    # Format Data Template sheet
    ws_data = wb['Data Template']

    # 1. Format header row (dark background, white text)
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")  # Dark slate gray
    header_font = Font(color="FFFFFF", bold=True)  # White text, bold

    for cell in ws_data[1]:  # First row
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 2. Set column widths
    for col in ws_data.columns:
        col_letter = col[0].column_letter
        ws_data.column_dimensions[col_letter].width = 15

    # 3. Hide column A (fid column)
    ws_data.column_dimensions['A'].hidden = True

    # 4. Center justify column C (sample_plot_number)
    for row in ws_data.iter_rows(min_row=2, min_col=3, max_col=3):  # Column C, skip header
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5. Freeze first row and columns B and C (freeze pane at D2)
    ws_data.freeze_panes = 'D2'

    # 6. Hide the Volumes sheet
    wb['Volumes'].sheet_state = 'hidden'

    wb.save(filepath)

    # Calculate file size
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)

    # Add validation enhancement if database session provided
    if db is not None:
        try:
            from .excel_validator import add_validation_to_excel
            import logging
            logger = logging.getLogger(__name__)
            logger.info("Adding validation to Excel file...")
            filepath, file_size_mb = add_validation_to_excel(filepath, db)
            logger.info(f"Validation added successfully! New size: {file_size_mb:.2f} MB")
        except Exception as e:
            # Log error but don't fail the export
            import logging
            import traceback
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to add validation to Excel: {e}")
            logger.error(traceback.format_exc())
            print(f"ERROR: Could not add validation to Excel: {e}")
            traceback.print_exc()
            # Continue with basic Excel export
    else:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning("Database session not provided - skipping Excel validation enhancement")

    return filepath, file_size_mb


def generate_synthetic_trees(
    calculation_id: uuid.UUID,
    db: Session,
    config: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[callable] = None
) -> Dict[str, Any]:
    """
    Main algorithm: Generate synthetic tree distribution from canopy height raster

    Args:
        calculation_id: UUID of calculation with boundary
        db: Database session
        config: Optional configuration overrides
        progress_callback: Optional function(percent, step) to report progress

    Returns:
        Dictionary with generation results and statistics
    """
    start_time = datetime.now()

    # Default configuration
    default_config = {
        'min_dbh_cm': MIN_DBH_CM,
        'min_height_m': MIN_HEIGHT_M,
        'max_trees_per_ha': MAX_TREES_PER_HA,
        'spatial_distribution': 'random',
        'algorithm_version': 'v1.0',
        'plot_buffer_meters': 12.62  # Default buffer for sample plot assignment (radius for 500m² plot)
    }
    config = {**default_config, **(config or {})}

    # Report progress
    def report(percent: int, step: str):
        if progress_callback:
            progress_callback(percent, step)

    # Step 1: Load calculation and data
    report(10, "Loading calculation data")
    calculation = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    if not calculation:
        raise ValueError(f"Calculation {calculation_id} not found")

    # Get species list from result_data
    result_data = calculation.result_data or {}
    species_list = result_data.get('potential_species', [])
    if not species_list:
        raise ValueError("No species data found in calculation")

    # FILTER: Only use tree species (exclude herbs, shrubs, non-woody plants)
    # Tree species have wood/timber/fuel/fodder uses
    tree_species_only = [
        sp for sp in species_list
        if sp.get('is_tree_species', True)  # Default TRUE for backward compatibility
    ]

    if not tree_species_only:
        raise ValueError(
            f"No tree species found in calculation. "
            f"Found {len(species_list)} total species, but none are classified as trees. "
            f"Tree species must have wood, timber, fuel, or fodder uses."
        )

    # Log species filtering for transparency
    if len(tree_species_only) < len(species_list):
        non_tree_count = len(species_list) - len(tree_species_only)
        print(f"INFO: Filtered out {non_tree_count} non-tree species (herbs, shrubs). "
              f"Using {len(tree_species_only)} tree species for model generation.")

    # Use filtered tree species for tree model generation
    species_list = tree_species_only

    forest_type = result_data.get('forest_type_dominant', 'Unknown')
    area_hectares = result_data.get('area_hectares', 0)

    # Step 1.2: Load species coefficients for volume calculations
    report(11, "Loading species coefficients")
    species_coefficients = {}

    # Load species coefficients from database
    from sqlalchemy import text as sql_text
    species_code_map = {}
    coef_query = sql_text("""
        SELECT scientific_name, species_code, a, b, c, s, m, bg, a1, b1, full_stem_merchantable
        FROM tree_species_coefficients
    """)
    coef_result = db.execute(coef_query)
    for row in coef_result:
        species_code_map[row.scientific_name] = row.species_code
        species_coefficients[row.scientific_name] = {
            'a': row.a,
            'b': row.b,
            'c': row.c,
            's': row.s,
            'm': row.m,
            'bg': row.bg,
            'a1': row.a1,
            'b1': row.b1,
            'full_stem_merchantable': bool(row.full_stem_merchantable) if row.full_stem_merchantable is not None else False,
        }

    # Step 1.5: Check if sampling design exists (REQUIRED)
    report(12, "Checking for sampling design")
    sampling_design = db.query(SamplingDesign).filter(
        SamplingDesign.calculation_id == calculation_id
    ).first()

    if not sampling_design or not sampling_design.points_geometry:
        raise ValueError(
            "Sample plots are required before generating tree distribution. "
            "Please create a sampling design first from the Sampling tab."
        )

    # Step 1.6: Extract plot centers and plot numbers from sampling design
    from geoalchemy2.shape import to_shape

    report(15, "Extracting sample plot centers")
    sample_points_geom = to_shape(sampling_design.points_geometry)
    plot_assignments = sampling_design.points_block_assignment or []
    buffer_deg = config['plot_buffer_meters'] / 111320.0
    plot_area_ha = (3.14159 * (config['plot_buffer_meters'] ** 2)) / 10000.0

    plot_centers = []
    if hasattr(sample_points_geom, 'geoms'):
        for idx, point in enumerate(sample_points_geom.geoms):
            plot_info = next((p for p in plot_assignments if p.get('point_index') == idx), None)
            plot_number = plot_info.get('plot_number', idx + 1) if plot_info else idx + 1
            plot_centers.append({
                'plot_number': str(plot_number),
                'center': (point.x, point.y),
                'buffer': point.buffer(buffer_deg)
            })
    else:
        plot_centers.append({
            'plot_number': '1',
            'center': (sample_points_geom.x, sample_points_geom.y),
            'buffer': sample_points_geom.buffer(buffer_deg)
        })

    # Step 2: Sample canopy height at each plot center from local TIFF
    report(20, "Sampling canopy height from raster")
    import rasterio
    from rasterio.sample import sample_gen

    tiff_path = r"D:\forest_management\canopy_height5m.tif"
    coords = [(pc['center'][0], pc['center'][1]) for pc in plot_centers]

    with rasterio.open(tiff_path) as src:
        values = list(sample_gen(src, coords))

    active_plots = []
    for pc, val in zip(plot_centers, values):
        height = float(val[0]) if val and val[0] is not None else 0
        if height > 0:
            pc['height'] = height
            active_plots.append(pc)

    if not active_plots:
        raise ValueError("No canopy height data found at sample plot locations")

    # Step 3: Generate trees per plot (within plot buffer area)
    report(30, f"Generating trees for {len(active_plots)} plots")
    trees = []
    tree_id = 1

    for idx, pc in enumerate(active_plots):
        canopy_height = pc['height']
        trees_per_ha = get_tree_density(canopy_height, forest_type, config['max_trees_per_ha'])
        if trees_per_ha == 0:
            continue

        num_trees = int(trees_per_ha * plot_area_ha)
        if num_trees == 0:
            continue

        for _ in range(num_trees):
            angle = random.uniform(0, 2 * 3.14159)
            dist = random.uniform(0, buffer_deg)
            x = pc['center'][0] + dist * math.cos(angle)
            y = pc['center'][1] + dist * math.sin(angle)

            # Select species
            species = weighted_random_choice(species_list)
            role = get_species_role(species.get('availability_rank', 3))

            # Assign height
            tree_height = assign_tree_height(canopy_height, species, role)

            # Calculate DBH
            dbh = calculate_dbh_from_height(tree_height, species)

            # Filter by thresholds
            if dbh < config['min_dbh_cm'] or tree_height < config['min_height_m']:
                continue  # Skip trees below threshold

            # Assign tree class
            tree_class = assign_tree_class(dbh, tree_height, species)

            # Calculate volumes using Forest Regulation 2079 formulas
            volumes = {'stem_volume': 0.0, 'branch_volume': 0.0, 'tree_volume': 0.0,
                      'gross_volume': 0.0, 'net_volume': 0.0, 'firewood_m3': 0.0}

            scientific_name = species.get('scientific_name')
            if scientific_name and scientific_name in species_coefficients:
                volumes = calculate_tree_volumes(
                    dbh=dbh,
                    height=tree_height,
                    tree_class=tree_class,
                    species_coefficients=species_coefficients[scientific_name]
                )

            # Create tree record (with pre-assigned plot number)
            trees.append({
                'tree_id': tree_id,
                'geometry': (x, y),
                'species_code': species_code_map.get(scientific_name, ''),
                'species_regulation': get_species_regulation(species_code_map.get(scientific_name, ''), scientific_name or ''),
                'species_scientific': scientific_name,
                'species_local': species.get('local_name'),
                'species_role': role,
                'height_m': tree_height,
                'dbh_cm': dbh,
                'tree_class': tree_class,
                'canopy_height_source': canopy_height,
                'forest_type': forest_type,
                'block_name': calculation.block_name or '',
                'generated_date': datetime.now().isoformat(),
                'model_version': config['algorithm_version'],
                'notes': 'SYNTHETIC DATA - Not ground survey',
                'sample_plot_number': pc['plot_number'],
                # Volume calculations (Forest Regulation 2079)
                'stem_volume': volumes['stem_volume'],
                'branch_volume': volumes['branch_volume'],
                'tree_volume': volumes['tree_volume'],
                'gross_volume': volumes['gross_volume'],
                'net_volume': volumes['net_volume'],
                'firewood_m3': volumes['firewood_m3']
            })
            tree_id += 1

    if not trees:
        raise ValueError("No trees generated within sample plot buffers - try increasing buffer distance or check sampling design")

    # Step 3.5: Spatial join - Assign correct block names to trees
    report(75, "Assigning block names via spatial join")
    trees = assign_block_names_to_trees(trees, result_data)

    total_trees_generated = len(trees)
    report(80, f"Generated {total_trees_generated} trees across {len(active_plots)} plots")

    # Step 4.5: Generate regeneration entries (1-10 cm DBH) for each sample plot
    report(87, "Generating regeneration entries")
    regeneration_entries = generate_regeneration_entries(
        sampling_design=sampling_design,
        species_list=species_list,
        buffer_meters=config['plot_buffer_meters'],
        tree_id_start=tree_id
    )

    # Assign block names to regeneration entries
    regeneration_entries = assign_block_names_to_trees(regeneration_entries, result_data)

    # Add regeneration to main tree list
    regeneration_count = len(regeneration_entries)
    trees.extend(regeneration_entries)
    report(88, f"Added {regeneration_count} regeneration entries to sample plots")

    # Step 5: Export to GPKG and Excel
    report(90, "Exporting to GPKG file")
    gpkg_filepath, gpkg_size_mb = export_to_gpkg(trees, calculation_id, db=db)

    # Add block number and total sample plots to each tree for Excel export
    # Get unique blocks and assign block numbers
    unique_blocks = list(dict.fromkeys(t.get('block_name', '') for t in trees if t.get('block_name')))
    block_to_number = {block: idx + 1 for idx, block in enumerate(unique_blocks)}
    
    # Count plots per block
    plots_per_block = {}
    for tree in trees:
        block_name = tree.get('block_name', 'Unknown')
        plot_num = tree.get('sample_plot_number')
        if plot_num and block_name:
            if block_name not in plots_per_block:
                plots_per_block[block_name] = set()
            for pn in str(plot_num).split(','):
                plots_per_block[block_name].add(pn.strip())
    
    # Add block_number and total_sample_plots to each tree
    for tree in trees:
        block_name = tree.get('block_name', '')
        tree['block_number'] = block_to_number.get(block_name, 0)
        tree['total_sample_plots'] = len(plots_per_block.get(block_name, set()))

    report(93, "Exporting to Excel file")
    excel_filepath, excel_size_mb = export_to_excel(trees, calculation_id, db=db)

    # Step 6: Calculate statistics (separate mature trees from regeneration)
    report(97, "Calculating statistics")

    # Separate mature trees (DBH >= 10cm) from regeneration (DBH < 10cm)
    mature_trees = [t for t in trees if t['dbh_cm'] >= 10]
    regeneration_trees = [t for t in trees if t['dbh_cm'] < 10]
    unestablished = [t for t in regeneration_trees if t['dbh_cm'] < 4]
    established = [t for t in regeneration_trees if t['dbh_cm'] >= 4]

    # Calculate effective sampling area (plot buffer area × number of plots)
    plot_area_m2 = 3.14159 * (config['plot_buffer_meters'] ** 2)  # π * r²
    total_plot_area_ha = (plot_area_m2 * sampling_design.total_points) / 10000.0

    # Statistics for mature trees only (heights only measured for mature trees)
    dbhs_mature = [t['dbh_cm'] for t in mature_trees] if mature_trees else [0]
    heights_mature = [t['height_m'] for t in mature_trees if t.get('height_m') is not None]
    if not heights_mature:
        heights_mature = [0]

    # DBH Class distribution per forest block
    # Group trees by block_name
    trees_by_block = {}
    for tree in trees:
        block_name = tree.get('block_name', 'Unknown')
        if block_name not in trees_by_block:
            trees_by_block[block_name] = []
        trees_by_block[block_name].append(tree)

    # Count plots per block (from sample_plot_number assignments)
    plots_per_block = {}
    for tree in trees:
        block_name = tree.get('block_name', 'Unknown')
        plot_num = tree.get('sample_plot_number')
        if plot_num:
            if block_name not in plots_per_block:
                plots_per_block[block_name] = set()
            # Handle comma-separated plot numbers
            for pn in str(plot_num).split(','):
                plots_per_block[block_name].add(pn.strip())

    # Calculate DBH class distribution per block
    block_dbh_distribution = {}
    for block_name, block_trees in trees_by_block.items():
        num_plots_in_block = len(plots_per_block.get(block_name, set())) or 1  # Avoid division by zero

        dbh_counts = {
            'unestablished_regen': len([t for t in block_trees if 1 <= t['dbh_cm'] < 4]),
            'established_regen': len([t for t in block_trees if 4 <= t['dbh_cm'] < 10]),
            'small_pole': len([t for t in block_trees if 10 <= t['dbh_cm'] < 20]),
            'large_pole': len([t for t in block_trees if 20 <= t['dbh_cm'] < 30]),
            'small_tree': len([t for t in block_trees if 30 <= t['dbh_cm'] < 40]),
            'medium_tree': len([t for t in block_trees if 40 <= t['dbh_cm'] < 50]),
            'large_tree': len([t for t in block_trees if 50 <= t['dbh_cm'] < 60]),
            'very_large_tree': len([t for t in block_trees if t['dbh_cm'] >= 60]),
        }

        # Average per plot for this block
        avg_regen = dbh_counts['unestablished_regen'] / num_plots_in_block
        avg_sapling = dbh_counts['established_regen'] / num_plots_in_block
        avg_pole = (dbh_counts['small_pole'] + dbh_counts['large_pole']) / num_plots_in_block
        avg_tree = (dbh_counts['small_tree'] + dbh_counts['medium_tree'] +
                    dbh_counts['large_tree'] + dbh_counts['very_large_tree']) / num_plots_in_block

        # Simplified 4-class system with per-hectare conversion
        # Expansion factors: Regeneration×1000, Sapling×400, Pole×100, Tree×20
        dbh_per_ha = {
            'regeneration_1_4cm': round(avg_regen * 1000, 1),  # Regeneration per ha
            'sapling_4_10cm': round(avg_sapling * 400, 1),      # Sapling per ha
            'pole_10_30cm': round(avg_pole * 100, 1),           # Pole per ha
            'tree_above_30cm': round(avg_tree * 20, 1),         # Tree per ha
        }

        # Calculate volumes per block (Forest Regulation 2079)
        # Separate pole (DBH 10-30cm) and tree (DBH >30cm) volumes
        pole_trees = [t for t in block_trees if 10 <= t.get('dbh_cm', 0) < 30]
        tree_trees = [t for t in block_trees if t.get('dbh_cm', 0) >= 30]

        # Pole volumes
        pole_net_volume = sum(t.get('net_volume', 0.0) for t in pole_trees)
        pole_firewood = sum(t.get('firewood_m3', 0.0) for t in pole_trees)
        
        # Tree volumes
        tree_net_volume = sum(t.get('net_volume', 0.0) for t in tree_trees)
        tree_firewood = sum(t.get('firewood_m3', 0.0) for t in tree_trees)

        # Total volumes
        total_stem_volume = sum(t.get('stem_volume', 0.0) for t in block_trees)
        total_branch_volume = sum(t.get('branch_volume', 0.0) for t in block_trees)
        total_tree_volume = sum(t.get('tree_volume', 0.0) for t in block_trees)
        total_gross_volume = sum(t.get('gross_volume', 0.0) for t in block_trees)
        total_net_volume = sum(t.get('net_volume', 0.0) for t in block_trees)
        total_firewood = sum(t.get('firewood_m3', 0.0) for t in block_trees)

        # Calculate per-hectare using plot factors (matching field inventory method)
        # Pole: plot factor 100, Tree: plot factor 20
        # Formula: volume_per_ha = (total_volume / num_plots) * plot_factor
        if num_plots_in_block > 0:
            # Calculate average per plot first, then multiply by factor
            pole_timber_per_ha = (pole_net_volume / num_plots_in_block) * 100
            pole_firewood_per_ha = (pole_firewood / num_plots_in_block) * 100
            tree_timber_per_ha = (tree_net_volume / num_plots_in_block) * 20
            tree_firewood_per_ha = (tree_firewood / num_plots_in_block) * 20
            
            # Legacy calculation for comparison
            plot_buffer_m = config.get('plot_buffer_meters', 10)
            plot_area_ha = (3.14159 * (plot_buffer_m ** 2)) / 10000.0
            total_sampled_area_ha = plot_area_ha * num_plots_in_block
            
            stem_volume_per_ha = total_stem_volume / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
            branch_volume_per_ha = total_branch_volume / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
            tree_volume_per_ha = total_tree_volume / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
            gross_volume_per_ha = total_gross_volume / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
            net_volume_per_ha = total_net_volume / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
            firewood_per_ha = total_firewood / total_sampled_area_ha if total_sampled_area_ha > 0 else 0
        else:
            pole_timber_per_ha = 0.0
            pole_firewood_per_ha = 0.0
            tree_timber_per_ha = 0.0
            tree_firewood_per_ha = 0.0
            stem_volume_per_ha = 0.0
            branch_volume_per_ha = 0.0
            tree_volume_per_ha = 0.0
            gross_volume_per_ha = 0.0
            net_volume_per_ha = 0.0
            firewood_per_ha = 0.0

        # Total growing stock (timber only) = pole timber + tree timber
        total_growing_stock = pole_timber_per_ha + tree_timber_per_ha
        
        # Total volume (timber + firewood)
        total_volume_per_ha = pole_timber_per_ha + pole_firewood_per_ha + tree_timber_per_ha + tree_firewood_per_ha

        block_dbh_distribution[block_name] = {
            'total_trees': len(block_trees),
            'num_plots': num_plots_in_block,
            'dbh_per_ha': dbh_per_ha,  # Per hectare values
            # Volume totals (Forest Regulation 2079)
            'stem_volume_m3': round(total_stem_volume, 2),
            'branch_volume_m3': round(total_branch_volume, 2),
            'tree_volume_m3': round(total_tree_volume, 2),
            'gross_volume_m3': round(total_gross_volume, 2),
            'net_volume_m3': round(total_net_volume, 2),
            'firewood_m3': round(total_firewood, 2),
            # Per-hectare values (legacy method - total volume / sampled area)
            'stem_volume_per_ha': round(stem_volume_per_ha, 2),
            'branch_volume_per_ha': round(branch_volume_per_ha, 2),
            'tree_volume_per_ha': round(tree_volume_per_ha, 2),
            'gross_volume_per_ha': round(gross_volume_per_ha, 2),
            'net_volume_per_ha': round(net_volume_per_ha, 2),
            'firewood_per_ha': round(firewood_per_ha, 2),
            # Per-hectare values (matching field inventory method - plot factor)
            'pole_timber_m3_per_ha': round(pole_timber_per_ha, 2),
            'pole_firewood_m3_per_ha': round(pole_firewood_per_ha, 2),
            'tree_timber_m3_per_ha': round(tree_timber_per_ha, 2),
            'tree_firewood_m3_per_ha': round(tree_firewood_per_ha, 2),
            # Total growing stock (pole timber + tree timber)
            'total_growing_stock_m3_per_ha': round(total_growing_stock, 2),
            # Total volume (timber + firewood) - for UI display
            'volume_per_ha': round(total_volume_per_ha, 2),
        }

    # Calculate overall volume totals (Forest Regulation 2079)
    # Calculate overall volume totals with pole/tree breakdown
    pole_trees_all = [t for t in trees if 10 <= t.get('dbh_cm', 0) < 30]
    tree_trees_all = [t for t in trees if t.get('dbh_cm', 0) >= 30]
    
    total_stem_volume_all = sum(t.get('stem_volume', 0.0) for t in trees)
    total_branch_volume_all = sum(t.get('branch_volume', 0.0) for t in trees)
    total_tree_volume_all = sum(t.get('tree_volume', 0.0) for t in trees)
    total_gross_volume_all = sum(t.get('gross_volume', 0.0) for t in trees)
    total_net_volume_all = sum(t.get('net_volume', 0.0) for t in trees)
    total_firewood_all = sum(t.get('firewood_m3', 0.0) for t in trees)
    
    # Pole/tree breakdown using plot factor method
    total_pole_net = sum(t.get('net_volume', 0.0) for t in pole_trees_all)
    total_pole_firewood = sum(t.get('firewood_m3', 0.0) for t in pole_trees_all)
    total_tree_net = sum(t.get('net_volume', 0.0) for t in tree_trees_all)
    total_tree_firewood = sum(t.get('firewood_m3', 0.0) for t in tree_trees_all)
    
    # Calculate per hectare using plot factors
    if sampling_design.total_points > 0:
        avg_pole_net_per_plot = total_pole_net / sampling_design.total_points
        avg_pole_firewood_per_plot = total_pole_firewood / sampling_design.total_points
        avg_tree_net_per_plot = total_tree_net / sampling_design.total_points
        avg_tree_firewood_per_plot = total_tree_firewood / sampling_design.total_points
        
        pole_timber_per_ha_all = avg_pole_net_per_plot * 100
        pole_firewood_per_ha_all = avg_pole_firewood_per_plot * 100
        tree_timber_per_ha_all = avg_tree_net_per_plot * 20
        tree_firewood_per_ha_all = avg_tree_firewood_per_plot * 20
    else:
        pole_timber_per_ha_all = 0
        pole_firewood_per_ha_all = 0
        tree_timber_per_ha_all = 0
        tree_firewood_per_ha_all = 0
    
    total_growing_stock_all = pole_timber_per_ha_all + tree_timber_per_ha_all
    total_volume_per_ha_all = pole_timber_per_ha_all + pole_firewood_per_ha_all + tree_timber_per_ha_all + tree_firewood_per_ha_all

    # DBH class per-hectare counts (overall forest level)
    if sampling_design.total_points > 0:
        overall_dbh_per_ha = {
            'regeneration_1_4cm': round((len(unestablished) / sampling_design.total_points) * 1000, 1),
            'sapling_4_10cm': round((len(established) / sampling_design.total_points) * 400, 1),
            'pole_10_30cm': round((len(pole_trees_all) / sampling_design.total_points) * 100, 1),
            'tree_above_30cm': round((len(tree_trees_all) / sampling_design.total_points) * 20, 1),
        }
    else:
        overall_dbh_per_ha = {}

    statistics = {
        'total_trees': len(trees),
        'total_trees_generated': total_trees_generated,  # Before filtering
        'trees_filtered_out': total_trees_generated - len(trees),
        'mature_trees': len(mature_trees),  # DBH >= 10cm
        'regeneration_total': len(regeneration_trees),  # DBH < 10cm
        'regeneration_unestablished': len(unestablished),  # 1-4cm
        'regeneration_established': len(established),  # 4-10cm
        'area_hectares': area_hectares,  # Total forest area
        'sampling_area_hectares': total_plot_area_ha,  # Area within plot buffers
        'trees_per_hectare': len(mature_trees) / total_plot_area_ha if total_plot_area_ha > 0 else 0,
        'regeneration_per_hectare': len(regeneration_trees) / total_plot_area_ha if total_plot_area_ha > 0 else 0,
        'min_dbh_cm': min(dbhs_mature),
        'max_dbh_cm': max(dbhs_mature),
        'mean_dbh_cm': sum(dbhs_mature) / len(dbhs_mature) if dbhs_mature else 0,
        'min_height_m': min(heights_mature),
        'max_height_m': max(heights_mature),
        'mean_height_m': sum(heights_mature) / len(heights_mature) if heights_mature else 0,
        'species_count': len(set(t['species_scientific'] for t in trees if t.get('species_scientific'))),
        'plot_buffer_meters': config['plot_buffer_meters'],
        'total_sample_plots': sampling_design.total_points,
        # Volume totals (Forest Regulation 2079)
        'total_stem_volume_m3': round(total_stem_volume_all, 2),
        'total_branch_volume_m3': round(total_branch_volume_all, 2),
        'total_tree_volume_m3': round(total_tree_volume_all, 2),
        'total_gross_volume_m3': round(total_gross_volume_all, 2),
        'total_net_volume_m3': round(total_net_volume_all, 2),
        'total_firewood_m3': round(total_firewood_all, 2),
        # Per-hectare values (legacy method - total volume / sampled area)
        'stem_volume_per_ha': round(total_stem_volume_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        'branch_volume_per_ha': round(total_branch_volume_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        'tree_volume_per_ha': round(total_tree_volume_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        'gross_volume_per_ha': round(total_gross_volume_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        'net_volume_per_ha': round(total_net_volume_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        'firewood_per_ha': round(total_firewood_all / total_plot_area_ha, 2) if total_plot_area_ha > 0 else 0,
        # Per-hectare values (matching field inventory method - plot factor)
        'pole_timber_m3_per_ha': round(pole_timber_per_ha_all, 2),
        'pole_firewood_m3_per_ha': round(pole_firewood_per_ha_all, 2),
        'tree_timber_m3_per_ha': round(tree_timber_per_ha_all, 2),
        'tree_firewood_m3_per_ha': round(tree_firewood_per_ha_all, 2),
        # Total growing stock (pole timber + tree timber)
        'total_growing_stock_m3_per_ha': round(total_growing_stock_all, 2),
        # Total volume (timber + firewood)
        'volume_per_ha': round(total_volume_per_ha_all, 2),
        'dbh_per_ha': overall_dbh_per_ha,  # Overall DBH class counts per hectare
        'block_dbh_distribution': block_dbh_distribution,  # Block-wise DBH class distribution
    }

    processing_time = (datetime.now() - start_time).total_seconds()

    report(100, "Complete")

    return {
        'gpkg_filepath': gpkg_filepath,
        'gpkg_filename': os.path.basename(gpkg_filepath),
        'gpkg_size_mb': gpkg_size_mb,
        'excel_filepath': excel_filepath,
        'excel_filename': os.path.basename(excel_filepath),
        'excel_size_mb': excel_size_mb,
        # Legacy compatibility (points to GPKG)
        'filepath': gpkg_filepath,
        'filename': os.path.basename(gpkg_filepath),
        'file_size_mb': gpkg_size_mb,
        'statistics': statistics,
        'processing_time_seconds': int(processing_time),
        'config': config
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ALL TREE EXPORT — Full-extent tree generation (1 row = 1 tree, flat format)
# ═══════════════════════════════════════════════════════════════════════════════

ALL_TREE_PIXEL_AREA_HA = 0.0025  # 5m × 5m = 25 m² = 0.0025 ha


def _probabilistic_round(value: float) -> int:
    """
    Round a float probabilistically.

    e.g. 1.25 → 1 (75% chance), 2 (25% chance)
    """
    base = int(math.floor(value))
    frac = value - base
    return base + (1 if random.random() < frac else 0)


def export_to_gpkg_flat(
    trees: List[Dict[str, Any]],
    calculation_id: uuid.UUID,
    db: Session = None,
    output_dir: str = "exports"
) -> Tuple[str, float]:
    """
    Export trees to GPKG in FLAT format (1 row = 1 tree).

    NOT the regulation format — no regen/sapling/pole/tree column split.

    Args:
        trees: List of tree dictionaries
        calculation_id: UUID of calculation
        db: Database session (optional, for forest name)
        output_dir: Directory to save GPKG files

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    forest_name = "forest"
    if db is not None:
        try:
            calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
            if calc and calc.forest_name:
                forest_name = calc.forest_name.replace(' ', '_')
                forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')
        except Exception:
            pass

    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{forest_name}_AllTrees_{timestamp}.gpkg"
    filepath = os.path.join(output_dir, filename)

    records = []
    for tree in trees:
        geom = tree.get('geometry')
        records.append({
            'tree_id': tree.get('tree_id'),
            'block_name': tree.get('block_name', ''),
            'species_code': tree.get('species_code', ''),
            'species_regulation': tree.get('species_regulation', ''),
            'species_scientific': tree.get('species_scientific', ''),
            'species_local': tree.get('species_local', ''),
            'species_role': tree.get('species_role', ''),
            'height_m': tree.get('height_m'),
            'dbh_cm': tree.get('dbh_cm'),
            'tree_class': tree.get('tree_class'),
            'canopy_height_source': tree.get('canopy_height_source'),
            'forest_type': tree.get('forest_type', ''),
            'stem_volume_m3': tree.get('stem_volume', 0.0),
            'branch_volume_m3': tree.get('branch_volume', 0.0),
            'tree_volume_m3': tree.get('tree_volume', 0.0),
            'gross_volume_m3': tree.get('gross_volume', 0.0),
            'net_volume_m3': tree.get('net_volume', 0.0),
            'firewood_m3': tree.get('firewood_m3', 0.0),
            'longitude': geom[0] if geom else None,
            'latitude': geom[1] if geom else None,
            'generated_date': tree.get('generated_date', ''),
            'model_version': tree.get('model_version', ''),
            'notes': tree.get('notes', ''),
            'geometry': Point(geom) if geom else None,
        })

    gdf = gpd.GeoDataFrame(records, crs='EPSG:4326')
    column_order = [
        'tree_id', 'block_name', 'species_code', 'species_regulation', 'species_scientific',
        'species_local', 'species_role', 'height_m', 'dbh_cm', 'tree_class',
        'canopy_height_source', 'forest_type',
        'stem_volume_m3', 'branch_volume_m3', 'tree_volume_m3',
        'gross_volume_m3', 'net_volume_m3', 'firewood_m3',
        'longitude', 'latitude', 'generated_date', 'model_version', 'notes',
        'geometry'
    ]
    gdf = gdf[[c for c in column_order if c in gdf.columns]]
    gdf.to_file(filepath, driver='GPKG', layer='all_trees')

    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
    return filepath, file_size_mb


def export_to_excel_flat(
    trees: List[Dict[str, Any]],
    calculation_id: uuid.UUID,
    db: Session = None,
    output_dir: str = "exports"
) -> Tuple[str, float]:
    """
    Export trees to Excel in FLAT format.

    Sheet 1: "Tree Data" — 1 row = 1 tree, flat columns
    Sheet 2: "Summary" — Block-wise statistics

    Args:
        trees: List of tree dictionaries
        calculation_id: UUID of calculation
        db: Database session
        output_dir: Directory to save Excel files

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    forest_name = "forest"
    if db is not None:
        try:
            calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
            if calc and calc.forest_name:
                forest_name = calc.forest_name.replace(' ', '_')
                forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')
        except Exception:
            pass

    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{forest_name}_AllTrees_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Sheet 1: Tree Data (flat)
    rows = []
    for tree in trees:
        geom = tree.get('geometry')
        rows.append({
            'tree_id': tree.get('tree_id'),
            'block_name': tree.get('block_name', ''),
            'species_code': tree.get('species_code', ''),
            'species_scientific': tree.get('species_scientific', ''),
            'species_local': tree.get('species_local', ''),
            'species_role': tree.get('species_role', ''),
            'height_m': tree.get('height_m'),
            'dbh_cm': tree.get('dbh_cm'),
            'tree_class': tree.get('tree_class'),
            'canopy_height_source': tree.get('canopy_height_source'),
            'forest_type': tree.get('forest_type', ''),
            'stem_volume_m3': tree.get('stem_volume', 0.0),
            'branch_volume_m3': tree.get('branch_volume', 0.0),
            'tree_volume_m3': tree.get('tree_volume', 0.0),
            'gross_volume_m3': tree.get('gross_volume', 0.0),
            'net_volume_m3': tree.get('net_volume', 0.0),
            'firewood_m3': tree.get('firewood_m3', 0.0),
            'longitude': geom[0] if geom else None,
            'latitude': geom[1] if geom else None,
        })

    df_trees = pd.DataFrame(rows)

    # Sheet 2: Summary by block
    summary_rows = []
    for block_name in df_trees['block_name'].unique():
        block_df = df_trees[df_trees['block_name'] == block_name]
        total = len(block_df)
        mature = len(block_df[block_df['dbh_cm'] >= 10])
        pole = len(block_df[(block_df['dbh_cm'] >= 10) & (block_df['dbh_cm'] < 30)])
        tree_large = len(block_df[block_df['dbh_cm'] >= 30])
        total_vol = block_df['tree_volume_m3'].sum()
        net_timber = block_df['net_volume_m3'].sum()
        firewood = block_df['firewood_m3'].sum()
        summary_rows.append({
            'block_name': block_name,
            'total_trees': total,
            'pole_trees_10_30cm': pole,
            'trees_above_30cm': tree_large,
            'total_volume_m3': round(total_vol, 2),
            'net_timber_m3': round(net_timber, 2),
            'firewood_m3': round(firewood, 2),
        })

    df_summary = pd.DataFrame(summary_rows)

    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        df_trees.to_excel(writer, sheet_name='Tree Data', index=False)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        workbook = writer.book
        header_fmt = workbook.add_format({
            'bg_color': '#2F4F4F', 'font_color': '#FFFFFF', 'bold': True,
            'align': 'center', 'valign': 'vcenter', 'border': 0,
        })
        for ws_name in ['Tree Data', 'Summary']:
            ws = writer.sheets[ws_name]
            for col_num, _ in enumerate(
                df_trees.columns if ws_name == 'Tree Data' else df_summary.columns
            ):
                ws.write(0, col_num, _, header_fmt)
            ws.autofilter(0, 0, len(df_trees) if ws_name == 'Tree Data' else len(df_summary),
                          len(df_trees.columns if ws_name == 'Tree Data' else df_summary.columns) - 1)
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
    return filepath, file_size_mb


def export_to_csv_flat(
    trees: List[Dict[str, Any]],
    calculation_id: uuid.UUID,
    db: Session = None,
    output_dir: str = "exports"
) -> Tuple[str, float]:
    """
    Export trees to CSV in FLAT format (1 row = 1 tree).

    Args:
        trees: List of tree dictionaries
        calculation_id: UUID of calculation
        db: Database session
        output_dir: Directory to save CSV files

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    forest_name = "forest"
    if db is not None:
        try:
            calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
            if calc and calc.forest_name:
                forest_name = calc.forest_name.replace(' ', '_')
                forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')
        except Exception:
            pass

    timestamp = datetime.now().strftime("%Y%m%d")
    filename = f"{forest_name}_AllTrees_{timestamp}.csv"
    filepath = os.path.join(output_dir, filename)

    rows = []
    for tree in trees:
        geom = tree.get('geometry')
        rows.append({
            'tree_id': tree.get('tree_id'),
            'block_name': tree.get('block_name', ''),
            'species_code': tree.get('species_code', ''),
            'species_scientific': tree.get('species_scientific', ''),
            'species_local': tree.get('species_local', ''),
            'species_role': tree.get('species_role', ''),
            'height_m': tree.get('height_m'),
            'dbh_cm': tree.get('dbh_cm'),
            'tree_class': tree.get('tree_class'),
            'canopy_height_source': tree.get('canopy_height_source'),
            'forest_type': tree.get('forest_type', ''),
            'stem_volume_m3': tree.get('stem_volume', 0.0),
            'branch_volume_m3': tree.get('branch_volume', 0.0),
            'tree_volume_m3': tree.get('tree_volume', 0.0),
            'gross_volume_m3': tree.get('gross_volume', 0.0),
            'net_volume_m3': tree.get('net_volume', 0.0),
            'firewood_m3': tree.get('firewood_m3', 0.0),
            'longitude': geom[0] if geom else None,
            'latitude': geom[1] if geom else None,
        })

    df = pd.DataFrame(rows)
    df.to_csv(filepath, index=False)

    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
    return filepath, file_size_mb


def generate_full_extent_trees(
    calculation_id: uuid.UUID,
    db: Session,
    config: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[callable] = None
) -> Dict[str, Any]:
    """
    Generate ALL trees within the forest boundary from canopy height raster.

    Unlike generate_synthetic_trees() which only creates trees within sample
    plot buffers, this function processes EVERY valid pixel in the forest
    boundary and creates trees across the entire extent.

    One row = one tree (flat format), not the regulation split format.

    Args:
        calculation_id: UUID of calculation with boundary
        db: Database session
        config: Optional configuration overrides
        progress_callback: Optional function(percent, step)

    Returns:
        Dictionary with generation results and statistics
    """
    start_time = datetime.now()

    default_config = {
        'min_dbh_cm': MIN_DBH_CM,
        'min_height_m': MIN_HEIGHT_M,
        'max_trees_per_ha': MAX_TREES_PER_HA,
        'algorithm_version': 'v1.0',
        'species_role_target_ratio': None,
    }
    config = {**default_config, **(config or {})}

    def report(percent: int, step: str):
        if progress_callback:
            progress_callback(percent, step)

    # Step 1: Load calculation
    report(5, "Loading calculation data")
    calculation = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    if not calculation:
        raise ValueError(f"Calculation {calculation_id} not found")

    result_data = calculation.result_data or {}
    species_list = result_data.get('potential_species', [])
    if not species_list:
        raise ValueError("No species data found in calculation")

    # Filter to tree species only
    tree_species_only = [sp for sp in species_list if sp.get('is_tree_species', True)]
    if not tree_species_only:
        raise ValueError(f"No tree species found. Total: {len(species_list)}, none classified as trees.")
    if len(tree_species_only) < len(species_list):
        print(f"INFO: Filtered out {len(species_list) - len(tree_species_only)} non-tree species.")
    species_list = tree_species_only

    # Step 2: Get boundary geometry
    report(10, "Loading boundary geometry")
    if not calculation.boundary_geom:
        raise ValueError("Calculation has no boundary geometry")

    from geoalchemy2.shape import to_shape
    boundary_shape = to_shape(calculation.boundary_geom)
    forest_type = result_data.get('forest_type_dominant', 'Unknown')

    # Step 3: Load species coefficients
    report(12, "Loading species coefficients")
    from sqlalchemy import text as sql_text
    species_coefficients = {}
    coef_query = sql_text("""
        SELECT scientific_name, species_code, a, b, c, s, m, bg, a1, b1,
               typical_hd_ratio_min, typical_hd_ratio_max,
               max_dbh_cm, max_height_m, growth_rate, full_stem_merchantable
        FROM tree_species_coefficients
    """)
    for row in db.execute(coef_query):
        species_coefficients[row.scientific_name] = {
            'species_code': row.species_code,
            'a': row.a, 'b': row.b, 'c': row.c,
            's': row.s, 'm': row.m, 'bg': row.bg,
            'a1': row.a1, 'b1': row.b1,
            'typical_hd_ratio_min': row.typical_hd_ratio_min,
            'typical_hd_ratio_max': row.typical_hd_ratio_max,
            'max_dbh_cm': row.max_dbh_cm, 'max_height_m': row.max_height_m,
            'growth_rate': row.growth_rate,
            'full_stem_merchantable': bool(row.full_stem_merchantable) if row.full_stem_merchantable is not None else False,
        }

    # Merge coefficients into species list for DBH/height calculations
    for sp in species_list:
        sci_name = sp.get('scientific_name')
        if sci_name and sci_name in species_coefficients:
            sp.update(species_coefficients[sci_name])

    # Step 4: Extract all pixels from TIFF within boundary
    report(20, "Extracting canopy height pixels from TIFF")
    import rasterio
    from rasterio.mask import mask as rio_mask
    from shapely.geometry import mapping

    tiff_path = r"D:\forest_management\canopy_height5m.tif"

    with rasterio.open(tiff_path) as src:
        out_image, out_transform = rio_mask(
            src,
            [mapping(boundary_shape)],
            crop=True,
            nodata=0,
            all_touched=False,
        )
        height_band = out_image[0]

    # Collect valid pixels with height > 0
    import numpy as np
    valid_rows, valid_cols = np.where(height_band > 0)
    valid_pixels = []
    for row, col in zip(valid_rows, valid_cols):
        height = float(height_band[row, col])
        if height < config['min_height_m']:
            continue
        # Get pixel bounds in geographic coordinates
        x_ul, y_ul = rasterio.transform.xy(out_transform, row, col, offset='ul')
        x_lr, y_lr = rasterio.transform.xy(out_transform, row, col, offset='lr')
        # Normalize bounds (handle negative x step)
        min_x, max_x = (x_ul, x_lr) if x_ul < x_lr else (x_lr, x_ul)
        min_y, max_y = (y_lr, y_ul) if y_lr < y_ul else (y_ul, y_lr)
        center_x = (min_x + max_x) / 2
        center_y = (min_y + max_y) / 2
        valid_pixels.append({
            'height': height,
            'center': (center_x, center_y),
            'bounds': (min_x, min_y, max_x, max_y),
        })

    if not valid_pixels:
        raise ValueError(
            "No valid canopy height pixels found within the forest boundary. "
            "Ensure the TIFF overlaps with the boundary and contains height values > 0."
        )

    report(30, f"Found {len(valid_pixels)} pixels. Generating trees...")
    print(f"INFO: Processing {len(valid_pixels)} valid pixels across the forest boundary.")

    # Step 5: Generate trees for each pixel
    trees = []
    tree_id = 1
    total_pixels = len(valid_pixels)

    # Pre-group species by role if ratio override is provided
    role_target_ratio = config.get('species_role_target_ratio')

    for idx, pixel in enumerate(valid_pixels):
        if (idx + 1) % max(1, total_pixels // 10) == 0:
            pct = 30 + int((idx + 1) / total_pixels * 45)
            report(pct, f"Generating trees: {idx + 1}/{total_pixels} pixels")

        canopy_height = pixel['height']
        trees_per_ha = get_tree_density(canopy_height, forest_type, config['max_trees_per_ha'])
        if trees_per_ha == 0:
            continue

        raw_trees_per_pixel = trees_per_ha * ALL_TREE_PIXEL_AREA_HA
        num_trees = _probabilistic_round(raw_trees_per_pixel)
        if num_trees == 0:
            continue

        for _ in range(num_trees):
            # Random position within pixel bounds
            x, y = generate_random_point_in_pixel(pixel['bounds'])

            # Select species (with optional ratio override)
            species = weighted_random_choice_with_ratio(species_list, role_target_ratio)
            role = get_species_role(species.get('availability_rank', 3))

            # Assign height
            tree_height = assign_tree_height(canopy_height, species, role)

            # Calculate DBH
            dbh = calculate_dbh_from_height(tree_height, species)

            # Filter by thresholds
            max_dbh = config.get('max_dbh_cm')
            if dbh < config['min_dbh_cm'] or (max_dbh and dbh > max_dbh) or tree_height < config['min_height_m']:
                continue

            # Assign tree class
            tree_class = assign_tree_class(dbh, tree_height, species)

            # Calculate volumes
            volumes = {'stem_volume': 0.0, 'branch_volume': 0.0, 'tree_volume': 0.0,
                       'gross_volume': 0.0, 'net_volume': 0.0, 'firewood_m3': 0.0}

            scientific_name = species.get('scientific_name')
            if scientific_name and scientific_name in species_coefficients:
                volumes = calculate_tree_volumes(
                    dbh=dbh,
                    height=tree_height,
                    tree_class=tree_class,
                    species_coefficients=species_coefficients[scientific_name]
                )

            trees.append({
                'tree_id': tree_id,
                'geometry': (x, y),
                'species_code': species.get('species_code'),
                'species_regulation': get_species_regulation(species.get('species_code'), scientific_name or ''),
                'species_scientific': scientific_name,
                'species_local': species.get('local_name'),
                'species_role': role,
                'height_m': tree_height,
                'dbh_cm': dbh,
                'tree_class': tree_class,
                'canopy_height_source': canopy_height,
                'forest_type': forest_type,
                'block_name': calculation.block_name or '',
                'generated_date': datetime.now().isoformat(),
                'model_version': config['algorithm_version'],
                'notes': 'SYNTHETIC DATA - Generated from canopy height model',
                'stem_volume': volumes['stem_volume'],
                'branch_volume': volumes['branch_volume'],
                'tree_volume': volumes['tree_volume'],
                'gross_volume': volumes['gross_volume'],
                'net_volume': volumes['net_volume'],
                'firewood_m3': volumes['firewood_m3'],
            })
            tree_id += 1

    if not trees:
        raise ValueError("No trees generated — all pixels resulted in zero trees after filtering. "
                         "Try lowering min_dbh_cm or min_height_m.")

    # Step 6: Spatial join — assign block names
    report(78, "Assigning block names via spatial join")
    trees = assign_block_names_to_trees(trees, result_data)

    total_trees = len(trees)
    report(80, f"Generated {total_trees} trees across the forest boundary")

    # Step 7: Export GPKG only (primary format)
    report(85, "Exporting to GPKG")
    gpkg_filepath, gpkg_size_mb = export_to_gpkg_flat(trees, calculation_id, db=db)
    # Excel and CSV skipped — GPKG is the canonical flat format

    # Step 8: Statistics
    report(95, "Calculating statistics")

    area_hectares = result_data.get('area_hectares', 0)
    dbhs = [t['dbh_cm'] for t in trees]
    heights = [t['height_m'] for t in trees if t.get('height_m') is not None]

    # Species role distribution
    role_dist = {}
    for t in trees:
        r = t.get('species_role', 'unknown')
        role_dist[r] = role_dist.get(r, 0) + 1

    # DBH class distribution
    dbh_classes = {
        'unestablished_regen_1_4cm': len([t for t in trees if t['dbh_cm'] < 4]),
        'established_regen_4_10cm': len([t for t in trees if 4 <= t['dbh_cm'] < 10]),
        'small_pole_10_20cm': len([t for t in trees if 10 <= t['dbh_cm'] < 20]),
        'large_pole_20_30cm': len([t for t in trees if 20 <= t['dbh_cm'] < 30]),
        'small_tree_30_40cm': len([t for t in trees if 30 <= t['dbh_cm'] < 40]),
        'medium_tree_40_50cm': len([t for t in trees if 40 <= t['dbh_cm'] < 50]),
        'large_tree_50_60cm': len([t for t in trees if 50 <= t['dbh_cm'] < 60]),
        'very_large_tree_60plus': len([t for t in trees if t['dbh_cm'] >= 60]),
    }

    # Block-wise stats
    trees_by_block = {}
    for t in trees:
        bname = t.get('block_name', 'Unknown')
        if bname not in trees_by_block:
            trees_by_block[bname] = []
        trees_by_block[bname].append(t)

    block_stats = {}
    for bname, btrees in trees_by_block.items():
        b_vol = sum(t.get('tree_volume', 0.0) for t in btrees)
        b_net = sum(t.get('net_volume', 0.0) for t in btrees)
        b_fire = sum(t.get('firewood_m3', 0.0) for t in btrees)
        block_stats[bname] = {
            'total_trees': len(btrees),
            'total_volume_m3': round(b_vol, 2),
            'net_timber_m3': round(b_net, 2),
            'firewood_m3': round(b_fire, 2),
        }

    statistics = {
        'total_trees': total_trees,
        'area_hectares': area_hectares,
        'trees_per_hectare': round(total_trees / area_hectares, 2) if area_hectares > 0 else 0,
        'min_dbh_cm': min(dbhs),
        'max_dbh_cm': max(dbhs),
        'mean_dbh_cm': round(sum(dbhs) / len(dbhs), 2),
        'min_height_m': min(heights) if heights else 0,
        'max_height_m': max(heights) if heights else 0,
        'mean_height_m': round(sum(heights) / len(heights), 2) if heights else 0,
        'species_count': len(set(t.get('species_scientific') for t in trees if t.get('species_scientific'))),
        'total_pixels_processed': len(valid_pixels),
        'species_role_distribution': role_dist,
        'dbh_class_distribution': dbh_classes,
        'block_wise_stats': block_stats,
        'total_stem_volume_m3': round(sum(t.get('stem_volume', 0.0) for t in trees), 2),
        'total_branch_volume_m3': round(sum(t.get('branch_volume', 0.0) for t in trees), 2),
        'total_tree_volume_m3': round(sum(t.get('tree_volume', 0.0) for t in trees), 2),
        'total_gross_volume_m3': round(sum(t.get('gross_volume', 0.0) for t in trees), 2),
        'total_net_volume_m3': round(sum(t.get('net_volume', 0.0) for t in trees), 2),
        'total_firewood_m3': round(sum(t.get('firewood_m3', 0.0) for t in trees), 2),
    }

    processing_time = (datetime.now() - start_time).total_seconds()
    report(100, "Complete")

    return {
        'gpkg_filepath': gpkg_filepath,
        'gpkg_filename': os.path.basename(gpkg_filepath),
        'gpkg_size_mb': gpkg_size_mb,
        'statistics': statistics,
        'processing_time_seconds': int(processing_time),
        'config': config,
    }


def extract_sample_plot_trees(
    all_tree_gpkg_path: str,
    calculation_id: uuid.UUID,
    db: Session,
    output_dir: str = "exports",
    output_filename: Optional[str] = None,
    plot_buffer_meters: float = 12.62,
) -> Dict[str, Any]:
    """
    Extract trees from the all-tree GPKG that fall within sample plot buffers.

    This provides a consistent sample-plot subset from the full-extent tree model,
    enabling plot-by-plot comparison with field inventory measurements.

    Args:
        all_tree_gpkg_path: Path to the all-tree GPKG file
        calculation_id: UUID of calculation
        db: Database session
        output_dir: Directory to save output
        plot_buffer_meters: Buffer radius in meters (default 12.62 ≈ 500m² plot)

    Returns:
        Dictionary with results including path and statistics
    """
    start_time = datetime.now()
    import geopandas as gpd
    from shapely.geometry import Point
    from geoalchemy2.shape import to_shape

    if not os.path.exists(all_tree_gpkg_path):
        raise ValueError(f"All-tree GPKG not found: {all_tree_gpkg_path}")

    # Step 1: Read the all-tree GPKG
    gdf_all = gpd.read_file(all_tree_gpkg_path, layer='all_trees')
    print(f"Loaded {len(gdf_all)} trees from all-tree GPKG")

    # Step 2: Get sampling design
    from ..models.sampling import SamplingDesign
    sampling_design = db.query(SamplingDesign).filter(
        SamplingDesign.calculation_id == calculation_id
    ).first()

    if not sampling_design or not sampling_design.points_geometry:
        raise ValueError("No sampling design found for this calculation")

    # Step 3: Extract plot centers and create buffers
    sample_points_geom = to_shape(sampling_design.points_geometry)
    plot_assignments = sampling_design.points_block_assignment or []
    buffer_deg = plot_buffer_meters / 111320.0

    plot_buffers = []
    if hasattr(sample_points_geom, 'geoms'):
        for idx, point in enumerate(sample_points_geom.geoms):
            plot_info = next((p for p in plot_assignments if p.get('point_index') == idx), None)
            plot_number = plot_info.get('plot_number', idx + 1) if plot_info else idx + 1
            plot_buffers.append({
                'plot_number': str(plot_number),
                'buffer': point.buffer(buffer_deg),
                'center': (point.x, point.y),
            })
    else:
        plot_buffers.append({
            'plot_number': '1',
            'buffer': sample_points_geom.buffer(buffer_deg),
            'center': (sample_points_geom.x, sample_points_geom.y),
        })

    print(f"Found {len(plot_buffers)} sample plots")

    # Step 4: Spatial filter — trees within each plot buffer
    gdf_filtered = gpd.GeoDataFrame(
        {'plot_number': [], 'tree_id': [], 'geometry': []}, crs='EPSG:4326'
    )
    for plot in plot_buffers:
        buffer_shape = plot['buffer']
        mask = gdf_all.within(buffer_shape)
        plot_trees = gdf_all[mask].copy()
        plot_trees['plot_number'] = plot['plot_number']
        gdf_filtered = gpd.GeoDataFrame(
            pd.concat([gdf_filtered, plot_trees], ignore_index=True),
            crs='EPSG:4326',
        )

    total_extracted = len(gdf_filtered)
    if total_extracted == 0:
        raise ValueError("No trees found within sample plot buffers")

    print(f"Extracted {total_extracted} trees within {len(plot_buffers)} plots")

    # Step 5: Move plot_number column to front
    cols = ['plot_number'] + [c for c in gdf_filtered.columns if c != 'plot_number' and c != 'geometry'] + ['geometry']
    gdf_filtered = gdf_filtered[[c for c in cols if c in gdf_filtered.columns]]

    # Step 6: Write output GPKG
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    forest_name = "forest"
    calc = db.query(Calculation).filter(Calculation.id == calculation_id).first()
    if calc and calc.forest_name:
        forest_name = calc.forest_name.replace(' ', '_')
        forest_name = ''.join(c for c in forest_name if c.isalnum() or c == '_')

    if output_filename:
        filename = output_filename
    else:
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"{forest_name}_AllTrees_SamplePlots_{timestamp}.gpkg"
    filepath = os.path.join(output_dir, filename)
    gdf_filtered.to_file(filepath, driver='GPKG', layer='sample_plot_trees')
    file_size_mb = os.path.getsize(filepath) / (1024 * 1024)

    # Step 7: Per-plot statistics
    plot_stats = {}
    for plot in plot_buffers:
        pn = plot['plot_number']
        plot_trees = gdf_filtered[gdf_filtered['plot_number'] == pn]
        plot_stats[pn] = {
            'tree_count': len(plot_trees),
            'mean_dbh_cm': round(plot_trees['dbh_cm'].mean(), 2) if len(plot_trees) > 0 else 0,
            'mean_height_m': round(plot_trees['height_m'].mean(), 2) if len(plot_trees) > 0 else 0,
            'total_volume_m3': round(plot_trees['tree_volume_m3'].sum(), 2) if len(plot_trees) > 0 and 'tree_volume_m3' in plot_trees.columns else 0,
        }

    processing_time = (datetime.now() - start_time).total_seconds()

    return {
        'filepath': filepath,
        'filename': filename,
        'size_mb': round(file_size_mb, 2),
        'total_trees_extracted': total_extracted,
        'total_plots': len(plot_buffers),
        'plot_statistics': plot_stats,
        'processing_time_seconds': int(processing_time),
    }
