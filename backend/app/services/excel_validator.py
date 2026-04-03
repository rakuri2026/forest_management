"""
Excel Validation Enhancement for Tree Model Export

Adds data validation rules to Excel export for field data editing.
- Numeric columns have strict range validation
- Species columns have dropdown suggestions with conditional formatting
- Users can enter custom values in species columns

Author: Forest Management System
Date: February 23, 2026
"""

from typing import Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string
from sqlalchemy.orm import Session
from sqlalchemy import text
import os


def add_validation_to_excel(
    excel_filepath: str,
    db: Session
) -> Tuple[str, float]:
    """
    Enhance existing Excel export with data validation rules

    Adds:
    - SPECIES_DATABASE hidden sheet with all tree species
    - Data validation rules for numeric columns (strict)
    - List validation for species columns (dropdown suggestions)
    - Conditional formatting for species (valid=green, invalid=red)
    - VALIDATION_GUIDE sheet with instructions

    Args:
        excel_filepath: Path to existing Excel file
        db: Database session for querying species

    Returns:
        Tuple of (filepath, file_size_mb)
    """
    # Load existing workbook
    wb = load_workbook(excel_filepath)
    ws_data = wb['Data Template']

    # Step 1: Create SPECIES_DATABASE sheet (hidden, for validation reference)
    ws_species = wb.create_sheet("SPECIES_DATABASE")
    ws_species.sheet_state = 'hidden'

    # Query all active tree species
    query = text("""
        SELECT
            species_code,
            scientific_name,
            local_name,
            is_tree_species,
            max_dbh_cm,
            max_height_m,
            economic_value
        FROM tree_species_coefficients
        WHERE is_active = TRUE
        ORDER BY
            species_code NULLS LAST,
            scientific_name
    """)

    result = db.execute(query)
    species_list = result.fetchall()

    # Write species database headers with styling
    headers = ['species_code', 'scientific_name', 'local_name', 'is_tree_species',
               'max_dbh_cm', 'max_height_m', 'economic_value']
    ws_species.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws_species[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write species data
    for sp in species_list:
        ws_species.append([
            sp[0] or '',  # species_code
            sp[1] or '',  # scientific_name
            sp[2] or '',  # local_name
            'TRUE' if sp[3] else 'FALSE',  # is_tree_species
            sp[4] or 100,  # max_dbh_cm
            sp[5] or 30,  # max_height_m
            sp[6] or 'Moderate'  # economic_value
        ])

    # Auto-adjust column widths in species sheet
    for column in ws_species.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_species.column_dimensions[column_letter].width = min(max_length + 2, 30)

    # Get column letters mapping
    col_map = {}
    for idx, cell in enumerate(ws_data[1], start=1):
        col_map[cell.value] = get_column_letter(idx)

    print(f"[Excel Validator] Found columns: {list(col_map.keys())}")
    print(f"[Excel Validator] Total columns: {len(col_map)}")

    # Get max row
    max_row = ws_data.max_row

    # Guard: If no data rows (only header), add a note and skip validation
    if max_row < 2:
        print(f"WARNING: Excel file has only header row (max_row={max_row}). Adding empty row for validation.")
        # Add a single empty row for validation to work
        ws_data.append([None] * len(col_map))
        max_row = ws_data.max_row

    # ===== NUMERIC VALIDATION (Strict) =====
    validation_rules = []

    # regen_dbh: minimum value=1, maximum value=3.99
    if 'regen_dbh' in col_map:
        validation_rules.append({
            'col': col_map['regen_dbh'],
            'type': 'decimal',
            'operator': 'between',
            'formula1': 1,
            'formula2': 3.99,
            'error_title': 'Invalid DBH',
            'error': 'Regeneration DBH must be between 1 and 3.99 cm'
        })

    # regen_count: min=1, max=15
    if 'regen_count' in col_map:
        validation_rules.append({
            'col': col_map['regen_count'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 1,
            'formula2': 15,
            'error_title': 'Invalid Count',
            'error': 'Regeneration count must be between 1 and 15'
        })

    # sapling_dbh_cm: minimum value=4, maximum value=9.99
    if 'sapling_dbh_cm' in col_map:
        validation_rules.append({
            'col': col_map['sapling_dbh_cm'],
            'type': 'decimal',
            'operator': 'between',
            'formula1': 4,
            'formula2': 9.99,
            'error_title': 'Invalid DBH',
            'error': 'Sapling DBH must be between 4 and 9.99 cm'
        })

    # sapling_count: min=1, max=10
    if 'sapling_count' in col_map:
        validation_rules.append({
            'col': col_map['sapling_count'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 1,
            'formula2': 10,
            'error_title': 'Invalid Count',
            'error': 'Sapling count must be between 1 and 10'
        })

    # pole_dbh_cm: minimum value=10, maximum value=29.99
    if 'pole_dbh_cm' in col_map:
        validation_rules.append({
            'col': col_map['pole_dbh_cm'],
            'type': 'decimal',
            'operator': 'between',
            'formula1': 10,
            'formula2': 29.99,
            'error_title': 'Invalid DBH',
            'error': 'Pole DBH must be between 10 and 29.99 cm'
        })

    # pole_height_m: minimum value=5, maximum value=25
    if 'pole_height_m' in col_map:
        validation_rules.append({
            'col': col_map['pole_height_m'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 5,
            'formula2': 25,
            'error_title': 'Invalid Height',
            'error': 'Pole height must be between 5 and 25 meters'
        })

    # tree_dbh_cm: minimum value=30, maximum value=200
    if 'tree_dbh_cm' in col_map:
        validation_rules.append({
            'col': col_map['tree_dbh_cm'],
            'type': 'decimal',
            'operator': 'between',
            'formula1': 30,
            'formula2': 200,
            'error_title': 'Invalid DBH',
            'error': 'Tree DBH must be between 30 and 200 cm'
        })

    # tree_height_m: minimum value=7, maximum value=50
    if 'tree_height_m' in col_map:
        validation_rules.append({
            'col': col_map['tree_height_m'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 7,
            'formula2': 50,
            'error_title': 'Invalid Height',
            'error': 'Tree height must be between 7 and 50 meters'
        })

    # firewood_kg_per_100sqm_per_year: minimum value=0, maximum value=500
    if 'firewood_kg_per_100sqm_per_year' in col_map:
        validation_rules.append({
            'col': col_map['firewood_kg_per_100sqm_per_year'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 0,
            'formula2': 500,
            'error_title': 'Invalid Firewood',
            'error': 'Firewood must be between 0 and 500 kg/100sqm/year'
        })

    # grass_kg_per_100sqm_per_year: minimum value=0, maximum value=500
    if 'grass_kg_per_100sqm_per_year' in col_map:
        validation_rules.append({
            'col': col_map['grass_kg_per_100sqm_per_year'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 0,
            'formula2': 500,
            'error_title': 'Invalid Grass',
            'error': 'Grass must be between 0 and 500 kg/100sqm/year'
        })

    # bedding_material_kg_per_100sqm_per_year: minimum value=0, maximum value=500
    if 'bedding_material_kg_per_100sqm_per_year' in col_map:
        validation_rules.append({
            'col': col_map['bedding_material_kg_per_100sqm_per_year'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 0,
            'formula2': 500,
            'error_title': 'Invalid Bedding Material',
            'error': 'Bedding material must be between 0 and 500 kg/100sqm/year'
        })

    # ntfp_kg_per_100sqm_per_year: minimum value=0, maximum value=100
    if 'ntfp_kg_per_100sqm_per_year' in col_map:
        validation_rules.append({
            'col': col_map['ntfp_kg_per_100sqm_per_year'],
            'type': 'whole',
            'operator': 'between',
            'formula1': 0,
            'formula2': 100,
            'error_title': 'Invalid NTFP',
            'error': 'NTFP must be between 0 and 100 kg/100sqm/year'
        })

    # Apply numeric validation rules
    for rule in validation_rules:
        dv = DataValidation(
            type=rule['type'],
            operator=rule.get('operator'),
            formula1=rule['formula1'],
            formula2=rule['formula2'],
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=rule['error_title'],
            error=rule['error']
        )
        dv.add(f"{rule['col']}2:{rule['col']}{max_row}")
        ws_data.add_data_validation(dv)
        print(f"[Excel Validator] Added validation for column {rule['col']}: {rule.get('error', 'list validation')}")

    print(f"[Excel Validator] Applied {len(validation_rules)} numeric validation rules")

    # pole_class and tree_class: list validation (1-4, A-D, a-d, i-iv, I-IV)
    class_values = "1,2,3,4,A,B,C,D,a,b,c,d,i,ii,iii,iv,I,II,III,IV"

    if 'pole_class' in col_map:
        dv_class = DataValidation(
            type='list',
            formula1=f'"{class_values}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='Invalid Class',
            error='Tree class must be one of: 1,2,3,4,A,B,C,D,a,b,c,d,i,ii,iii,iv,I,II,III,IV'
        )
        dv_class.add(f"{col_map['pole_class']}2:{col_map['pole_class']}{max_row}")
        ws_data.add_data_validation(dv_class)

    if 'tree_class' in col_map:
        dv_class = DataValidation(
            type='list',
            formula1=f'"{class_values}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='Invalid Class',
            error='Tree class must be one of: 1,2,3,4,A,B,C,D,a,b,c,d,i,ii,iii,iv,I,II,III,IV'
        )
        dv_class.add(f"{col_map['tree_class']}2:{col_map['tree_class']}{max_row}")
        ws_data.add_data_validation(dv_class)

    # ===== SPECIES VALIDATION - Dropdown with Auto-conversion =====
    # Create combined species list (scientific + local names) in SPECIES_DATABASE for dropdown

    # Add a new column header for combined display
    species_db_last_col = get_column_letter(ws_species.max_column + 1)
    ws_species[f'{species_db_last_col}1'] = 'display_name_combined'
    ws_species[f'{species_db_last_col}1'].fill = header_fill
    ws_species[f'{species_db_last_col}1'].font = header_font

    # Populate combined display names "Scientific Name (Local Name)"
    for row_idx in range(2, len(species_list) + 2):
        scientific = ws_species[f'B{row_idx}'].value or ''
        local = ws_species[f'C{row_idx}'].value or ''
        if scientific and local:
            ws_species[f'{species_db_last_col}{row_idx}'] = f"{scientific} ({local})"
        elif scientific:
            ws_species[f'{species_db_last_col}{row_idx}'] = scientific
        else:
            ws_species[f'{species_db_last_col}{row_idx}'] = local

    # Create SPECIES_LOOKUP sheet for dropdowns (hidden)
    ws_lookup = wb.create_sheet("SPECIES_LOOKUP")
    ws_lookup.sheet_state = 'hidden'

    # Build combined list: scientific names + local names
    all_species_names = []
    for sp in species_list:
        scientific = sp[1] or ''
        local = sp[2] or ''
        if scientific:
            all_species_names.append(scientific)
        if local and local != scientific:  # Add local name only if different
            all_species_names.append(local)

    # Remove duplicates and sort
    all_species_names = sorted(list(set(all_species_names)))

    # Write to SPECIES_LOOKUP sheet
    ws_lookup['A1'] = 'species_list'
    ws_lookup['A1'].font = header_font
    ws_lookup['A1'].fill = header_fill
    for idx, name in enumerate(all_species_names, start=2):
        ws_lookup[f'A{idx}'] = name

    lookup_range = f"SPECIES_LOOKUP!$A$2:$A${len(all_species_names) + 1}"

    # Get number of species rows for validation range
    species_count = len(species_list)

    print(f"[Excel Validator] Created species lookup list with {len(all_species_names)} entries")

    # ===== SPECIES DATA VALIDATION - Dropdown =====
    # Add dropdown validation to species columns
    species_columns = [
        'regen_species_scientific',
        'sapling_species_scientific',
        'pole_species_scientific',
        'tree_species_scientific'
    ]

    for species_col_name in species_columns:
        if species_col_name in col_map:
            col_letter = col_map[species_col_name]

            # Add autocomplete validation (showDropDown=False enables type-ahead)
            dv_species = DataValidation(
                type='list',
                formula1=lookup_range,
                allow_blank=True,
                showDropDown=False,  # This enables autocomplete/type-ahead behavior!
                showErrorMessage=False  # Allow custom entries
            )
            # Add helpful prompts
            dv_species.prompt = 'Type scientific or local name. List will filter as you type.'
            dv_species.promptTitle = 'Species Entry'
            dv_species.error = 'Species not found in database. You can still use this value or check spelling.'
            dv_species.errorTitle = 'Species Validation'

            dv_species.add(f"{col_letter}2:{col_letter}{max_row}")
            ws_data.add_data_validation(dv_species)

            print(f"[Excel Validator] Added autocomplete validation to {col_letter} ({species_col_name})")

    # ===== SPECIES CONDITIONAL FORMATTING =====
    # Add conditional formatting for species columns (green for valid, red for invalid)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")

    print(f"[Excel Validator] Adding species conditional formatting for {len(species_columns)} columns")

    for species_col_name in species_columns:
        if species_col_name in col_map:
            col_letter = col_map[species_col_name]
            col_num = column_index_from_string(col_letter)

            # Use proper range string format for conditional formatting
            range_str = f"{col_letter}2:{col_letter}{max_row}"

            try:
                # Green for valid species (found in database)
                # Check if species is in scientific_name (col B) or local_name (col C) of SPECIES_DATABASE
                green_formula = f'OR(ISNUMBER(MATCH({col_letter}2,SPECIES_DATABASE!$B$2:$B${species_count+1},0)),ISNUMBER(MATCH({col_letter}2,SPECIES_DATABASE!$C$2:$C${species_count+1},0)))'
                green_rule = FormulaRule(formula=[green_formula], fill=green_fill, font=green_font)
                ws_data.conditional_formatting.add(range_str, green_rule)

                # Red for invalid species (not found in database)
                red_formula = f'AND(LEN({col_letter}2)>0,NOT(ISNUMBER(MATCH({col_letter}2,SPECIES_DATABASE!$B$2:$B${species_count+1},0))),NOT(ISNUMBER(MATCH({col_letter}2,SPECIES_DATABASE!$C$2:$C${species_count+1},0))))'
                red_rule = FormulaRule(formula=[red_formula], fill=red_fill, font=red_font)
                ws_data.conditional_formatting.add(range_str, red_rule)

                print(f"[Excel Validator] Added conditional formatting to {col_letter} ({species_col_name})")

            except Exception as e:
                print(f"[Excel Validator] Warning: Could not add conditional formatting to {species_col_name}: {e}")

    # ===== AUTO-CONVERSION HELPER COLUMNS =====
    # Add helper columns at the end that show scientific name conversions
    # Users can reference these to verify conversions

    print(f"[Excel Validator] Adding auto-conversion helper columns at end of sheet")

    # Get the last column number
    last_col_num = ws_data.max_column

    # Add helper columns for each species column
    helper_col_num = last_col_num + 1

    for species_col_name in species_columns:
        if species_col_name in col_map:
            col_letter = col_map[species_col_name]
            helper_col_letter = get_column_letter(helper_col_num)

            # Set header for helper column with informative name
            helper_header = f"✓ {species_col_name.replace('_scientific', '')}_CONVERTED"
            ws_data[f"{helper_col_letter}1"] = helper_header
            ws_data[f"{helper_col_letter}1"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue
            ws_data[f"{helper_col_letter}1"].font = Font(color="FFFFFF", bold=True, size=9)
            ws_data[f"{helper_col_letter}1"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Set column width
            ws_data.column_dimensions[helper_col_letter].width = 20

            # Add formula to convert local name to scientific name
            # Formula logic:
            # 1. If cell is empty, return empty
            # 2. If value is in scientific_name column, return as is (already correct)
            # 3. If value is in local_name column, lookup and return scientific_name
            # 4. Otherwise return the value as is with note

            for row in range(2, max_row + 1):
                cell_ref = f"{col_letter}{row}"
                formula = f'=IF(ISBLANK({cell_ref}),"",IF(ISNUMBER(MATCH({cell_ref},SPECIES_DATABASE!$B$2:$B${species_count+1},0)),{cell_ref},IF(ISNUMBER(MATCH({cell_ref},SPECIES_DATABASE!$C$2:$C${species_count+1},0)),INDEX(SPECIES_DATABASE!$B$2:$B${species_count+1},MATCH({cell_ref},SPECIES_DATABASE!$C$2:$C${species_count+1},0)),{cell_ref}&" (⚠ not in DB)")))'
                ws_data[f"{helper_col_letter}{row}"] = formula

            print(f"[Excel Validator] Added auto-conversion helper column for {species_col_name} at column {helper_col_letter}")

            helper_col_num += 1

    # ===== VALIDATION_GUIDE SHEET =====
    ws_guide = wb.create_sheet("VALIDATION_GUIDE")

    guide_content = [
        ['Field Data Validation Guide', ''],
        ['', ''],
        ['SPECIES ENTRY SYSTEM - AUTO-COMPLETE (NEW!):', ''],
        ['How to enter species (just like surname column!):', ''],
        ['1. Click on species column cell', 'Start typing scientific OR local name'],
        ['2. As you type, list filters automatically', 'E.g., type "Sal" → see "Sal" in filtered list'],
        ['3. Select from filtered list or continue typing', 'E.g., type "Shorea" → see all Shorea species'],
        ['4. Press Enter to confirm', 'Cell turns green if valid, red if not in database'],
        ['', ''],
        ['Auto-conversion helper columns:', ''],
        ['- Enter local name (e.g., "Sal")', 'Check blue column at end to see "Shorea robusta"'],
        ['- Enter scientific name (e.g., "Shorea robusta")', 'Helper column shows same name'],
        ['- Green highlight = valid species', 'Red highlight = not found in database'],
        ['', ''],
        ['HELPER COLUMNS (Blue headers at end):', ''],
        ['✓ regen_CONVERTED, etc.', 'Shows scientific name for any entry (local or scientific)'],
        ['Purpose', 'Verify that local names are correctly matched to scientific names'],
        ['Usage', 'Reference these columns to check conversions before import'],
        ['', ''],
        ['NUMERIC VALIDATION (Strict - Excel will reject invalid values):', ''],
        ['Column', 'Allowed Range'],
        ['regen_dbh', '1 to 3.99 cm'],
        ['regen_count', '1 to 15'],
        ['sapling_dbh_cm', '4 to 9.99 cm'],
        ['sapling_count', '1 to 10'],
        ['pole_dbh_cm', '10 to 29.99 cm'],
        ['pole_height_m', '5 to 25 meters'],
        ['pole_class', '1,2,3,4,A,B,C,D,a,b,c,d,i,ii,iii,iv,I,II,III,IV'],
        ['tree_dbh_cm', '30 to 200 cm'],
        ['tree_height_m', '7 to 50 meters'],
        ['tree_class', '1,2,3,4,A,B,C,D,a,b,c,d,i,ii,iii,iv,I,II,III,IV'],
        ['firewood_kg_per_100sqm_per_year', '0 to 500 kg/100sqm/year'],
        ['grass_kg_per_100sqm_per_year', '0 to 500 kg/100sqm/year'],
        ['bedding_material_kg_per_100sqm_per_year', '0 to 500 kg/100sqm/year'],
        ['ntfp_kg_per_100sqm_per_year', '0 to 100 kg/100sqm/year'],
        ['', ''],
        ['SIZE CLASS DEFINITIONS:', ''],
        ['Regeneration', 'DBH: 1-3.99 cm'],
        ['Sapling', 'DBH: 4-9.99 cm'],
        ['Pole', 'DBH: 10-29.99 cm'],
        ['Tree', 'DBH: 30-200 cm'],
        ['', ''],
        ['TREE CLASS (Nepal Standard):', ''],
        ['1 or I', 'Large trees (DBH≥40cm or H≥16m)'],
        ['2 or II', 'Mid-large trees (DBH 25-40cm, H 12-16m)'],
        ['3 or III', 'Mid-small trees (DBH 15-25cm, H 8-12m)'],
        ['4 or IV', 'Small trees (DBH<15cm or H<8m)'],
        ['A,B,C,D or a,b,c,d', 'Alternative classification system'],
        ['', ''],
        ['SN COLUMNS:', ''],
        ['regen_sn, sapling_sn, pole_sn, tree_sn', 'Keep existing values (no validation)'],
        ['', ''],
        ['RESOURCE HARVEST COLUMNS (Per Sample Plot):', ''],
        ['firewood, grass, bedding_material, ntfp', 'Default values populated once per sample plot (first row only)'],
        ['Note:', 'Edit values only in the first row of each sample plot'],
        ['', ''],
        ['IMPORTANT:', ''],
        ['- Numeric columns: Excel will reject values outside range', ''],
        ['- Species columns: Dropdown + auto-conversion from local to scientific name', ''],
        ['- Check SPECIES_DATABASE sheet for complete species list', ''],
        ['- Helper columns (hidden) show scientific name conversions', ''],
        ['- You can leave cells blank if no data for that category', ''],
    ]

    for row_data in guide_content:
        ws_guide.append(row_data)

    ws_guide.column_dimensions['A'].width = 50
    ws_guide.column_dimensions['B'].width = 40

    bold_font = Font(bold=True, size=12)
    ws_guide['A1'].font = Font(bold=True, size=14)
    ws_guide['A3'].font = bold_font   # SPECIES ENTRY SYSTEM - AUTO-COMPLETE
    ws_guide['A15'].font = bold_font  # HELPER COLUMNS (adjusted)
    ws_guide['A20'].font = bold_font  # NUMERIC VALIDATION
    ws_guide['A35'].font = bold_font  # SIZE CLASS DEFINITIONS
    ws_guide['A41'].font = bold_font  # TREE CLASS
    ws_guide['A48'].font = bold_font  # SN COLUMNS
    ws_guide['A51'].font = bold_font  # RESOURCE HARVEST COLUMNS
    ws_guide['A55'].font = bold_font  # IMPORTANT

    # Define fill colors for the guide
    green_fill_guide = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill_guide = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill_guide = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Add color examples in guide (adjusted row numbers)
    ws_guide['A12'].fill = green_fill_guide  # Green highlight example
    ws_guide['A13'].fill = red_fill_guide   # Red highlight example

    # Save enhanced workbook
    print(f"[Excel Validator] Saving workbook to {excel_filepath}")
    wb.save(excel_filepath)
    print(f"[Excel Validator] Workbook saved successfully")

    # Calculate new file size
    file_size_mb = os.path.getsize(excel_filepath) / (1024 * 1024)

    print(f"[Excel Validator] Validation enhancement completed successfully!")
    print(f"[Excel Validator] File size: {file_size_mb:.2f} MB")
    print(f"[Excel Validator] Data validations in workbook: {len(ws_data.data_validations.dataValidation)}")

    return excel_filepath, file_size_mb
