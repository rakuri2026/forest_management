"""
Household Information API endpoints
Handles household survey data for community forest user groups
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from uuid import UUID
import io
import os
import csv
import tempfile
from datetime import datetime
from decimal import Decimal

from ..core.database import get_db
from ..utils.auth import get_current_user
from ..utils.number_format import normalize_nepali_digits
from ..models.user import User
from ..models.calculation import Calculation
from ..models.household_information import HouseholdInformation
from ..models.caste_classification import CasteClassification
from ..models.forest_committee import ForestUserCommittee, AdvisoryCommittee, FinancialCommittee
from ..schemas.household_info import (
    HouseholdInfoCreate,
    HouseholdInfoUpdate,
    HouseholdInfoResponse,
    HouseholdUploadResponse,
    HouseholdUploadValidation,
    HouseholdSummary,
    CasteClassificationResponse,
    SurnameSuggestion,
    TemplateDownloadOptions
)
from ..services.household_calculations import HouseholdCalculations
from ..services.forest_committee_service import ForestCommitteeValidation

router = APIRouter(prefix="/api/household")


# ============================================================================
# Template Download
# ============================================================================

@router.post("/calculations/{calculation_id}/template")
async def download_household_template(
    calculation_id: UUID,
    options: TemplateDownloadOptions,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Download Excel template for household information entry

    Args:
        calculation_id: UUID of the calculation (community forest)
        options: Template download options (land unit, include coordinates)

    Returns:
        Excel file with headers, data validation, and formulas
    """
    # Verify calculation exists and belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
        from openpyxl.worksheet.datavalidation import DataValidation

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Household Information"

        # Color definitions
        auto_calc_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
        manual_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green

        # Define columns (Nepali headers)
        columns = [
            ("A", "घर नंवर", "House No"),
            ("B", "थर", "Surname"),
            ("C", "घरमुली पुरूष", "Male Head"),
            ("D", "घरमुली महिला", "Female Head"),
            ("E", "टोल ठेगाना", "Address"),
        ]

        # Add coordinate columns if requested
        col_offset = 5
        if options.include_coordinates:
            columns.extend([
                ("F", "अक्षाँस", "Latitude"),
                ("G", "देशान्तर", "Longitude"),
            ])
            col_offset = 7

        # Continue with other columns
        columns.extend([
            (chr(65 + col_offset), "महिला संख्या", "Female Count"),
            (chr(65 + col_offset + 1), "पुरूष संख्या", "Male Count"),
            (chr(65 + col_offset + 2), f"जमिन ({options.land_unit})", f"Land ({options.land_unit})"),
            (chr(65 + col_offset + 3), "पेशा (वनमा आश्रीत)", "Forest Occupation (Yes/No)"),
            (chr(65 + col_offset + 4), "पेशा (अन्य)", "Other Occupation (Yes/No)"),
            (chr(65 + col_offset + 5), "गाइ गोरू", "Cow/Ox Count"),
            (chr(65 + col_offset + 6), "भैसी राँगा", "Buffalo Count"),
            (chr(65 + col_offset + 7), "बाख्रा भेडा", "Goat/Sheep Count"),
            (chr(65 + col_offset + 8), "बन पैदाबारको माग (काठ क्यू.फि.)", "Timber (cft)"),
            (chr(65 + col_offset + 9), "बन पैदाबारको माग (पोल)", "Poles"),
            (chr(65 + col_offset + 10), "बन पैदाबारको माग (दाउरा) भारी", "Firewood (Bhari)"),
            (chr(65 + col_offset + 11), "बन पैदाबारको माग (घाँस) भारी", "Grass (Bhari)"),
            (chr(65 + col_offset + 12), "बन पैदाबारको माग (सोत्तर) भारी", "Bedding (Bhari)"),
            (chr(65 + col_offset + 13), "जातिय वर्गिकरण", "Caste Classification"),
            (chr(65 + col_offset + 14), "अन्य समूहमा सदस्यता", "Other Group (Yes/No)"),
            (chr(65 + col_offset + 15), "सम्पन्नताको स्तर", "Prosperity Level"),
            (chr(65 + col_offset + 16), "कैफियत", "Remarks"),
        ])

        # Style definitions
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(italic=True, size=9)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Row 1: Nepali headers
        for col_letter, nepali_text, _ in columns:
            cell = ws[f"{col_letter}1"]
            cell.value = nepali_text
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Row 2: English descriptions
        for col_letter, _, english_text in columns:
            cell = ws[f"{col_letter}2"]
            cell.value = english_text
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Define column letters for formulas (MUST BE BEFORE THEY ARE USED)
        house_no_col = "A"
        surname_col = "B"
        firewood_col = chr(65 + col_offset + 10)
        grass_col = chr(65 + col_offset + 11)
        bedding_col = chr(65 + col_offset + 12)
        female_col = chr(65 + col_offset)
        male_col = chr(65 + col_offset + 1)
        land_col = chr(65 + col_offset + 2)
        occupation_col = chr(65 + col_offset + 3)
        other_occ_col = chr(65 + col_offset + 4)
        cow_ox_col = chr(65 + col_offset + 5)
        buffalo_col = chr(65 + col_offset + 6)
        goat_sheep_col = chr(65 + col_offset + 7)
        timber_col = chr(65 + col_offset + 8)
        pole_col = chr(65 + col_offset + 9)
        caste_col = chr(65 + col_offset + 13)
        other_group_col = chr(65 + col_offset + 14)
        prosperity_col = chr(65 + col_offset + 15)

        # Set column widths and apply color coding with borders
        # Auto-calculated columns: House No, Firewood, Grass, Bedding, Caste, Prosperity
        auto_calc_cols = [house_no_col, firewood_col, grass_col, bedding_col, caste_col, prosperity_col]

        # Border styles
        dark_green_border = Border(
            left=Side(style='thin', color='00704E'),
            right=Side(style='thin', color='00704E'),
            top=Side(style='thin', color='00704E'),
            bottom=Side(style='thin', color='00704E')
        )
        dark_red_border = Border(
            left=Side(style='thin', color='8B0000'),
            right=Side(style='thin', color='8B0000'),
            top=Side(style='thin', color='8B0000'),
            bottom=Side(style='thin', color='8B0000')
        )

        for col_letter, _, _ in columns:
            ws.column_dimensions[col_letter].width = 15

            # Apply color coding and borders to data rows (3-501)
            for row in range(3, 501):
                cell = ws[f"{col_letter}{row}"]
                if col_letter in auto_calc_cols:
                    cell.fill = auto_calc_fill  # Light red for auto-calculated
                    cell.border = dark_red_border  # Dark red border
                else:
                    cell.fill = manual_fill  # Light green for manual input
                    cell.border = dark_green_border  # Dark green border

        # Add data validation for occupation dropdowns
        # पेशा (वनमा आश्रित) dropdown
        forest_occ_dv = DataValidation(
            type="list",
            formula1='"उच्च (NTFP संकलन),उच्च (NTFP/दाउरा बिक्री),उच्च (पूर्ण आश्रित),छैन,सामान्य (काठ/दाउरा),सामान्य (घाँस/दाउरा)"',
            allow_blank=True
        )
        forest_occ_dv.error = 'कृपया सूचीबाट चयन गर्नुहोस्'
        forest_occ_dv.errorTitle = 'अमान्य इनपुट'
        ws.add_data_validation(forest_occ_dv)
        forest_occ_dv.add(f"{occupation_col}3:{occupation_col}500")

        # पेशा (अन्य) dropdown
        other_occ_dv = DataValidation(
            type="list",
            formula1='"कृषि मात्र (Agriculture only),छैन (None / Unable to work),दैनिक ज्यालादारी (Wage labor),वैदेशिक रोजगार (Remittance),व्यापार / सरकारी जागिर"',
            allow_blank=True
        )
        other_occ_dv.error = 'कृपया सूचीबाट चयन गर्नुहोस्'
        other_occ_dv.errorTitle = 'अमान्य इनपुट'
        ws.add_data_validation(other_occ_dv)
        other_occ_dv.add(f"{other_occ_col}3:{other_occ_col}500")

        # छ/छैन dropdown for other fields
        nepali_yes_no_dv = DataValidation(type="list", formula1='"छ,छैन"', allow_blank=True)
        nepali_yes_no_dv.error = 'कृपया छ वा छैन चयन गर्नुहोस्'
        nepali_yes_no_dv.errorTitle = 'अमान्य इनपुट'

        ws.add_data_validation(nepali_yes_no_dv)
        # Note: Other Group column now uses formula for default value, no dropdown needed

        # STRICT Integer validation for number fields - BLOCKS non-numeric input
        integer_dv = DataValidation(
            type="whole",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle='stop',  # STOP means user CANNOT enter invalid data
            error='कृपया सङ्ख्या मात्र प्रविष्ट गर्नुहोस् (केवल ०-९)\nPlease enter numbers only (0-9)',
            errorTitle='अमान्य इनपुट / Invalid Input',
            showInputMessage=True,
            promptTitle='सङ्ख्या प्रविष्ट गर्नुहोस्',
            prompt='केवल पूर्ण सङ्ख्या (०, १, २, ३...) प्रविष्ट गर्नुहोस्\nEnter whole numbers only (0, 1, 2, 3...)'
        )
        ws.add_data_validation(integer_dv)
        integer_dv.add(f"{female_col}3:{female_col}500")
        integer_dv.add(f"{male_col}3:{male_col}500")
        integer_dv.add(f"{cow_ox_col}3:{cow_ox_col}500")
        integer_dv.add(f"{buffalo_col}3:{buffalo_col}500")
        integer_dv.add(f"{goat_sheep_col}3:{goat_sheep_col}500")
        integer_dv.add(f"{pole_col}3:{pole_col}500")

        # STRICT Decimal validation for land area and timber - BLOCKS non-numeric input
        decimal_dv = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle='stop',  # STOP means user CANNOT enter invalid data
            error='कृपया दशमलव सङ्ख्या प्रविष्ट गर्नुहोस् (जस्तै: १.५, ०.७५)\nPlease enter decimal numbers (e.g., 1.5, 0.75)',
            errorTitle='अमान्य इनपुट / Invalid Input',
            showInputMessage=True,
            promptTitle='दशमलव सङ्ख्या',
            prompt='दशमलव सङ्ख्या प्रविष्ट गर्नुहोस् (जस्तै: १५.५, ०.७५)\nEnter decimal numbers (e.g., 15.5, 0.75)'
        )
        ws.add_data_validation(decimal_dv)
        decimal_dv.add(f"{land_col}3:{land_col}500")
        decimal_dv.add(f"{timber_col}3:{timber_col}500")

        # Create Parameters sheet for configurable calculation values
        ws_params = wb.create_sheet("Parameters", 0)  # Insert as first sheet
        ws_params.sheet_state = 'visible'

        # Add parameters with defaults
        param_data = [
            ["Calculation Parameters", "मापदण्ड"],
            ["", ""],
            ["Parameter", "Value", "Unit", "Description (English)", "विवरण (नेपाली)"],
            ["", "", "", "", ""],
            ["FIREWOOD CALCULATION", "", "", "Adjust based on geography (Terai/Hills/Mountain)", "भूगोल अनुसार समायोजन गर्नुहोस् (तराई/पहाड/हिमाल)"],
            ["Person Daily Need", 250, "gm/day", "Daily firewood per person", "प्रति व्यक्ति दैनिक दाउरा"],
            ["Livestock Daily Need", 600, "gm/day", "Daily firewood per cattle", "प्रति गाईवस्तु दैनिक दाउरा"],
            ["Bhari Weight", 30, "kg", "Weight of one bhari (Forest Regulation 2079)", "एक भारीको तौल (वन नियमावली २०७९)"],
            ["", "", "", "", ""],
            ["GRASS CALCULATION", "", "", "Adjust based on pasture availability", "चरण क्षेत्र अनुसार समायोजन गर्नुहोस्"],
            ["Cow/Ox Daily Need", 20, "kg/day", "Daily grass per cow/ox", "प्रति गाई/गोरु दैनिक घाँस"],
            ["Buffalo Daily Need", 30, "kg/day", "Daily grass per buffalo", "प्रति भैंसी दैनिक घाँस"],
            ["Goat/Sheep Daily Need", 5, "kg/day", "Daily grass per goat/sheep", "प्रति बाख्रा/भेडा दैनिक घाँस"],
            ["Days Per Year", 365, "days", "Calculation period", "गणना अवधि"],
            ["Bhari Weight", 30, "kg", "Weight of one bhari (Forest Regulation 2079)", "एक भारीको तौल (वन नियमावली २०७९)"],
            ["", "", "", "", ""],
            ["BEDDING MATERIAL CALCULATION", "", "", "Adjust based on culture and farming", "संस्कृति र खेती अनुसार समायोजन गर्नुहोस्"],
            ["Cattle Daily Need", 10, "kg/day", "Daily bedding per cattle", "प्रति गाईवस्तु दैनिक सोत्तर"],
            ["Days Per Year", 365, "days", "Calculation period", "गणना अवधि"],
            ["Bhari Weight", 30, "kg", "Weight of one bhari (Forest Regulation 2079)", "एक भारीको तौल (वन नियमावली २०७९)"],
            ["", "", "", "", ""],
            ["DEFAULT VALUES", "", "", "", ""],
            ["Timber Demand", 5, "cft/year", "Default timber per household", "प्रति घरधुरी काठ"],
            ["Pole Demand", 5, "nos/year", "Default poles per household", "प्रति घरधुरी पोल"],
            ["", "", "", "", ""],
            ["PROSPERITY CLASSIFICATION THRESHOLDS", "", "", "Adjust thresholds as needed", "आवश्यकता अनुसार परिवर्तन गर्नुहोस्"],
            ["सम्पन्न (Prosperous) - Land Min", 10, "ropani", "Minimum land for prosperous", "सम्पन्नको न्यूनतम जमिन"],
            ["सम्पन्न (Prosperous) - Land Max", 999, "ropani", "Maximum land for prosperous", "सम्पन्नको अधिकतम जमिन"],
            ["सम्पन्न (Prosperous) - Cattle Min", 2, "nos", "Minimum cattle for prosperous", "सम्पन्नको न्यूनतम गाईवस्तु"],
            ["सम्पन्न (Prosperous) - Occupation", "छैन", "text", "Forest occupation = छैन for prosperous", "सम्पन्नको वन पेशा"],
            ["", "", "", "", ""],
            ["मध्यम (Medium) - Land Min", 5, "ropani", "Minimum land for medium", "मध्यमको न्यूनतम जमिन"],
            ["मध्यम (Medium) - Land Max", 20, "ropani", "Maximum land for medium", "मध्यमको अधिकतम जमिन"],
            ["मध्यम (Medium) - Cattle Min", 1, "nos", "Minimum cattle for medium", "मध्यमको न्यूनतम गाईवस्तु"],
            ["", "", "", "", ""],
            ["विपन्न (Poor) - Land Min", 1, "ropani", "Minimum land for poor", "विपन्नको न्यूनतम जमिन"],
            ["विपन्न (Poor) - Land Max", 10, "ropani", "Maximum land for poor", "विपन्नको अधिकतम जमिन"],
            ["विपन्न (Poor) - Forest Occupation", "उच्च,सामान्य", "text", "Forest occupation for poor", "विपन्नको वन पेशा"],
            ["", "", "", "", ""],
            ["अति विपन्न (Very Poor) - Land Max", 5, "ropani", "Maximum land for very poor", "अति विपन्नको अधिकतम जमिन"],
            ["अति विपन्न (Very Poor) - Cattle Max", 2, "nos", "Maximum cattle for very poor", "अति विपन्नको अधिकतम गाईवस्तु"],
            ["अति विपन्न (Very Poor) - Occupation", "उच्च", "text", "Forest occupation for very poor", "अति विपन्नको वन पेशा"],
        ]

        for row_idx, row_data in enumerate(param_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_params.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                elif row_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif row_idx in [5, 10, 17, 22]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                elif col_idx == 2 and row_idx > 3 and value and isinstance(value, (int, float)):
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        ws_params.column_dimensions['A'].width = 25
        ws_params.column_dimensions['B'].width = 12
        ws_params.column_dimensions['C'].width = 12
        ws_params.column_dimensions['D'].width = 35
        ws_params.column_dimensions['E'].width = 35

        # Protect Parameters sheet - only allow editing values in column B
        # Lock ALL cells first
        for row in ws_params.iter_rows(min_row=1, max_row=30, min_col=1, max_col=5):
            for cell in row:
                cell.protection = Protection(locked=True)

        # Unlock ONLY the editable parameter values (column B, specific rows)
        # B6:B8 (Firewood parameters)
        for row_idx in [6, 7, 8]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B11:B15 (Grass parameters)
        for row_idx in [11, 12, 13, 14, 15]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B18:B20 (Bedding parameters)
        for row_idx in [18, 19, 20]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B23:B24 (Default values)
        for row_idx in [23, 24]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B27:B30 (सम्पन्न prosperity parameters)
        for row_idx in [27, 28, 29, 30]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B32:B34 (मध्यम prosperity parameters)
        for row_idx in [32, 33, 34]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B36:B37 (विपन्न prosperity parameters)
        for row_idx in [36, 37]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # B40:B41 (अति विपन्न prosperity parameters)
        for row_idx in [40, 41]:
            ws_params.cell(row=row_idx, column=2).protection = Protection(locked=False)

        # No sheet protection - only cell-level locks
        # Users can protect manually if needed
        # Cell locks still work: B6:B8, B11:B15, B18:B20, B23:B24 editable, rest locked

        # Apply formulas and defaults for rows 3-501
        for row in range(3, 501):
            # Auto-numbering: House number (increments if surname is filled)
            if row == 3:
                ws[f"{house_no_col}{row}"] = f'=IF({surname_col}{row}<>"",ROW()-2,"")'
            else:
                ws[f"{house_no_col}{row}"] = f'=IF({surname_col}{row}<>"",IF({surname_col}{row-1}<>"",{house_no_col}{row-1}+1,1),"")'

            # Caste classification VLOOKUP (supports both Nepali and English surnames)
            ws[f"{caste_col}{row}"] = f'=IFERROR(VLOOKUP({surname_col}{row},CasteDB!$A$2:$B$500,2,FALSE),IFERROR(VLOOKUP({surname_col}{row},CasteDB!$C$2:$D$500,2,FALSE),""))'

            # Timber default from Parameters sheet (only if House No is not blank)
            ws[f"{timber_col}{row}"] = f'=IF({house_no_col}{row}<>"",Parameters!$B$23,"")'

            # Pole default from Parameters sheet (only if House No is not blank)
            ws[f"{pole_col}{row}"] = f'=IF({house_no_col}{row}<>"",Parameters!$B$24,"")'

            # Firewood formula using parameters from Parameters sheet
            # ((female + male) * person_daily + (cow + buffalo) * livestock_daily) / bhari_weight
            ws[f"{firewood_col}{row}"] = f'=IF(AND({female_col}{row}<>"",{male_col}{row}<>""),ROUND((({female_col}{row}+{male_col}{row})*Parameters!$B$6+({cow_ox_col}{row}+{buffalo_col}{row})*Parameters!$B$7)/Parameters!$B$8,2),"")'

            # Grass formula using parameters
            # ((cow * cow_daily + buffalo * buffalo_daily + goat_sheep * goat_daily) * days) / bhari_weight
            ws[f"{grass_col}{row}"] = f'=IF({cow_ox_col}{row}<>"",ROUND((({cow_ox_col}{row}*Parameters!$B$11+{buffalo_col}{row}*Parameters!$B$12+{goat_sheep_col}{row}*Parameters!$B$13)*Parameters!$B$14)/Parameters!$B$15,2),"")'

            # Bedding formula using parameters
            # ((cow + buffalo) * cattle_daily * days) / bhari_weight
            ws[f"{bedding_col}{row}"] = f'=IF({cow_ox_col}{row}<>"",ROUND((({cow_ox_col}{row}+{buffalo_col}{row})*Parameters!$B$18*Parameters!$B$19)/Parameters!$B$20,2),"")'

            # Prosperity level auto-calculation using Parameters sheet
            # Uses configurable thresholds from Parameters sheet
            # B27: सम्पन्न Land Min, B28: सम्पन्न Land Max, B29: सम्पन्न Cattle Min, B30: सम्पन्न Occupation
            # B32: मध्यम Land Min, B33: मध्यम Land Max, B34: मध्यम Cattle Min
            # B36: विपन्न Land Min, B37: विपन्न Land Max
            # B40: अति विपन्न Land Max, B41: अति विपन्न Cattle Max
            prosperity_formula = f'''=IF({land_col}{row}="","",
IF(AND({land_col}{row}>=Parameters!$B$27,{occupation_col}{row}=Parameters!$B$30,{cow_ox_col}{row}+{buffalo_col}{row}>=Parameters!$B$29),"सम्पन्न",
IF(AND({land_col}{row}>=Parameters!$B$32,{land_col}{row}<=Parameters!$B$33,{cow_ox_col}{row}+{buffalo_col}{row}>=Parameters!$B$34),"मध्यम",
IF(AND({land_col}{row}>=Parameters!$B$36,{land_col}{row}<=Parameters!$B$37),"विपन्न",
IF(OR({land_col}{row}<=Parameters!$B$40,{cow_ox_col}{row}+{buffalo_col}{row}<=Parameters!$B$41),"अति विपन्न","मध्यम")))))'''
            ws[f"{prosperity_col}{row}"] = prosperity_formula.replace('\n', '')

            # Other Group (अन्य समूहमा सदस्यता) - default to 'छैन' only if House No is not blank
            ws[f"{other_group_col}{row}"] = f'=IF({house_no_col}{row}<>"","छैन","")'

        # Load caste classification data and create hidden sheet
        caste_csv_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
            "testData", "households_information", "caste_classification.csv"
        )

        ws_caste = wb.create_sheet("CasteDB")
        ws_caste.sheet_state = 'hidden'  # Hide the sheet

        if os.path.exists(caste_csv_path):
            with open(caste_csv_path, 'r', encoding='utf-8-sig') as f:
                csv_reader = csv.reader(f)
                next(csv_reader)  # Skip header row

                # Add headers to hidden sheet (5 columns: Nepali, English, Combined surnames for dropdown)
                ws_caste['A1'] = 'surname_ne'
                ws_caste['B1'] = 'classification_ne'
                ws_caste['C1'] = 'surname_en'
                ws_caste['D1'] = 'classification_ne_copy'
                ws_caste['E1'] = 'surname_combined'

                # Collect all unique surnames for dropdown
                all_surnames = set()
                
                # Load data (both Nepali and English surnames)
                row_idx = 2
                for row in csv_reader:
                    if len(row) >= 6:
                        classification_ne = row[0].strip()
                        surname_ne = row[2].strip()
                        surname_en = row[5].strip()

                        # Columns A-B: Nepali surname lookup
                        ws_caste[f'A{row_idx}'] = surname_ne
                        ws_caste[f'B{row_idx}'] = classification_ne

                        # Columns C-D: English surname lookup
                        ws_caste[f'C{row_idx}'] = surname_en
                        ws_caste[f'D{row_idx}'] = classification_ne

                        # Collect only Nepali surnames for dropdown (user requirement)
                        if surname_ne:
                            all_surnames.add(surname_ne)
                        row_idx += 1

                # Store the lookup data end row
                lookup_end_row = row_idx - 1

                # Add combined surname list in column E (separate rows for each surname)
                sorted_surnames = sorted(all_surnames, key=lambda x: x.lower())
                dropdown_start_row = row_idx
                for idx, surname in enumerate(sorted_surnames):
                    ws_caste.cell(row=row_idx + idx, column=5, value=surname)
                
                dropdown_end_row = row_idx + len(sorted_surnames) - 1

        # Create named range for surname auto-suggestion using combined column E
        # Use column E which has combined surnames for the dropdown
        surname_formula = f'CasteDB!$E${dropdown_start_row}:$E${dropdown_end_row}'
        
        # Add data validation for surname column (column B) with auto-suggestion
        surname_dv = DataValidation(
            type="list",
            formula1=surname_formula,
            allow_blank=True,
            showDropDown=False
        )
        surname_dv.error = 'कृपया सूचीबाट चयन गर्नुहोस्'
        surname_dv.errorTitle = 'अमान्य इनपुट'
        surname_dv.prompt = 'थर छान्नुहोस् वा मैले लेखेको थर प्रयोग गर्नुहोस्'
        surname_dv.promptTitle = 'थर / Surname'
        
        ws.add_data_validation(surname_dv)
        surname_dv.add(f"{surname_col}3:{surname_col}500")

        # Freeze header rows
        ws.freeze_panes = "A3"

        # ═══════════════════════════════════════════════════════════════════
        # SHEET PROTECTION STRATEGY
        # ═══════════════════════════════════════════════════════════════════
        # Household Information sheet:
        #   - Headers (rows 1-2): LOCKED (cannot edit)
        #   - Auto-calculated columns (red background): LOCKED (read-only)
        #   - Manual entry columns (green background), rows 3-500: UNLOCKED (editable)
        #   - All cells beyond row 500: LOCKED
        #   → Result: Only A3:Z500 (manual entry columns) are editable
        # ═══════════════════════════════════════════════════════════════════

        # Get manual entry columns (not auto-calculated)
        manual_entry_cols = [col_letter for col_letter, _, _ in columns if col_letter not in auto_calc_cols]

        # Lock ALL cells in the Household Information sheet (A1:XFD1048576)
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=1, max_col=26):
            for cell in row:
                cell.protection = Protection(locked=True)

        # Explicitly unlock A3:Z500 (editable range)
        for col_idx in range(1, 27):
            for row_idx in range(3, 501):
                ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)

        # No sheet protection - only cell-level locks
        # Users can unprotect manually if needed
        # Cell locks still work: A3:Z500 editable, rest locked

        # Convert data range to Excel Table named "user_info"
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Get the last column letter
        last_col_letter = columns[-1][0]

        # Create table spanning from A1 to last_column + 500 rows
        table = Table(displayName="user_info", ref=f"A1:{last_col_letter}500")
        table_style = TableStyleInfo(
            name="TableStyleLight9",  # Green-themed table style
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = table_style
        # Disable autoFilter on the table itself
        table.autoFilter = None
        
        # Disable autoFilter on worksheet BEFORE adding table
        ws.auto_filter.ref = None
        ws.add_table(table)
        
        # Ensure autoFilter stays disabled after table is added
        ws.auto_filter.ref = None

        # Set print area to define the usable data range
        # This guides Excel and users to the intended data area
        ws.print_area = f'A1:{last_col_letter}500'

        # Set tab color to indicate this is the main data sheet
        ws.sheet_properties.tabColor = "70AD47"  # Green color

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        ws_instructions.column_dimensions['A'].width = 80

        instructions = [
            "Household Information Form - Instructions",
            "घरधुरी जानकारी फाराम - निर्देशन",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "IMPORTANT: CONFIGURABLE PARAMETERS",
            "महत्वपूर्ण: समायोज्य मापदण्डहरू",
            "═══════════════════════════════════════════════════════════════════",
            "",
            "The 'Parameters' sheet contains calculation values that you can adjust",
            "based on your geography (Terai/Hills/Mountains) and local conditions:",
            "",
            "'Parameters' पानामा गणना मानहरू छन् जुन तपाईंको भूगोल (तराई/पहाड/हिमाल)",
            "र स्थानीय अवस्था अनुसार समायोजन गर्न सक्नुहुन्छ:",
            "",
            "FIREWOOD CALCULATION (दाउरा गणना):",
            "  • Person Daily Need (प्रति व्यक्ति दैनिक): 250 gm/day (default)",
            "  • Livestock Daily Need (प्रति गाईवस्तु दैनिक): 600 gm/day (default)",
            "  • Adjust for climate: Higher in cold regions, lower in warm regions",
            "    जलवायु अनुसार: चिसो क्षेत्रमा बढी, न्यानो क्षेत्रमा कम",
            "",
            "GRASS CALCULATION (घाँस गणना):",
            "  • Cow/Ox Daily Need: 20 kg/day (default)",
            "  • Buffalo Daily Need: 30 kg/day (default)",
            "  • Goat/Sheep Daily Need: 5 kg/day (default)",
            "  • Adjust for pasture availability and grazing practices",
            "    चरण क्षेत्र र चराइ अभ्यास अनुसार समायोजन गर्नुहोस्",
            "",
            "BEDDING MATERIAL (सोत्तर सामग्री):",
            "  • Cattle Daily Need: 10 kg/day (default)",
            "  • Adjust based on farming practices and culture",
            "    खेती अभ्यास र संस्कृति अनुसार समायोजन गर्नुहोस्",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "COLOR CODING / रङ कोडिङ:",
            "═══════════════════════════════════════════════════════════════════",
            "  🔴 Light Red + Dark Red Border = Auto-calculated (uses Parameters)",
            "     हल्का रातो + गाढा रातो बोर्डर = स्वतः गणना (Parameters प्रयोग गर्छ)",
            "",
            "  🟢 Light Green + Dark Green Border = Manual entry required",
            "     हल्का हरियो + गाढा हरियो बोर्डर = म्यानुअल प्रविष्टि आवश्यक",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "EXCEL TABLE FORMAT",
            "═══════════════════════════════════════════════════════════════════",
            "  • Data is formatted as Excel Table named 'user_info'",
            "  • Use built-in filtering and sorting features",
            "  • Table automatically expands when you add data",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "SHEET PROTECTION / पाना सुरक्षा",
            "═══════════════════════════════════════════════════════════════════",
            "WHAT IS PROTECTED:",
            "  • Header rows (1-2) - LOCKED (cannot edit or delete)",
            "    शीर्षक पङ्क्तिहरू (१-२) - लक (सम्पादन वा मेटाउन सकिँदैन)",
            "",
            "  • Auto-calculated columns (red background) - LOCKED (read-only)",
            "    स्वतः गणना स्तम्भहरू (रातो पृष्ठभूमि) - लक (पढ्न मात्र)",
            "",
            "WHAT YOU CAN EDIT:",
            "  • Manual entry columns (green background) - UNLOCKED (rows 3-500)",
            "    म्यानुअल प्रविष्टि स्तम्भहरू (हरियो पृष्ठभूमि) - अनलक (पङ्क्ति ३-५०००)",
            "",
            "  • Up to 500 households (rows 3-500) are ready for data entry",
            "    500 घरधुरी सम्म (पङ्क्ति 3-500) डाटा प्रविष्टि को लागि तयार छ",
            "",
            "  • Rows beyond 500 and columns beyond Z are protected",
            "    पङ्क्ति 500 पछि र स्तम्भ Z पछि सुरक्षित छन्",
            "",
            "FOR MORE THAN 500 HOUSEHOLDS:",
            "  1. Review tab → Unprotect Sheet (no password needed)",
            "     Review ट्याब → Unprotect Sheet (पासवर्ड आवश्यक छैन)",
            "",
            "  2. Enter data in additional rows (beyond row 500)",
            "     थप पङ्क्तिहरूमा डाटा प्रविष्ट गर्नुहोस् (पङ्क्ति 500 पछि)",
            "",
            "  3. (Optional) Re-protect: Review → Protect Sheet",
            "     (वैकल्पिक) पुन: सुरक्षित: Review → Protect Sheet",
            "",
            "DATA VALIDATION (STRICTLY ENFORCED):",
            "  • Number fields ONLY accept numeric input (0-9)",
            "    सङ्ख्या क्षेत्रहरूले केवल सङ्ख्यात्मक इनपुट (०-९) स्वीकार गर्छन्",
            "",
            "  • Text (strings) will be REJECTED in number columns",
            "    सङ्ख्या स्तम्भहरूमा पाठ (strings) अस्वीकार गरिनेछ",
            "",
            "  • Error message will show if invalid data is entered",
            "    अमान्य डाटा प्रविष्ट गरेमा त्रुटि सन्देश देखिनेछ",
            "",
            "OTHER SHEETS:",
            "  • Parameters: Only column B values editable",
            "    Parameters: केवल स्तम्भ B मानहरू सम्पादन योग्य",
            "",
            "  • Instructions: Fully protected (read-only)",
            "    Instructions: पूर्ण सुरक्षित (पढ्न मात्र)",
            "",
            "  • CasteDB: Hidden (lookup data)",
            "    CasteDB: लुकाइएको (लुकअप डाटा)",
            "",
            "  • वन उपभोक्ता समिति: Forest User Committee (max 15 members)",
            "    Main committee with position, gender, caste validation",
            "",
            "  • सल्लाहाकार समिति: Advisory Committee (max 10 members, optional)",
            "    Advisory committee information",
            "",
            "  • आर्थिक समिति: Financial Committee (max 10 members, optional)",
            "    Financial committee information",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "STEP-BY-STEP GUIDE",
            "═══════════════════════════════════════════════════════════════════",
            "",
            "1. ADJUST PARAMETERS (Optional - if your area differs from defaults)",
            "   मापदण्ड समायोजन (वैकल्पिक - यदि तपाईंको क्षेत्र पूर्वनिर्धारितबाट फरक छ)",
            "   → Go to 'Parameters' sheet",
            "   → Change values in column B based on your local conditions",
            "   → Calculations will update automatically",
            "",
            "2. FILL HOUSEHOLD DATA (Starting from Row 3)",
            "   घरधुरी डाटा भर्नुहोस् (पङ्क्ति ३ देखि)",
            "",
            "   AUTO-CALCULATED FIELDS (Light Red - Do Not Edit):",
            "   स्वतः गणना हुने क्षेत्रहरू (हल्का रातो - सम्पादन नगर्नुहोस्):",
            "",
            "   a) House Number (घर नंवर): Auto-numbered when surname entered",
            "      थर प्रविष्ट गर्दा स्वतः नम्बर हुन्छ",
            "",
            "   b) Caste Classification (जातिय वर्गिकरण):",
            "      • Auto-filled from database when surname is entered",
            "      • Supports both Nepali and English surnames",
            "      • थर प्रविष्ट गर्दा डाटाबेसबाट स्वतः भरिन्छ",
            "      • नेपाली र अङ्ग्रेजी दुवै थर समर्थित",
            "",
            "   c) Timber & Poles: Uses default from Parameters sheet",
            "      काठ र पोल: Parameters पानाबाट पूर्वनिर्धारित प्रयोग गर्छ",
            "",
            "   d) Firewood (दाउरा): Uses Parameters sheet values",
            "      Formula: ((Female+Male)×PersonDaily + (Cow+Buffalo)×LivestockDaily) ÷ Bhari",
            "",
            "   e) Grass (घाँस): Uses Parameters sheet values",
            "      Formula: ((Cow×CowDaily + Buffalo×BuffaloDaily + Goat×GoatDaily)×Days) ÷ Bhari",
            "",
            "   f) Bedding (सोत्तर): Uses Parameters sheet values",
            "      Formula: ((Cow+Buffalo)×CattleDaily×Days) ÷ Bhari",
            "",
            "   g) Prosperity Level (सम्पन्नताको स्तर):",
            "      • सम्पन्न: >15 ropani, >5 cattle, non-forest occupation",
            "      • मध्यम: 5-15 ropani, 2-4 cattle",
            "      • विपन्न: 1-5 ropani, 2-3 cattle, forest occupation",
            "      • अति विपन्न: 0-1 ropani, 1-2 goats/sheep, forest occupation",
            "",
            "   MANUAL ENTRY FIELDS (Light Green):",
            "   म्यानुअल प्रविष्टि क्षेत्रहरू (हल्का हरियो):",
            "   • Surname, Names, Address, Population counts, Land area, Livestock",
            "   • Occupation fields: Select छ/छैन from dropdown",
            "   • All numeric fields have validation (only accept numbers)",
            "",
            "3. DATA VALIDATION:",
            "   • Dropdown fields: छ/छैन (for occupation and group membership)",
            "   • Number fields: Only accept whole numbers ≥ 0",
            "   • Decimal fields: Only accept decimal numbers ≥ 0 (land, timber)",
            "",
            "4. AFTER FILLING:",
            "   • Save the file (Ctrl+S or File → Save)",
            "   • Return to the system and upload",
            "   • System will validate and import all data",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "FILE INFORMATION",
            "═══════════════════════════════════════════════════════════════════",
            f"Land Unit: {options.land_unit}",
            f"Coordinates: {'Included' if options.include_coordinates else 'Not included'}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Forest: {calculation.forest_name or 'N/A'}",
            f"Capacity: Up to 500 households",
            "",
            "═══════════════════════════════════════════════════════════════════",
            "TIPS / सुझावहरू:",
            "═══════════════════════════════════════════════════════════════════",
            "• Formulas are pre-filled up to row 500",
            "  सूत्रहरू पङ्क्ति 500 सम्म पूर्व-भरिएका छन्",
            "",
            "• Use Excel's filter feature (Data → Filter) to sort/search",
            "  क्रमबद्ध/खोजी गर्न Excel को फिल्टर सुविधा प्रयोग गर्नुहोस्",
            "",
            "• Hidden sheet 'CasteDB' contains surname-caste mappings",
            "  लुकेको पाना 'CasteDB' मा थर-जाति म्यापिङ छ",
            "",
            "• For technical support, contact your system administrator",
            "  प्राविधिक सहयोगको लागि, आफ्नो प्रणाली प्रशासकलाई सम्पर्क गर्नुहोस्",
        ]

        for i, instruction in enumerate(instructions, start=1):
            cell = ws_instructions[f"A{i}"]
            cell.value = instruction
            if i == 1:
                cell.font = Font(bold=True, size=14)
            elif instruction.startswith("   "):
                cell.font = Font(size=10, italic=True)
            cell.alignment = Alignment(wrap_text=True)

        # Lock ALL cells in the used range of the Instructions sheet
        for row in ws_instructions.iter_rows(min_row=1, max_row=500, min_col=1, max_col=10):
            for cell in row:
                cell.protection = Protection(locked=True)

        # No sheet protection - cell locks still work
        # All cells in Instructions are locked, users can unprotect manually if needed

        # ====================================================================
        # CREATE FOREST USER COMMITTEE SHEETS
        # ====================================================================

        # Sheet 5: Main Forest User Committee (वन उपभोक्ता समिति)
        ws_main_committee = wb.create_sheet("वन उपभोक्ता समिति")
        ws_main_committee.column_dimensions['A'].width = 8   # सि.नं.
        ws_main_committee.column_dimensions['B'].width = 12  # लिङ्ग
        ws_main_committee.column_dimensions['C'].width = 18  # पद
        ws_main_committee.column_dimensions['D'].width = 18  # जातिय वर्ग
        ws_main_committee.column_dimensions['E'].width = 25  # नाम
        ws_main_committee.column_dimensions['F'].width = 25  # ठेगाना
        ws_main_committee.column_dimensions['G'].width = 15  # मोवाइल नंवर

        # Header row 1 (Nepali)
        main_headers_ne = ["सि.नं.", "लिङ्ग", "पद", "जातिय वर्ग", "नाम", "ठेगाना", "मोवाइल नंवर"]
        main_headers_en = ["S.No.", "Gender", "Position", "Caste Category", "Name", "Address", "Mobile"]

        for col_idx, header in enumerate(main_headers_ne, start=1):
            cell = ws_main_committee.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.font = Font(bold=True, color="FFFFFF")

        # Header row 2 (English)
        for col_idx, header in enumerate(main_headers_en, start=1):
            cell = ws_main_committee.cell(row=2, column=col_idx, value=header)
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Add data validation for gender (column B)
        gender_validation = DataValidation(type="list", formula1='"महिला,पुरूष"', allow_blank=False)
        gender_validation.error = "कृपया सूचीबाट छान्नुहोस्: महिला, पुरूष"
        gender_validation.errorTitle = "अवैध प्रविष्टि"
        ws_main_committee.add_data_validation(gender_validation)
        gender_validation.add(f"B3:B17")  # Rows 3-17 (max 15 members)

        # Add data validation for position (column C)
        position_validation = DataValidation(type="list",
            formula1='"अध्यक्ष,उपाध्यक्ष,कोषाध्यक्ष,सह कोषाध्यक्ष,सचिव,सह सचिव,सदस्य"',
            allow_blank=False)
        position_validation.error = "कृपया सूचीबाट छान्नुहोस्"
        position_validation.errorTitle = "अवैध पद"
        ws_main_committee.add_data_validation(position_validation)
        position_validation.add(f"C3:C17")

        # Add data validation for caste category (column D)
        caste_validation = DataValidation(type="list",
            formula1='"जनजाती,आदिवासी,दलित,सिमान्तकृत,अन्य"',
            allow_blank=False)
        caste_validation.error = "कृपया सूचीबाट छान्नुहोस्"
        caste_validation.errorTitle = "अवैध जातिय वर्ग"
        ws_main_committee.add_data_validation(caste_validation)
        caste_validation.add(f"D3:D17")

        # Fill in starting rows with serial numbers (1-15)
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        green_border = Border(
            left=Side(style='thin', color='70AD47'),
            right=Side(style='thin', color='70AD47'),
            top=Side(style='thin', color='70AD47'),
            bottom=Side(style='thin', color='70AD47')
        )

        for row_num in range(3, 18):  # Rows 3-17 (15 members)
            serial_no = row_num - 2
            ws_main_committee.cell(row=row_num, column=1, value=serial_no)

            # Apply green background and border to all editable cells
            for col_idx in range(1, 8):
                cell = ws_main_committee.cell(row=row_num, column=col_idx)
                cell.fill = green_fill
                cell.border = green_border
                cell.protection = Protection(locked=False)
                cell.alignment = Alignment(horizontal="center" if col_idx <= 4 else "left")

        # Lock header rows
        for row_num in [1, 2]:
            for col_idx in range(1, 8):
                ws_main_committee.cell(row=row_num, column=col_idx).protection = Protection(locked=True)

        # Sheet 6: Advisory Committee (सल्लाहाकार समिति)
        ws_advisory = wb.create_sheet("सल्लाहाकार समिति")
        ws_advisory.column_dimensions['A'].width = 8   # सि.नं.
        ws_advisory.column_dimensions['B'].width = 30  # नाम
        ws_advisory.column_dimensions['C'].width = 30  # ठेगाना
        ws_advisory.column_dimensions['D'].width = 15  # मोवाइल नंवर

        # Header rows
        advisory_headers_ne = ["सि.नं.", "नाम", "ठेगाना", "मोवाइल नंवर"]
        advisory_headers_en = ["S.No.", "Name", "Address", "Mobile"]

        for col_idx, header in enumerate(advisory_headers_ne, start=1):
            cell = ws_advisory.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, header in enumerate(advisory_headers_en, start=1):
            cell = ws_advisory.cell(row=2, column=col_idx, value=header)
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Fill rows 3-12 (max 10 members)
        for row_num in range(3, 13):
            serial_no = row_num - 2
            ws_advisory.cell(row=row_num, column=1, value=serial_no)

            for col_idx in range(1, 5):
                cell = ws_advisory.cell(row=row_num, column=col_idx)
                cell.fill = green_fill
                cell.border = green_border
                cell.protection = Protection(locked=False)
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")

        # Lock header rows
        for row_num in [1, 2]:
            for col_idx in range(1, 5):
                ws_advisory.cell(row=row_num, column=col_idx).protection = Protection(locked=True)

        # Sheet 7: Financial Committee (आर्थिक समिति)
        ws_financial = wb.create_sheet("आर्थिक समिति")
        ws_financial.column_dimensions['A'].width = 8   # सि.नं.
        ws_financial.column_dimensions['B'].width = 30  # नाम
        ws_financial.column_dimensions['C'].width = 30  # ठेगाना
        ws_financial.column_dimensions['D'].width = 15  # मोवाइल नंवर

        # Header rows
        financial_headers_ne = ["सि.नं.", "नाम", "ठेगाना", "मोवाइल नंवर"]
        financial_headers_en = ["S.No.", "Name", "Address", "Mobile"]

        for col_idx, header in enumerate(financial_headers_ne, start=1):
            cell = ws_financial.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_idx, header in enumerate(financial_headers_en, start=1):
            cell = ws_financial.cell(row=2, column=col_idx, value=header)
            cell.font = Font(italic=True, size=9, color="666666")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Fill rows 3-12 (max 10 members)
        for row_num in range(3, 13):
            serial_no = row_num - 2
            ws_financial.cell(row=row_num, column=1, value=serial_no)

            for col_idx in range(1, 5):
                cell = ws_financial.cell(row=row_num, column=col_idx)
                cell.fill = green_fill
                cell.border = green_border
                cell.protection = Protection(locked=False)
                cell.alignment = Alignment(horizontal="center" if col_idx == 1 else "left")

        # Lock header rows
        for row_num in [1, 2]:
            for col_idx in range(1, 5):
                ws_financial.cell(row=row_num, column=col_idx).protection = Protection(locked=True)

        # Save to temporary file
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        # Return file
        filename = f"household_template_{calculation.forest_name or 'data'}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return FileResponse(
            tmp_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
            background=None
        )

    except ImportError as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"openpyxl library not installed: {str(e)}"
        )
    except Exception as e:
        import traceback
        print("="*80)
        print("TEMPLATE GENERATION ERROR:")
        print("="*80)
        traceback.print_exc()
        print("="*80)
        raise HTTPException(status_code=500, detail=f"Error generating template: {str(e)}")


# ============================================================================
# Upload Excel File
# ============================================================================

@router.post("/calculations/{calculation_id}/upload", response_model=HouseholdUploadResponse)
async def upload_household_data(
    calculation_id: UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload filled Excel file with household data

    Args:
        calculation_id: UUID of the calculation
        file: Excel file with household data

    Returns:
        Upload result with validation errors/warnings
    """
    # Verify calculation exists and belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Note: Upsert mode enabled - uploads will update existing records (by house_no)
    # or insert new records. No need to delete existing data first.

    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx, .xls) or CSV file (.csv)")

    try:
        # Read uploaded file
        contents = await file.read()

        # Determine file type and parse accordingly
        is_csv = file.filename.endswith('.csv')

        if is_csv:
            # Parse CSV file
            import csv
            from io import StringIO

            # Decode bytes to string
            text_content = contents.decode('utf-8-sig')  # utf-8-sig handles BOM
            csv_reader = csv.reader(StringIO(text_content))
            rows = list(csv_reader)

            # Parse headers from row 1 (Nepali)
            headers = [cell.strip() for cell in rows[0] if cell]

            # Data rows start from row 3 (skip English headers in row 2)
            data_rows = rows[2:] if len(rows) > 2 else []

        else:
            # Parse Excel file
            from openpyxl import load_workbook

            # data_only=True: Read calculated values instead of formula strings
            # This prevents reading "=IF(...)" formulas as text
            wb = load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active

            # Parse headers from row 1 (Nepali)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            # Data rows from row 3 onwards (row 2 is English headers)
            data_rows = list(ws.iter_rows(min_row=3, values_only=True))

        # Expected header mappings (Nepali -> field name)
        header_map = {
            "घर नंवर": "house_no",
            "थर": "surname",
            "घरमुली पुरूष": "household_head_male",
            "घरमुली महिला": "household_head_female",
            "टोल ठेगाना": "address_tole",
            "अक्षाँस": "latitude",
            "देशान्तर": "longitude",
            "महिला संख्या": "female_count",
            "पुरूष संख्या": "male_count",
            "जमिन (ropani)": "land_area",
            "जमिन (kaththa)": "land_area",
            "पेशा (वनमा आश्रीत)": "forest_based_occupation",
            "पेशा (अन्य)": "other_occupation",
            "गाइ गोरू": "cow_ox_count",
            "भैसी राँगा": "buffalo_count",
            "बाख्रा भेडा": "goat_sheep_count",
            "बन पैदाबारको माग (काठ क्यू.फि.)": "timber_demand_cft",
            "बन पैदाबारको माग (पोल)": "pole_demand",
            "बन पैदाबारको माग (दाउरा) भारी": "firewood_demand_bhari",
            "बन पैदाबारको माग (घाँस) भारी": "grass_demand_bhari",
            "बन पैदाबारको माग (सोत्तर) भारी": "bedding_demand_bhari",
            "जातिय वर्गिकरण": "caste_classification_ne",
            "अन्य समूहमा सदस्यता": "other_group_membership",
            "सम्पन्नताको स्तर": "prosperity_level",
            "कैफियत": "remarks"
        }

        # Determine land unit from headers
        land_unit = None
        for header in headers:
            if "जमिन (ropani)" in header:
                land_unit = "ropani"
            elif "जमिन (kaththa)" in header:
                land_unit = "kaththa"

        # Parse data rows (starting from row 3, skip row 2 which is English headers)
        validations = []
        records_to_import = []

        for row_idx, row in enumerate(data_rows, start=3):
            # Skip empty rows
            if not any(row):
                continue

            validation = HouseholdUploadValidation(
                row_number=row_idx,
                is_valid=True,
                errors=[],
                warnings=[],
                data={}
            )

            # Parse row data
            row_data = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    field_name = header_map.get(header)

                    if field_name and cell_value is not None:
                        # Normalize Devanagari digits → Arabic for string values
                        if isinstance(cell_value, str):
                            cell_value = normalize_nepali_digits(cell_value)
                        # Convert to string and strip whitespace for CSV
                        if is_csv and isinstance(cell_value, str):
                            cell_value = cell_value.strip()
                            # Skip empty strings
                            if cell_value == '':
                                continue
                        # Handle occupation fields with new dropdown options
                        if field_name == 'forest_based_occupation':
                            # True if not "छैन" (has forest-based occupation)
                            if isinstance(cell_value, str):
                                row_data[field_name] = cell_value.strip() != 'छैन'
                            else:
                                row_data[field_name] = bool(cell_value)
                        elif field_name == 'other_occupation':
                            # True if any value selected (not empty and not "छैन")
                            if isinstance(cell_value, str):
                                val = cell_value.strip()
                                row_data[field_name] = val != '' and val != 'छैन (None / Unable to work)'
                            else:
                                row_data[field_name] = bool(cell_value)
                        elif field_name == 'other_group_membership':
                            # Convert छ/छैन or Yes/No to boolean
                            if isinstance(cell_value, str):
                                row_data[field_name] = cell_value.strip() in ['छ', 'Yes', 'yes', 'true', '1']
                            else:
                                row_data[field_name] = bool(cell_value)
                        else:
                            row_data[field_name] = cell_value

            # Add land unit
            if land_unit:
                row_data['land_unit'] = land_unit

            # Set default values for boolean flags if not present
            if 'firewood_auto_calculated' not in row_data:
                row_data['firewood_auto_calculated'] = True
            if 'grass_auto_calculated' not in row_data:
                row_data['grass_auto_calculated'] = True
            if 'bedding_auto_calculated' not in row_data:
                row_data['bedding_auto_calculated'] = True
            if 'caste_classification_manual' not in row_data:
                row_data['caste_classification_manual'] = False
            if 'prosperity_auto_suggested' not in row_data:
                row_data['prosperity_auto_suggested'] = True

            # Validate required fields
            if not row_data.get('house_no'):
                validation.errors.append("House number is required")
                validation.is_valid = False
            if not row_data.get('surname'):
                validation.errors.append("Surname is required")
                validation.is_valid = False
            if row_data.get('female_count') is None:
                validation.errors.append("Female count is required")
                validation.is_valid = False
            if row_data.get('male_count') is None:
                validation.errors.append("Male count is required")
                validation.is_valid = False

            # Validate numeric fields
            numeric_fields = ['house_no', 'female_count', 'male_count', 'cow_ox_count',
                            'buffalo_count', 'goat_sheep_count', 'pole_demand']
            for field in numeric_fields:
                if field in row_data and row_data[field] is not None:
                    try:
                        row_data[field] = int(row_data[field])
                        if row_data[field] < 0:
                            validation.errors.append(f"{field} cannot be negative")
                            validation.is_valid = False
                    except (ValueError, TypeError):
                        validation.errors.append(f"Invalid value for {field}")
                        validation.is_valid = False

            # Validate decimal fields
            decimal_fields = ['latitude', 'longitude', 'land_area', 'timber_demand_cft',
                            'firewood_demand_bhari', 'grass_demand_bhari', 'bedding_demand_bhari']
            for field in decimal_fields:
                if field in row_data and row_data[field] is not None:
                    try:
                        row_data[field] = Decimal(str(row_data[field]))
                    except:
                        validation.errors.append(f"Invalid decimal value for {field}")
                        validation.is_valid = False

            # Perform automatic calculations if valid so far
            if validation.is_valid:
                try:
                    calculated_data, warnings = HouseholdCalculations.validate_and_calculate(
                        row_data, db
                    )
                    row_data.update(calculated_data)
                    validation.warnings.extend(warnings)
                except Exception as e:
                    validation.errors.append(f"Calculation error: {str(e)}")
                    validation.is_valid = False

            validation.data = row_data
            validations.append(validation)

            if validation.is_valid:
                records_to_import.append(row_data)

        # Import valid records with upsert logic (update if house_no exists, insert if new)
        imported_count = 0
        updated_count = 0

        if records_to_import:
            # Get all existing house numbers for this calculation
            existing_households = db.query(HouseholdInformation).filter(
                HouseholdInformation.calculation_id == calculation_id
            ).all()

            # Create a map of house_no -> household object for quick lookup
            existing_by_house_no = {h.house_no: h for h in existing_households}

            # Process each record
            for idx, record_data in enumerate(records_to_import):
                house_no = record_data.get('house_no')

                # Check if this house_no already exists
                if house_no in existing_by_house_no:
                    # UPDATE existing record
                    existing_household = existing_by_house_no[house_no]

                    # Update all fields
                    for key, value in record_data.items():
                        setattr(existing_household, key, value)

                    updated_count += 1

                    # Add warning to validation for this row
                    for validation in validations:
                        if validation.data and validation.data.get('house_no') == house_no:
                            validation.warnings.append(
                                f"⚠️ House #{house_no} already exists - record will be UPDATED"
                            )
                            break
                else:
                    # INSERT new record
                    household = HouseholdInformation(
                        calculation_id=calculation_id,
                        created_by=current_user.id,
                        **record_data
                    )
                    db.add(household)
                    imported_count += 1

            db.commit()

        # ====================================================================
        # Parse and Import Committee Data (Optional)
        # ====================================================================
        committee_imported_count = 0
        committee_validations = []

        # Only parse committee data from Excel files (not CSV)
        if not is_csv:
            # Header mappings for committee sheets
            main_committee_headers = {
                "सि.नं.": "serial_no",
                "लिङ्ग": "gender",
                "पद": "position",
                "जातिय वर्ग": "caste_category",
                "नाम": "name",
                "ठेगाना": "address",
                "मोवाइल नंवर": "mobile"
            }

            advisory_financial_headers = {
                "सि.नं.": "serial_no",
                "नाम": "name",
                "ठेगाना": "address",
                "मोवाइल नंवर": "mobile"
            }

            # Parse Main Committee (वन उपभोक्ता समिति)
            if "वन उपभोक्ता समिति" in wb.sheetnames:
                try:
                    ws_main = wb["वन उपभोक्ता समिति"]
                    main_records, main_validations = ForestCommitteeValidation.parse_committee_sheet(
                        ws_main, main_committee_headers, 'main'
                    )

                    # Validate composition (50% women rule, position uniqueness)
                    if main_records:
                        _, composition_warnings = ForestCommitteeValidation.validate_committee_composition(main_records)
                        if composition_warnings:
                            # Add composition warnings to the first validation result
                            if main_validations:
                                main_validations[0]['warnings'].extend(composition_warnings)

                    # Import main committee records
                    for record in main_records:
                        member = ForestUserCommittee(
                            calculation_id=calculation_id,
                            created_by=current_user.id,
                            **record
                        )
                        db.add(member)
                        committee_imported_count += 1

                    committee_validations.extend(main_validations)
                except Exception as e:
                    print(f"Error parsing main committee: {str(e)}")
                    import traceback
                    traceback.print_exc()

            # Parse Advisory Committee (सल्लाहाकार समिति)
            if "सल्लाहाकार समिति" in wb.sheetnames:
                try:
                    ws_advisory = wb["सल्लाहाकार समिति"]
                    advisory_records, advisory_validations = ForestCommitteeValidation.parse_committee_sheet(
                        ws_advisory, advisory_financial_headers, 'advisory'
                    )

                    # Import advisory committee records
                    for record in advisory_records:
                        member = AdvisoryCommittee(
                            calculation_id=calculation_id,
                            created_by=current_user.id,
                            **record
                        )
                        db.add(member)
                        committee_imported_count += 1

                    committee_validations.extend(advisory_validations)
                except Exception as e:
                    print(f"Error parsing advisory committee: {str(e)}")
                    import traceback
                    traceback.print_exc()

            # Parse Financial Committee (आर्थिक समिति)
            if "आर्थिक समिति" in wb.sheetnames:
                try:
                    ws_financial = wb["आर्थिक समिति"]
                    financial_records, financial_validations = ForestCommitteeValidation.parse_committee_sheet(
                        ws_financial, advisory_financial_headers, 'financial'
                    )

                    # Import financial committee records
                    for record in financial_records:
                        member = FinancialCommittee(
                            calculation_id=calculation_id,
                            created_by=current_user.id,
                            **record
                        )
                        db.add(member)
                        committee_imported_count += 1

                    committee_validations.extend(financial_validations)
                except Exception as e:
                    print(f"Error parsing financial committee: {str(e)}")
                    import traceback
                    traceback.print_exc()

            # Commit committee data
            if committee_imported_count > 0:
                try:
                    db.commit()
                    # Invalidate OP data cache so {{uc_members}} etc. show fresh data
                    from app.models.op_data_cache import OpDataCache
                    db.query(OpDataCache).filter(
                        OpDataCache.calculation_id == calculation_id
                    ).delete(synchronize_session=False)
                    db.commit()
                except Exception as e:
                    db.rollback()
                    print(f"Error committing committee data: {str(e)}")
                    import traceback
                    traceback.print_exc()

        return HouseholdUploadResponse(
            success=(imported_count + updated_count) > 0,
            total_rows=len(validations),
            valid_rows=len(records_to_import),
            invalid_rows=len(validations) - len(records_to_import),
            records_imported=imported_count,
            records_updated=updated_count,
            validations=validations
        )

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Required library not installed: {str(e)}"
        )
    except UnicodeDecodeError as e:
        raise HTTPException(
            status_code=400,
            detail="CSV file encoding error. Please save your file as UTF-8 encoded CSV."
        )
    except Exception as e:
        db.rollback()
        import traceback
        print("="*80)
        print("FILE UPLOAD ERROR:")
        print("="*80)
        traceback.print_exc()
        print("="*80)
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


# ============================================================================
# CRUD Operations
# ============================================================================

@router.get("/calculations/{calculation_id}/households", response_model=List[HouseholdInfoResponse])
def get_households(
    calculation_id: UUID,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all household records for a calculation"""

    # Verify calculation belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Get households with computed total_population
    households = db.query(HouseholdInformation).filter(
        HouseholdInformation.calculation_id == calculation_id
    ).offset(skip).limit(limit).all()

    # Add computed field
    results = []
    for h in households:
        h_dict = {
            **{c.name: getattr(h, c.name) for c in h.__table__.columns},
            "total_population": h.female_count + h.male_count
        }
        results.append(HouseholdInfoResponse(**h_dict))

    return results


@router.get("/households/{household_id}", response_model=HouseholdInfoResponse)
def get_household(
    household_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a single household by ID"""

    household = db.query(HouseholdInformation).filter(
        HouseholdInformation.id == household_id
    ).first()

    if not household:
        raise HTTPException(status_code=404, detail="Household not found")

    # Verify user owns the calculation
    calculation = db.query(Calculation).filter(
        Calculation.id == household.calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=403, detail="Access denied")

    # Return with computed field
    h_dict = {
        **{c.name: getattr(household, c.name) for c in household.__table__.columns},
        "total_population": household.female_count + household.male_count
    }
    return HouseholdInfoResponse(**h_dict)


@router.post("/calculations/{calculation_id}/households", response_model=HouseholdInfoResponse)
def create_household(
    calculation_id: UUID,
    household_data: HouseholdInfoCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new household record"""

    # Verify calculation belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Perform automatic calculations
    data_dict = household_data.dict()
    calculated_data, warnings = HouseholdCalculations.validate_and_calculate(data_dict, db)
    data_dict.update(calculated_data)

    # Create household
    household = HouseholdInformation(
        calculation_id=calculation_id,
        created_by=current_user.id,
        **data_dict
    )
    db.add(household)
    db.commit()
    db.refresh(household)

    # Return with computed field
    h_dict = {
        **{c.name: getattr(household, c.name) for c in household.__table__.columns},
        "total_population": household.female_count + household.male_count
    }
    return HouseholdInfoResponse(**h_dict)


@router.put("/households/{household_id}", response_model=HouseholdInfoResponse)
def update_household(
    household_id: UUID,
    household_update: HouseholdInfoUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a household record"""

    household = db.query(HouseholdInformation).filter(
        HouseholdInformation.id == household_id
    ).first()

    if not household:
        raise HTTPException(status_code=404, detail="Household not found")

    # Verify user owns the calculation
    calculation = db.query(Calculation).filter(
        Calculation.id == household.calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=403, detail="Access denied")

    # Update fields
    update_data = household_update.dict(exclude_unset=True)
    
    # Debug log
    print(f"DEBUG: Received update data: {update_data}")

    # Handle manual edits - mark fields as manually entered
    if 'prosperity_level' in update_data:
        update_data['prosperity_auto_suggested'] = False
    if 'caste_classification_ne' in update_data:
        update_data['caste_classification_manual'] = True

    # Perform automatic calculations if relevant fields changed (but not if only manual override fields changed)
    needs_recalc = any(key in update_data for key in [
        'female_count', 'male_count', 'cow_ox_count', 'buffalo_count', 'goat_sheep_count',
        'surname', 'land_area', 'land_unit'
    ])

    # Don't recalculate if user manually edited auto-calculated fields
    auto_edited_fields = ['firewood_demand_bhari', 'grass_demand_bhari', 'bedding_demand_bhari',
                          'prosperity_level', 'caste_classification_ne', 'forest_based_occupation']
    if any(key in update_data for key in auto_edited_fields):
        needs_recalc = False
        # Ensure auto-calculated flags are set to False for manually edited fields
        if 'firewood_demand_bhari' in update_data:
            update_data['firewood_auto_calculated'] = False
        if 'grass_demand_bhari' in update_data:
            update_data['grass_auto_calculated'] = False
        if 'bedding_demand_bhari' in update_data:
            update_data['bedding_auto_calculated'] = False

    if needs_recalc:
        # Get current data and merge with updates
        current_data = {c.name: getattr(household, c.name) for c in household.__table__.columns}
        current_data.update(update_data)

        calculated_data, warnings = HouseholdCalculations.validate_and_calculate(current_data, db)
        update_data.update(calculated_data)

    # Apply updates
    for key, value in update_data.items():
        setattr(household, key, value)

    household.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(household)

    # Return with computed field
    h_dict = {
        **{c.name: getattr(household, c.name) for c in household.__table__.columns},
        "total_population": household.female_count + household.male_count
    }
    return HouseholdInfoResponse(**h_dict)


@router.delete("/households/{household_id}")
def delete_household(
    household_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a household record"""

    household = db.query(HouseholdInformation).filter(
        HouseholdInformation.id == household_id
    ).first()

    if not household:
        raise HTTPException(status_code=404, detail="Household not found")

    # Verify user owns the calculation
    calculation = db.query(Calculation).filter(
        Calculation.id == household.calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=403, detail="Access denied")

    db.delete(household)
    db.commit()

    return {"message": "Household deleted successfully"}


@router.delete("/calculations/{calculation_id}/households")
def delete_all_households(
    calculation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete all household records for a calculation (before re-upload)"""

    # Verify calculation belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Delete all households
    deleted_count = db.query(HouseholdInformation).filter(
        HouseholdInformation.calculation_id == calculation_id
    ).delete()

    db.commit()

    return {"message": f"Deleted {deleted_count} household records"}


# ============================================================================
# Summary Statistics
# ============================================================================

@router.get("/calculations/{calculation_id}/summary", response_model=HouseholdSummary)
def get_household_summary(
    calculation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get aggregate statistics for all households in a calculation"""

    # Verify calculation belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Get all households
    households = db.query(HouseholdInformation).filter(
        HouseholdInformation.calculation_id == calculation_id
    ).all()

    if not households:
        raise HTTPException(status_code=404, detail="No household data found")

    # Calculate totals
    total_households = len(households)
    total_male = sum(h.male_count for h in households)
    total_female = sum(h.female_count for h in households)
    total_population = total_male + total_female
    total_cow_ox = sum(h.cow_ox_count for h in households)
    total_buffalo = sum(h.buffalo_count for h in households)
    total_goat_sheep = sum(h.goat_sheep_count for h in households)
    total_timber = sum(h.timber_demand_cft for h in households)
    total_poles = sum(h.pole_demand for h in households)
    total_firewood = sum(h.firewood_demand_bhari or 0 for h in households)
    total_grass = sum(h.grass_demand_bhari or 0 for h in households)
    total_bedding = sum(h.bedding_demand_bhari or 0 for h in households)

    # Average land area
    land_areas = [h.land_area for h in households if h.land_area is not None]
    avg_land = Decimal(sum(land_areas) / len(land_areas)) if land_areas else None

    # Caste distribution
    caste_dist = {}
    for h in households:
        if h.caste_classification_ne:
            caste_dist[h.caste_classification_ne] = caste_dist.get(h.caste_classification_ne, 0) + 1

    # Prosperity distribution
    prosperity_dist = {}
    for h in households:
        prosperity_dist[h.prosperity_level] = prosperity_dist.get(h.prosperity_level, 0) + 1

    # Forest dependent households
    forest_dependent = sum(1 for h in households if h.forest_based_occupation)

    return HouseholdSummary(
        total_households=total_households,
        total_population=total_population,
        total_male=total_male,
        total_female=total_female,
        total_cow_ox=total_cow_ox,
        total_buffalo=total_buffalo,
        total_goat_sheep=total_goat_sheep,
        total_timber_demand_cft=total_timber,
        total_pole_demand=total_poles,
        total_firewood_demand_bhari=total_firewood,
        total_grass_demand_bhari=total_grass,
        total_bedding_demand_bhari=total_bedding,
        avg_land_area=avg_land,
        caste_distribution=caste_dist,
        prosperity_distribution=prosperity_dist,
        forest_dependent_households=forest_dependent
    )


# ============================================================================
# Export Analysis
# ============================================================================

@router.get("/calculations/{calculation_id}/export")
async def export_household_analysis(
    calculation_id: UUID,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export household data and analysis to Excel"""

    # Verify calculation belongs to user
    calculation = db.query(Calculation).filter(
        Calculation.id == calculation_id,
        Calculation.user_id == current_user.id
    ).first()

    if not calculation:
        raise HTTPException(status_code=404, detail="Calculation not found")

    # Get households and summary
    households = db.query(HouseholdInformation).filter(
        HouseholdInformation.calculation_id == calculation_id
    ).all()

    if not households:
        raise HTTPException(status_code=404, detail="No household data to export")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import PieChart, BarChart, Reference

        wb = Workbook()

        # Sheet 1: Household Data
        ws_data = wb.active
        ws_data.title = "Household Data"

        # Headers
        headers = [
            "House No", "Surname", "Male Head", "Female Head", "Address",
            "Latitude", "Longitude", "Female Count", "Male Count", "Total Population",
            "Land Area", "Land Unit", "Forest Occupation", "Other Occupation",
            "Cow/Ox", "Buffalo", "Goat/Sheep", "Timber (cft)", "Poles",
            "Firewood (Bhari)", "Grass (Bhari)", "Bedding (Bhari)",
            "Caste Classification", "Other Group", "Prosperity", "Remarks"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data rows
        for row_idx, h in enumerate(households, start=2):
            ws_data.cell(row=row_idx, column=1, value=h.house_no)
            ws_data.cell(row=row_idx, column=2, value=h.surname)
            ws_data.cell(row=row_idx, column=3, value=h.household_head_male)
            ws_data.cell(row=row_idx, column=4, value=h.household_head_female)
            ws_data.cell(row=row_idx, column=5, value=h.address_tole)
            ws_data.cell(row=row_idx, column=6, value=float(h.latitude) if h.latitude else None)
            ws_data.cell(row=row_idx, column=7, value=float(h.longitude) if h.longitude else None)
            ws_data.cell(row=row_idx, column=8, value=h.female_count)
            ws_data.cell(row=row_idx, column=9, value=h.male_count)
            ws_data.cell(row=row_idx, column=10, value=h.female_count + h.male_count)
            ws_data.cell(row=row_idx, column=11, value=float(h.land_area) if h.land_area else None)
            ws_data.cell(row=row_idx, column=12, value=h.land_unit)
            ws_data.cell(row=row_idx, column=13, value="छ" if h.forest_based_occupation else "छैन")
            ws_data.cell(row=row_idx, column=14, value="छ" if h.other_occupation else "छैन")
            ws_data.cell(row=row_idx, column=15, value=h.cow_ox_count)
            ws_data.cell(row=row_idx, column=16, value=h.buffalo_count)
            ws_data.cell(row=row_idx, column=17, value=h.goat_sheep_count)
            ws_data.cell(row=row_idx, column=18, value=float(h.timber_demand_cft))
            ws_data.cell(row=row_idx, column=19, value=h.pole_demand)
            ws_data.cell(row=row_idx, column=20, value=float(h.firewood_demand_bhari) if h.firewood_demand_bhari else None)
            ws_data.cell(row=row_idx, column=21, value=float(h.grass_demand_bhari) if h.grass_demand_bhari else None)
            ws_data.cell(row=row_idx, column=22, value=float(h.bedding_demand_bhari) if h.bedding_demand_bhari else None)
            ws_data.cell(row=row_idx, column=23, value=h.caste_classification_ne)
            ws_data.cell(row=row_idx, column=24, value="छ" if h.other_group_membership else "छैन")
            ws_data.cell(row=row_idx, column=25, value=h.prosperity_level)
            ws_data.cell(row=row_idx, column=26, value=h.remarks)

        # Sheet 2: Summary Statistics
        ws_summary = wb.create_sheet("Summary")

        summary_data = [
            ["Household Summary Statistics", ""],
            ["", ""],
            ["Total Households", len(households)],
            ["Total Population", sum(h.female_count + h.male_count for h in households)],
            ["  - Male", sum(h.male_count for h in households)],
            ["  - Female", sum(h.female_count for h in households)],
            ["", ""],
            ["Livestock", ""],
            ["Total Cows/Oxen", sum(h.cow_ox_count for h in households)],
            ["Total Buffaloes", sum(h.buffalo_count for h in households)],
            ["Total Goats/Sheep", sum(h.goat_sheep_count for h in households)],
            ["", ""],
            ["Forest Product Demands (Yearly)", ""],
            ["Timber (cft)", float(sum(h.timber_demand_cft for h in households))],
            ["Poles", sum(h.pole_demand for h in households)],
            ["Firewood (Bhari)", float(sum(h.firewood_demand_bhari or 0 for h in households))],
            ["Grass (Bhari)", float(sum(h.grass_demand_bhari or 0 for h in households))],
            ["Bedding Material (Bhari)", float(sum(h.bedding_demand_bhari or 0 for h in households))],
            ["", ""],
            ["Forest Dependent Households", sum(1 for h in households if h.forest_based_occupation)],
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws_summary.cell(row=row_idx, column=1).font = Font(bold=True, size=14)

        # Auto-adjust column widths
        for ws in [ws_data, ws_summary]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to temporary file
        with tempfile.NamedTemporaryFile(mode='wb', suffix='.xlsx', delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name

        filename = f"household_analysis_{calculation.forest_name or 'data'}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        return FileResponse(
            tmp_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating export: {str(e)}")


# ============================================================================
# Utility Endpoints
# ============================================================================

@router.get("/surnames", response_model=List[SurnameSuggestion])
def get_surname_suggestions(
    q: str = Query(..., min_length=1),
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db)
):
    """Get surname suggestions for autocomplete"""

    results = db.query(CasteClassification).filter(
        CasteClassification.surname_ne.ilike(f"%{q}%")
    ).limit(limit).all()

    return [
        SurnameSuggestion(
            surname_ne=r.surname_ne,
            surname_en=r.surname_en,
            classification_ne=r.classification_ne,
            caste_ne=r.caste_ne
        )
        for r in results
    ]


@router.get("/caste-lookup/{surname}", response_model=List[CasteClassificationResponse])
def lookup_caste_by_surname(
    surname: str,
    db: Session = Depends(get_db)
):
    """Lookup caste classification by surname (may return multiple results)"""

    results = db.query(CasteClassification).filter(
        CasteClassification.surname_ne == surname
    ).all()

    if not results:
        raise HTTPException(status_code=404, detail=f"Surname '{surname}' not found in database")

    return results
